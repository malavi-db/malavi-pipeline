"""Render one submission as a self-contained HTML review document for a curator.

The curator this is written for is a biologist with a browser and an email client. They
are not going to run a command, read JSON, or open a terminal to find out why a check
fired. Everything they need is on this one page, in the order they work in: what needs
judgment first, then the evidence, then everything that was submitted.

Three rules shape the implementation.

**Escape everything.** The values on this page came from a public Google Form and are
rendered into a document a curator opens from ``file://``, where a script would run with
local-file privileges. A spreadsheet cell containing markup must never become markup. That
is not a theoretical concern for a database that invites strangers to upload workbooks.

**Self-contained.** Inline CSS, no fonts, images, scripts or fetches from anywhere. The
page must render identically from a local file, from a mail attachment, and from behind a
login, because which of those a curator uses is a deployment decision that should not
change what they see.

**Show the submitter's words next to ours.** Wherever the system reinterpreted a value --
normalized a name, read a column by position -- both forms appear. A curator correcting a
submission has to be able to tell what the submitter typed from what we made of it.
"""
from __future__ import annotations

import html
import os
import re
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

from ._league_gothic_b64 import WOFF_BASE64
from .checks import CHECKS, CheckRun, Outcome, Severity
from .config import load_config
from .verdicts import prefill_url

# The site's own design tokens, copied from docs/assets/css/malavi.css rather than
# approximated: violet-biased neutrals, the Okabe-Ito parasite-genus colors, and the
# ok/warn/stop trio. A curator moving between the public site and this document should
# not feel they have changed software.
#
# The tokens are duplicated here rather than imported because the report must be one
# self-contained file that renders from file://, from a mail attachment and from behind a
# login. If they drift, test_report_html catches it.
_STYLE = """
@font-face {
  font-family:"League Gothic";
  src:url(data:font/woff;base64,__LEAGUE_GOTHIC__) format("woff");
  font-weight:400; font-style:normal; font-display:block;
}
:root {
  --paper:#F4F3F8; --surface:#FFFFFF; --surface-2:#ECEAF3; --surface-3:#E2DFEC;
  --ink:#17141F; --ink-2:#443E55; --ink-3:#6E677F;
  --rule:#D8D4E4; --rule-soft:#E6E3EE;
  --accent:#5B4BA6; --accent-ink:#FFFFFF; --accent-wash:#EDEAF8;
  --ok:#1F7A4D; --warn:#9A6410; --stop:#A63A3A;
  --ok-wash:#E4F2EA; --warn-wash:#F7EEDC; --stop-wash:#F7E7E7;
  --info:#0072B2; --info-wash:#D9E9F5;
  --shadow:0 1px 2px rgba(23,20,31,.05), 0 8px 24px -12px rgba(23,20,31,.18);
  --radius:10px; --measure:68ch;
}
@media (prefers-color-scheme: dark) {
  :root {
    --paper:#17141F; --surface:#201C2B; --surface-2:#2A2537; --surface-3:#332C42;
    --ink:#ECEAF4; --ink-2:#B4AECA; --ink-3:#8B8499;
    --rule:#362F47; --rule-soft:#2B2539;
    --accent:#A99CF0; --accent-ink:#17141F; --accent-wash:#292244;
    --ok:#5FC08C; --warn:#DCA84A; --stop:#E38080;
    --ok-wash:#1C3129; --warn-wash:#332A18; --stop-wash:#35211F;
    --info:#3B82D0; --info-wash:#16263B;
    --shadow:0 1px 2px rgba(0,0,0,.4), 0 8px 24px -12px rgba(0,0,0,.7);
  }
}
* { box-sizing:border-box; }
body {
  margin:0; background:var(--paper); color:var(--ink);
  font-family: ui-sans-serif, system-ui, -apple-system, "Segoe UI", Roboto,
               "Helvetica Neue", Arial, sans-serif;
  font-size:16px; line-height:1.6; -webkit-font-smoothing:antialiased;
}
.wrap { max-width:1140px; margin:0 auto; padding:38px 20px 96px; }
.display { font-family:"League Gothic", ui-sans-serif, sans-serif; font-weight:400;
  letter-spacing:.01em; text-transform:uppercase; line-height:.92; text-wrap:balance; }
.mono { font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
  font-variant-numeric: tabular-nums; }
.sci { font-style:italic; }
h1 { font-size:clamp(34px,5vw,50px); margin:0 0 8px; }
h2 { font-size:30px; margin:44px 0 6px; }
h3 { font-size:17px; margin:26px 0 8px; letter-spacing:-.01em; }
p { margin:0 0 12px; max-width:var(--measure); color:var(--ink-2); }
.eyebrow { font-size:11px; letter-spacing:.15em; text-transform:uppercase;
  color:var(--ink-3); margin:0 0 10px; }
.lead { font-size:17.5px; }
.sub { color:var(--ink-3); font-size:13.5px; margin:0 0 20px; }
/* Definition lists that state a vocabulary the page then uses: the triage tiers and
   the sequence QC calls. Kept narrow so a definition reads as a caption, not a table. */
/* A cell stating something the curator has to resolve before the row can be accepted --
   currently an alternative name pointing at a lineage MalAvi does not hold. */
.bad { background:var(--stop-wash); color:var(--stop); font-weight:600; }
/* The per-window parentage behind a chimera call, printed under the finding it
   explains rather than in a section of its own. */
.rowflags { font-size:10px; color:var(--ink-3); line-height:1.25; }
.chimera { margin:8px 0 2px; font-size:12.5px; color:var(--ink-2); }
.chimera ul { margin:4px 0 4px 18px; padding:0; }
.chimera li { margin:1px 0; }
.chimera .caveat { color:var(--ink-3); font-style:italic; margin:4px 0 0; }
.alnkey { font-size:12px; color:var(--ink-3); margin:0 0 18px; }
.alnkey .sw { display:inline-block; min-width:16px; text-align:center; font-weight:700;
  border-radius:3px; padding:1px 4px; margin:0 5px 0 14px; font-family:var(--mono, monospace); }
.alnkey .sw:first-child { margin-left:0; }
.alnkey .sw.k-nonsyn { background:var(--stop-wash); color:var(--stop); }
.alnkey .sw.k-transversion { background:var(--warn-wash); color:var(--warn); }
.alnkey .sw.k-transition { background:var(--ok-wash); color:var(--ok); }
.alnkey .sw.rare { text-decoration:underline; text-decoration-thickness:2px; }
.tierkey { margin:0 0 18px; font-size:13px; color:var(--ink-2); }
.tierkey dt { font-weight:600; margin:8px 0 0; }
.tierkey dd { margin:1px 0 0 0; color:var(--ink-3); }

.card { background:var(--surface); border:1px solid var(--rule);
  border-radius:var(--radius); padding:22px; box-shadow:var(--shadow); }
.tiles { display:flex; gap:10px; flex-wrap:wrap; margin:14px 0 18px; }
.tile { border:1px solid var(--rule); border-radius:10px; padding:12px 18px;
  min-width:120px; background:var(--surface); }
.tile .n { display:block; font-family:"League Gothic",Impact,sans-serif; font-size:34px;
  line-height:1; }
.tile .k { display:block; font-size:11px; letter-spacing:.08em; text-transform:uppercase;
  color:var(--ink-3); margin-top:6px; }
.appendix-head { margin-top:26px; }
.taken-name { font-family:var(--mono, monospace); font-weight:700; color:var(--stop); }
.taken-note { color:var(--stop); font-size:13px; }
/* Fixed layout, or the browser hands the host-name column as much width as its longest
   binomial wants and the lineage columns run off the right edge of the paper -- which is
   invisible in HTML, because it scrolls, and lost in the PDF, which does not. The row
   header is capped and allowed to wrap for the same reason. */
table.matrix { border-collapse:collapse; font-size:10px; table-layout:fixed;
  width:100%; }
table.matrix th, table.matrix td { border:1px solid var(--rule); padding:4px 2px;
  text-align:center; overflow-wrap:anywhere; }
/* break-word, not anywhere: a binomial should wrap at the space between genus and
   species, and only split a word when it genuinely cannot fit. "Aix galericula / ta" is
   what anywhere gives you. */
table.matrix th.rowhead { text-align:left; white-space:normal; font-weight:400;
  background:var(--surface-2); text-transform:none; letter-spacing:0;
  font-size:10.5px; width:30%; overflow-wrap:break-word; }
table.matrix th.mono { font-size:9px; letter-spacing:0; }
table.matrix td.hit { font-weight:700; background:var(--ok-wash); }
table.matrix td.miss { color:var(--ink-3); }
h2.display { margin-top:34px; }
h3.place { margin:20px 0 6px; font-size:14px; font-weight:700; }
ul.files { margin:6px 0 14px 18px; font-size:13px; }
ul.files a { color:var(--accent); }

/* A group that came through clean, and the tally of quiet checks. Green because the
   report should be able to say good news as clearly as it says bad. */
/* The disposition line. Status is carried by the WORD, not by the colour -- a reader who
   cannot distinguish the tints still gets the whole message. */
.disposition { border:2px solid var(--rule); border-radius:10px; padding:14px 18px;
  margin:12px 0 22px; }
.disposition .d-label { display:block; font-weight:700; font-size:16px; }
.disposition .d-note { display:block; margin-top:6px; font-size:13px; color:var(--ink-2); }
.disposition.blocked { border-color:var(--stop); background:var(--stop-wash); }
.disposition.incomplete { border-color:var(--warn); background:var(--warn-wash); }
.disposition.review { border-color:var(--accent); background:var(--accent-wash); }
.disposition.clear { border-color:var(--ok); background:var(--ok-wash); }

.check.clear { border-left-color:var(--ok); background:var(--ok-wash); }
.pill.clear { background:var(--ok); color:#fff; }

h3.group { margin:26px 0 10px; font-family:"League Gothic",Impact,sans-serif;
  text-transform:uppercase; letter-spacing:.02em; font-size:20px;
  border-bottom:2px solid var(--rule); padding-bottom:4px; }
a.jump { display:inline-block; margin:8px 12px 0 0; padding:4px 11px;
  border:1px solid var(--accent); border-radius:6px; color:var(--accent);
  font-size:11px; font-weight:700; text-decoration:none; white-space:nowrap; }

/* FASTA and any long unbroken string. Without the break rules a 479-character sequence
   is one word, and one word cannot wrap: it simply runs off the page and out of the PDF. */
pre.fasta, .fasta { font-family:var(--mono, "DejaVu Sans Mono", monospace); font-size:10px;
  line-height:1.45; background:var(--surface-2); border:1px solid var(--rule);
  border-radius:8px; padding:12px 14px; white-space:pre-wrap; word-break:break-all;
  overflow-wrap:anywhere; }

/* The reproduced submission. Given a tinted ground and a rule down the side so it is
   obvious at a glance that the report has stopped talking and started quoting. */
.appendix { background:var(--surface-2); border-left:4px solid var(--ink-3);
  padding:18px 20px 4px; margin-top:34px; border-radius:0 10px 10px 0; }
.appendix table { font-size:10px; }
.appendix td, .appendix th { word-break:break-all; overflow-wrap:anywhere;
  max-width:220px; }

/* PDF bookmarks, so the reader can navigate a nine-page document from the sidebar
   instead of scrolling it. */
h1 { bookmark-level:1; bookmark-label:content(); }
h2 { bookmark-level:2; bookmark-label:content(); }
h3.group, h3.appendix-head { bookmark-level:3; bookmark-label:content(); }
.check .title { font-weight:700; font-size:15px; }

/* The block a curator acts from. Given its own weight because it is the only part of
   this document that asks for something, and it has to survive being printed. */
.verdict { border:2px solid var(--accent); border-radius:10px; padding:18px 20px;
  margin:12px 0 26px; background:var(--accent-wash); break-inside:avoid; }
.verdict p { margin:0 0 10px; }
.verdict p:last-child { margin-bottom:0; }
.verdict-link { display:inline-block; background:var(--accent); color:#fff;
  text-decoration:none; font-weight:700; padding:10px 18px; border-radius:8px; }
.verdict-note { font-size:13px; color:var(--ink-2); }

.banner { background:var(--stop-wash); border:1px solid var(--stop);
  border-radius:var(--radius); padding:16px 20px; margin:20px 0; color:var(--ink-2); }
.banner b { color:var(--stop); }
.note { background:var(--surface-2); border-left:3px solid var(--accent);
  border-radius:0 var(--radius) var(--radius) 0; padding:12px 16px; margin:14px 0;
  font-size:13.5px; color:var(--ink-2); max-width:var(--measure); }

.meta { display:grid; grid-template-columns:max-content 1fr; gap:6px 22px;
  background:var(--surface); border:1px solid var(--rule); border-radius:var(--radius);
  padding:18px 22px; font-size:13.5px; box-shadow:var(--shadow); }
.meta dt { font-size:10.5px; letter-spacing:.08em; text-transform:uppercase;
  color:var(--ink-3); align-self:center; }
.meta dd { margin:0; color:var(--ink-2); }

.check { background:var(--surface); border:1px solid var(--rule);
  border-radius:var(--radius); padding:14px 18px; margin:10px 0;
  box-shadow:var(--shadow); border-left:3px solid var(--rule); }
.check.finding { border-left-color:var(--warn); }
.check.finding.blocking { border-left-color:var(--stop); }
.check.finding.info { border-left-color:var(--info); }
.check.error { border-left-color:var(--stop); background:var(--stop-wash); }
.check.skip { border-left-color:var(--ink-3); background:var(--surface-2);
  box-shadow:none; }
.check .hd { display:flex; gap:10px; align-items:center; flex-wrap:wrap; }
.check .id { font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
  font-size:13.5px; font-weight:600; color:var(--ink); }
.check .count { color:var(--ink-3); font-size:12px; margin-left:auto;
  font-variant-numeric:tabular-nums; }
.asserts { font-size:13.5px; color:var(--ink-2); margin:6px 0 0;
  max-width:var(--measure); }
.pill { font-size:10.5px; letter-spacing:.08em; text-transform:uppercase;
  padding:3px 10px; border-radius:999px; font-weight:600; }
.pill.blocking { background:var(--stop-wash); color:var(--stop); }
.pill.warning { background:var(--warn-wash); color:var(--warn); }
.pill.info { background:var(--info-wash); color:var(--info); }

.finds { margin:12px 0 0; padding:0; list-style:none; }
.finds li { padding:9px 0; border-top:1px solid var(--rule-soft); font-size:13.5px;
  color:var(--ink-2); }
.finds .subj { font-weight:600; color:var(--ink);
  font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
  font-size:13px; }
.finds .where { color:var(--ink-3); font-size:11.5px; }

.passes { background:var(--surface); border:1px solid var(--rule);
  border-radius:var(--radius); overflow:hidden; }
.passes .row { display:flex; gap:14px; align-items:baseline; padding:9px 18px;
  border-bottom:1px solid var(--rule-soft); font-size:13px; color:var(--ink-3); }
.passes .row:last-child { border-bottom:0; }
.passes .row b { font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
  font-weight:600; color:var(--ink-2); min-width:250px; font-size:12.5px; }
.passes .row .count { margin-left:auto; font-variant-numeric:tabular-nums;
  white-space:nowrap; }

.tablewrap { overflow-x:auto; border:1px solid var(--rule);
  border-radius:var(--radius); background:var(--surface); margin:12px 0; }
table { border-collapse:collapse; width:100%; font-size:13.5px; }
th { background:var(--surface-2); text-align:left; font-size:10.5px; letter-spacing:.08em;
  text-transform:uppercase; color:var(--ink-3); font-weight:600; padding:9px 12px;
  white-space:nowrap; border-bottom:1px solid var(--rule); position:sticky; top:0; }
td { padding:8px 12px; border-bottom:1px solid var(--rule-soft); color:var(--ink-2);
  white-space:nowrap; vertical-align:top; }
tr:last-child td { border-bottom:0; }
tbody tr:hover td { background:var(--surface-2); }
td.wrapcell { white-space:normal; min-width:280px; }
td.rowno { color:var(--ink-3); font-variant-numeric:tabular-nums; text-align:right; }
/* The Ref a curator quotes on the verdict form. Deliberately the most legible cell in
   the row: it is the one value they have to copy by eye without mistyping. */
td.handle { font-family: ui-monospace, Menlo, Consolas, monospace; font-weight:700;
  color:var(--accent); background:var(--accent-wash); text-align:center;
  letter-spacing:.02em; white-space:nowrap; }
td.mono, .mono td { font-family: ui-monospace, Menlo, Consolas, monospace;
  color:var(--ink); font-weight:600; }
.was { color:var(--ink-3); text-decoration:line-through; }

/* ---- Alignment grid ----
   A monospace matrix of single characters. The differing base carries the weight; the
   dots are deliberately faint so the eye lands on what changed rather than on the
   scaffolding around it. */
table.aln { border-collapse:collapse; font-family: ui-monospace, SFMono-Regular, Menlo,
  Consolas, monospace; font-size:12px; width:auto; }
table.aln th.seqname { background:var(--surface); text-align:left; padding:3px 14px 3px 10px;
  font-size:11.5px; letter-spacing:0; text-transform:none; color:var(--ink);
  white-space:nowrap; position:static; border-bottom:1px solid var(--rule-soft); }
table.aln th.seqname.query { font-weight:700; }
table.aln th.seqname .dist { color:var(--ink-3); font-weight:400; margin-left:10px;
  font-size:10.5px; }
table.aln th.pos { background:var(--surface-2); padding:5px 3px; text-align:center;
  border-bottom:1px solid var(--rule); font-size:9.5px; letter-spacing:0;
  color:var(--ink-3); font-weight:500; white-space:nowrap; }
table.aln td.b { text-align:center; padding:3px 3px; min-width:26px;
  border-bottom:1px solid var(--rule-soft); color:var(--ink-2); }
table.aln td.b.query { background:var(--accent-wash); color:var(--ink); font-weight:700; }
table.aln td.b.diff { background:var(--warn-wash); color:var(--warn); font-weight:700; }
table.aln td.b.same { color:var(--ink-3); }
table.aln td.b.nodata { color:var(--rule); }
/* What each difference does, shaded on the submitted row. Three mutually exclusive kinds
   plus a rare-base marker, which is an underline rather than a fourth colour so it can
   sit on top of any of them. These must come after td.b.query: they carry the same
   specificity as it does, so order is what decides, and defining them further up the
   sheet is how they silently did nothing the first time. */
table.aln td.b.k-nonsyn { background:var(--stop-wash); color:var(--stop); font-weight:700; }
table.aln td.b.k-transversion { background:var(--warn-wash); color:var(--warn); font-weight:700; }
table.aln td.b.k-transition { background:var(--ok-wash); color:var(--ok); font-weight:700; }
table.aln td.b.rare { text-decoration:underline; text-decoration-thickness:2px; }

footer { margin-top:56px; padding-top:18px; border-top:1px solid var(--rule);
  color:var(--ink-3); font-size:12.5px; max-width:var(--measure); }

/* ---- Paged media -----------------------------------------------------------
   The same document is delivered as a PDF, because it reaches curators through
   Google Drive and Drive has not rendered HTML since 2016 -- without a PDF every
   curator downloads a file instead of clicking a link.

   Paged media breaks things a scrolling page never notices: a finding split
   across a page boundary, a table whose header appears only on the first of six
   pages, and no way to tell which page you are on. Screen colors also print
   badly, so the print rules pin the light palette rather than trusting whatever
   the renderer decided about a dark-mode preference. */
@media print {
  :root {
    --paper:#FFFFFF; --surface:#FFFFFF; --surface-2:#F4F3F8; --surface-3:#ECEAF3;
    --ink:#17141F; --ink-2:#3A3548; --ink-3:#635C73;
    --rule:#C9C4D8; --rule-soft:#DEDAE8;
    --accent:#4A3C90; --shadow:none;
    --ok:#1A6B43; --warn:#845309; --stop:#8F2F2F;
  }
  body { font-size:10.5pt; line-height:1.45; background:#fff; }
  .wrap { max-width:none; padding:0; }

  @page {
    size:A4; margin:16mm 14mm 18mm;
    @bottom-center {
      content:"__PAGE_LABEL__ — page " counter(page) " of " counter(pages);
      font-family:ui-sans-serif, sans-serif; font-size:8pt; color:#635C73;
    }
  }

  h1 { font-size:26pt; }
  h2 { font-size:17pt; break-after:avoid; }
  h3 { font-size:11pt; break-after:avoid; }

  /* A finding split across a page boundary is a finding a curator misreads. */
  .check, .stat, .meta, .banner, .note { break-inside:avoid; }
  section { break-inside:auto; }

  /* Six pages of table with the header on page one is unreadable. */
  thead { display:table-header-group; }
  tr { break-inside:avoid; }
  th { position:static; }

  /* Nothing scrolls on paper: wide content has to wrap or shrink instead. */
  .tablewrap { overflow:visible; }
  table { font-size:8.5pt; }
  /* The alignment is read character by character; shrinking it with the data tables
     would defeat the point of drawing it. */
  table.aln { font-size:9pt; }
  table.aln td.b { min-width:22px; padding:2px; }
  table.aln th.pos { font-size:7.5pt; padding:3px 1px; }
  .alignment { break-inside:avoid; }
  th, td { padding:4px 6px; white-space:normal; word-break:break-word; }
  td.wrapcell { min-width:0; }

  a { color:var(--ink); text-decoration:none; }
  footer { break-before:avoid; }
}
"""

_OUTCOME_WORD = {
    Outcome.ERROR: "could not run",
    Outcome.FINDING: "found something",
    Outcome.SKIP: "skipped",
    Outcome.PASS: "passed",
}


def _stylesheet(page_label: str = "MalAvi curator report") -> str:
    """The stylesheet with the font blob and the printed page label substituted in.

    The font blob is kept out of the CSS literal so the tokens above stay readable; 17 KB
    of base64 in the middle of a stylesheet makes it impossible to review.

    ``page_label`` is the name printed in the footer of every page, and it exists because
    this stylesheet is shared with the release edition report -- a document that is not a
    curator report and must not say it is on all fourteen of its pages.

    Both substitutions are **asserted**, not attempted. An unasserted ``str.replace``
    that matches nothing is the exact failure that once shipped an unstyled verdict block:
    it degrades to a silently wrong document rather than to an error, and the only reader
    who would ever notice is the one holding the printout.
    """
    stylesheet = _STYLE
    for placeholder, value in (("__LEAGUE_GOTHIC__", WOFF_BASE64),
                               ("__PAGE_LABEL__", page_label)):
        if placeholder not in stylesheet:
            raise AssertionError(
                f"the stylesheet has no {placeholder} placeholder, so {placeholder!r} "
                f"would be dropped silently")
        stylesheet = stylesheet.replace(placeholder, value)
    return stylesheet


def esc(value: Any) -> str:
    """HTML-escape any value. The only way text reaches this page."""
    if value is None:
        return ""
    return html.escape(str(value), quote=True)


def _where(source: Optional[Dict[str, Any]],
           known_file: Optional[str] = None) -> str:
    """Render a source reference as the place a curator should look.

    A source that names only the file is suppressed when that file is the workbook the
    report is already about: repeating the same long filename beside every finding is
    noise that crowds out the references that do locate something. Sheet and row always
    survive, because those are the ones worth following.
    """
    if not source:
        return ""
    if source.get("sheet") and source.get("row"):
        return f"{source['sheet']}, row {source['row']}"
    for key in ("sheet", "table"):
        if source.get(key):
            return str(source[key])
    file_name = source.get("file")
    if file_name and file_name != known_file:
        return str(file_name)
    return ""


def _severity_class(check_id: str) -> str:
    check = CHECKS.get(check_id)
    return check.severity.value if check else "warning"


# ---------------------------------------------------------------------------------------
# Sections
# ---------------------------------------------------------------------------------------

def _header(submission: Dict[str, Any], run: CheckRun,
            metadata: Optional[Dict[str, Any]]) -> str:
    reference = submission.get("reference") or {}
    submitter = submission.get("submitter") or {}
    provenance = submission.get("provenance") or {}
    run_provenance = run.provenance or {}

    # No headline. A curator opening this needs to know what to do, not to be presented
    # with a paper's title set in 40pt -- the title is a fact about the submission, so it
    # belongs in the box with the other facts.
    bits: List[str] = [
        '<h1 class="display">MalAvi submission review</h1>']

    if run.incomplete:
        # This must be the first thing a curator sees. A partial run presented without
        # this banner reads as a clean bill of health, which is the single most
        # dangerous thing this document could do.
        bits.append(
            '<div class="banner"><b>Validation is incomplete.</b> At least one check '
            'could not be executed, so the absence of a finding from it means nothing. '
            'The checks concerned are listed below with what went wrong. Do not treat '
            'this submission as screened until they have run.</div>')

    # What a curator needs to know about the submission, and nothing about the software.
    # Schema version, code revision and tool version are real provenance and are kept --
    # at the very end, where somebody debugging a report can find them and nobody reading
    # one has to step over them.
    rows: List[str] = [
        ("Paper", esc(reference.get("title"))),
        ("DOI", esc(reference.get("doi"))),
        ("Submitted by", esc(submitter.get("name"))),
        ("Institution", esc(submitter.get("institution"))),
        ("Email", esc(submitter.get("email"))),
        ("Received", esc((metadata or {}).get("_fetched_at"))),
        ("Workbook", esc(provenance.get("workbook"))),
        ("Checked against release", esc(run_provenance.get("release"))),
    ]
    cells = "".join(f"<dt>{label}</dt><dd>{value or '—'}</dd>"
                    for label, value in rows if value or label in
                    ("Checked against release", "Network checks"))
    bits.append(f'<dl class="meta">{cells}</dl>')
    return "\n".join(bits)


#: The screen's two ways of saying "this name is not available": the release already owns
#: it, or a submission still in the queue asked for it first. The report treats them the
#: same when offering a free alternative — the submitter has to be given another name
#: either way — and differently when saying why, which the finding headline handles.
_NAME_UNAVAILABLE_CODES = ("name_already_in_malavi", "name_claimed_by_another_submission")


def _claimed_names(screen: Optional[Any]) -> set:
    """Proposed names the screen found are not available to this submitter."""
    reports = screen if isinstance(screen, list) else ([screen] if screen else [])
    out = set()
    for report in reports:
        for issue in report.get("issues", []):
            if issue.get("code") in _NAME_UNAVAILABLE_CODES and issue.get("subject"):
                out.add(str(issue["subject"]))
    return out


def _name_suggestions(proposed: Sequence[Dict[str, Any]], taken: set,
                      screen: Optional[Any] = None) -> Dict[str, str]:
    """A free name for each proposed name that is not free, keyed by the taken name.

    Offered in the report rather than left to a follow-up email, so that approving the
    submission approves the correction with it. The suggestion comes from the same code
    the website offers submitters, because a curator told one name and a submitter told
    another is worse than neither being told anything.
    """
    if not taken:
        return {}
    # Prefer what the screen recorded: that is the value a curator was shown and therefore
    # the one they approved. Recomputing here could quietly produce a different name if the
    # release changed between screening and rendering.
    recorded = {}
    for report in (screen if isinstance(screen, list) else ([screen] if screen else [])):
        recorded.update(report.get("name_suggestions") or {})
    if recorded:
        return {k: v for k, v in recorded.items() if k in taken}
    try:
        from .naming import suggest_name
        from .gate import load_snapshot
        names = (load_snapshot() or {}).get("lineages") or []
    except Exception:                                   # noqa: BLE001
        return {}
    if not names:
        return {}

    out: Dict[str, str] = {}
    for entry in proposed:
        name = str(entry.get("lineage_name") or "")
        if name not in taken:
            continue
        suggestion = suggest_name(entry.get("host_species") or "", names)
        if suggestion.ok and suggestion.proposal:
            out[name] = suggestion.proposal
    return out


def _provenance_section(run: CheckRun) -> str:
    """Which software produced this report. Last, because only a maintainer needs it."""
    p = run.provenance or {}
    rows = [("Schema version", p.get("schema_version")),
            ("Curation code", p.get("code_version")),
            ("Tool version", p.get("tool_version")),
            ("Network checks", "enabled" if p.get("online") else "off"),
            ("Report generated", p.get("generated"))]
    cells = "".join(f"<dt>{label}</dt><dd>{esc(value) or '—'}</dd>" for label, value in rows)
    return ('<h2 class="display">Report production details</h2>'
            f'<dl class="meta">{cells}</dl>')


DISPOSITIONS = {
    "blocked": ("Needs correction before curation",
                "Something here has to be resolved before these records can go in."),
    "incomplete": ("Validation incomplete",
                   "At least one check could not run, so part of this submission is "
                   "unexamined. Absence of a finding from those checks means nothing."),
    "review": ("Ready for curator review",
               "The automated checks completed."),
    "clear": ("No automated problems found",
              "The automated checks completed and raised nothing. That is not the same as "
              "correct — it is the part a program can see."),
}


def _disposition(run: CheckRun) -> str:
    """Where this submission stands, in one line, before any detail.

    Deliberately never says "valid" or "approved". These checks support a curator's
    judgment and do not replace it, and a report that announced a submission as valid
    would be inviting exactly the rubber-stamp the review process exists to prevent.
    """
    results = list(run.results)
    blocking = any(r.outcome is Outcome.FINDING
                   and _severity_class(r.check_id) == "blocking" for r in results)
    errored = any(r.outcome is Outcome.ERROR for r in results)
    findings = any(r.outcome is Outcome.FINDING for r in results)
    skipped = any(r.outcome is Outcome.SKIP for r in results)

    if blocking:
        state = "blocked"
    elif errored:
        state = "incomplete"
    elif findings or skipped:
        state = "review"
    else:
        state = "clear"

    label, explain = DISPOSITIONS[state]
    return (f'<div class="disposition {state}">'
            f'<span class="d-label">{esc(label)}</span>'
            f'<span class="d-note">{esc(explain)}</span></div>')


def _summary_section(submission: Dict[str, Any], screen: Optional[Any]) -> str:
    """What is in this submission, before any opinion about it.

    First, because it is a curator's first question and the old layout answered it on page
    six, behind every check the machine had run. A report that leads with its own output
    rather than with the thing being reviewed teaches the reader that it is written for the
    program, not for them.
    """
    proposed = submission.get("proposed_lineages") or []
    records = submission.get("records") or []
    vectors = submission.get("vectors") or []
    reference = submission.get("reference") or {}

    sequences = 0
    reports = screen if isinstance(screen, list) else ([screen] if screen else [])
    for report in reports:
        sequences += len(report.get("sequences", []))

    # Which lineages were found in which host, so the host list answers the question a
    # curator actually has rather than just listing species.
    by_host: Dict[str, set] = {}
    for record in records:
        host = record.get("host_species")
        if host:
            by_host.setdefault(host, set()).add(str(record.get("lineage_name") or ""))
    countries = sorted({r.get("country") for r in records if r.get("country")})

    # Names the screen says MalAvi already owns. Shown in red with a real alternative, so
    # accepting the report accepts the fix too rather than starting a separate exchange.
    taken = _claimed_names(screen)
    suggestions = _name_suggestions(proposed, taken, screen)
    # Names whose sequence turned out to be a lineage MalAvi already has. These are not
    # new lineages at all, so they get told that rather than offered a rename.
    not_new: Dict[str, str] = {}
    for report in (screen if isinstance(screen, list) else ([screen] if screen else [])):
        recorded = report.get("not_new_lineages") or {}
        if isinstance(recorded, dict):
            not_new.update(recorded)
        else:                       # older screens recorded a bare list of names
            not_new.update({n: "" for n in recorded})

    def tile(number, label):
        return (f'<div class="tile"><span class="n">{number}</span>'
                f'<span class="k">{esc(label)}</span></div>')

    tiles = "".join([
        tile(len(proposed), "new lineage" + ("s" if len(proposed) != 1 else "")),
        tile(sequences, "sequence" + ("s" if sequences != 1 else "")),
        tile(len(records), "host record" + ("s" if len(records) != 1 else "")),
        tile(len(vectors), "vector record" + ("s" if len(vectors) != 1 else "")),
    ])

    lines = []
    if proposed:
        parts = []
        for entry in proposed:
            name = str(entry.get("lineage_name") or "")
            if name in not_new:
                match = not_new.get(name) or "a lineage MalAvi already has"
                parts.append(
                    f'<span class="taken-name">{esc(name)}</span>'
                    f'<span class="taken-note"> (Not a new lineage &mdash; this sequence '
                    f'is <b>{esc(match)}</b>, which MalAvi already holds. Record it '
                    f'against {esc(match)} rather than naming it.)</span>')
            elif name in taken:
                fix = suggestions.get(name)
                note = (f" (Name already taken; suggesting &rarr; <b>{esc(fix)}</b>)"
                        if fix else " (Name already taken)")
                parts.append(f'<span class="taken-name">{esc(name)}</span>'
                             f'<span class="taken-note">{note}</span>')
            else:
                parts.append(f'<span class="mono">{esc(name)}</span>')
        lines.append("<dt>Names claimed</dt><dd>" + "<br>".join(parts) + "</dd>")
    if by_host:
        shown = []
        for host in sorted(by_host)[:6]:
            found = ", ".join(sorted(n for n in by_host[host] if n))
            shown.append(f"<i>{esc(host)}</i>"
                         + (f" <span class='mono'>({esc(found)})</span>" if found else ""))
        extra = (f" and {len(by_host) - 6} more" if len(by_host) > 6 else "")
        lines.append("<dt>Hosts</dt><dd>" + "<br>".join(shown) + esc(extra) + "</dd>")
    if countries:
        # The country alone is rarely the answer to "where"; the site is what a curator
        # recognises and what makes a record checkable against the paper.
        by_country: Dict[str, set] = {}
        for record in records:
            country = record.get("country")
            if country:
                site = (record.get("site") or "").strip()
                by_country.setdefault(country, set())
                if site:
                    by_country[country].add(site)
        shown = []
        for country in sorted(by_country)[:8]:
            sites = sorted(by_country[country])
            shown.append(esc(country)
                         + (f" ({esc(', '.join(sites))})" if sites else ""))
        extra = f" and {len(by_country) - 8} more" if len(by_country) > 8 else ""
        lines.append("<dt>Where</dt><dd>" + "<br>".join(shown) + esc(extra) + "</dd>")
    if reference.get("journal") or reference.get("year"):
        cite = " ".join(filter(None, [str(reference.get("journal") or ""),
                                      str(reference.get("year") or "")]))
        lines.append(f"<dt>Published in</dt><dd>{esc(cite)}</dd>")

    detail = f"<dl class='meta'>{''.join(lines)}</dl>" if lines else ""
    return ('<h2 class="display">What was submitted</h2>'
            f'<div class="tiles">{tiles}</div>{detail}')


def _verdict_section(submission_id: Optional[str], revision: int) -> str:
    """The block a curator acts from: what this submission is, and one link to answer.

    The link carries the submission id and the revision, so neither is typed. That is not
    convenience: an approval has to be bound to the version of this report the curator
    actually read, and a hand-typed identifier attaches a decision to whatever was typed.
    The first live test of the verdict form produced a submission id of "testing1,2,3",
    which is what typing one looks like in practice.

    Without an id — a report generated outside the submission tree, say — the link is
    omitted rather than pointing at a blank form, and the reason is stated. A form that
    cannot record a usable verdict is worse than no link, because the curator only finds
    out after they have read everything and pressed Submit.
    """
    if not submission_id:
        return ('<h2 class="display">Recording your decision</h2>'
                '<p class="note">This report was generated outside a submission '
                'directory, so it has no submission identifier and no verdict link. '
                'Generate it from the submission itself to get one.</p>')

    try:
        review = load_config().get("review") or {}
        url = _safe_url(prefill_url(review["verdict_form_url"],
                                    review["verdict_form_entries"],
                                    submission_id, revision))
        if not url:
            raise KeyError("verdict_form_url is not an allowed https URL")
    except (KeyError, TypeError):
        return ('<h2 class="display">Recording your decision</h2>'
                '<p class="note">The verdict form is not configured in '
                'config/project.yml, so no link could be built.</p>')

    return f"""<h2 class="display">Recording your decision</h2>
<div class="verdict">
  <p><b>{esc(submission_id)}</b> &middot; revision {revision}</p>
  <p><a class="verdict-link" href="{esc(url)}">Record your decision on this submission</a></p>
  <p class="verdict-note">Please click the link above to submit your report. If you accept
  the submission and there are suggested name changes for novel lineages, you are approving
  those suggestions as they appear in this report.</p>
</div>"""


def _standing_section(entry: Optional[Any]) -> str:
    """The flags and corrections already on this submission, **with the ids that name them**.

    Why this exists. Three of the verdict form's five actions need an identifier typed into
    a box: "Which flag are you withdrawing?", "Which hold are you clearing?" and "Which
    correction are you approving?". Those ids -- ``V1``, ``C1`` -- are minted in the review
    ledger, which is gitignored and which no curator can open. So the actions existed, the
    ledger enforced their rules, and there was no way for the person using the form to learn
    what to type. A curator could raise a flag and then not be able to withdraw it; a lead
    could not approve a correction, which is the gate every correction stops at.

    Rendered from the ledger entry rather than recomputed, so "standing" here means exactly
    what it means to the ledger: ``Verdict.blocks`` is a blocking verdict that has not been
    retracted or overridden, and a pending correction is one no lead has approved yet.

    Nothing is shown when nothing is standing. A report on a clean submission should not
    carry an empty table explaining machinery the reader does not need.

    This section is for curators and goes no further: the report is delivered to the curator
    registry and nowhere else. The reason text quotes the submitter's data, which is why the
    ledger it comes from is gitignored -- the audience here is the people who wrote it.
    """
    if entry is None:
        return ""

    # Duck-typed on purpose. The rule about what blocks belongs to the ledger, and reading
    # it off the objects keeps this from becoming a second opinion about which flags count.
    standing = [verdict for verdict in getattr(entry, "verdicts", []) or []
                if getattr(verdict, "blocks", False)]
    pending = [correction for correction in getattr(entry, "corrections", []) or []
               if not getattr(correction, "approved", False)
               and not getattr(correction, "applied", False)]
    approved = [correction for correction in getattr(entry, "corrections", []) or []
                if getattr(correction, "approved", False)
                and not getattr(correction, "applied", False)]

    if not standing and not pending and not approved:
        return ""

    blocks = ['<h2 class="display">Flags and corrections on this submission</h2>']

    if standing:
        rows = "\n".join(
            f"<tr><td class=\"handle\">{esc(verdict.id)}</td>"
            f"<td>{esc(verdict.curator)}</td>"
            f"<td>{esc(_short_date(verdict.at))}</td>"
            f"<td>revision {esc(verdict.revision)}</td>"
            f"<td>{esc(verdict.reason_text or verdict.reason_code or '—')}</td></tr>"
            for verdict in standing)
        blocks.append(
            '<p>These flags are blocking this submission. It cannot be approved while any '
            'of them stands.</p>'
            '<div class="tablewrap"><table>'
            '<thead><tr><th>Flag</th><th>Raised by</th><th>When</th><th>On</th>'
            '<th>Why</th></tr></thead>'
            f'<tbody>{rows}</tbody></table></div>'
            '<p class="note">Quote the <b>Flag</b> id on the verdict form. Use '
            '&ldquo;Withdraw a flag you placed yourself&rdquo; for your own, or &ldquo;Clear '
            'another curator&rsquo;s hold&rdquo; if you are a lead clearing somebody '
            'else&rsquo;s.</p>')

    if pending:
        rows = "\n".join(
            f"<tr><td class=\"handle\">{esc(correction.id)}</td>"
            f"<td>{esc(correction.by)}</td>"
            f"<td>{esc(_short_date(correction.at))}</td>"
            f"<td>{esc(correction.change)}</td></tr>"
            for correction in pending)
        blocks.append(
            '<p>These corrections are waiting for a lead curator to approve them. Nothing '
            'is applied until one does.</p>'
            '<div class="tablewrap"><table>'
            '<thead><tr><th>Correction</th><th>Proposed by</th><th>When</th>'
            '<th>What should change</th></tr></thead>'
            f'<tbody>{rows}</tbody></table></div>'
            '<p class="note">A lead quotes the <b>Correction</b> id under &ldquo;Approve a '
            'correction&rdquo; on the verdict form.</p>')

    if approved:
        rows = "\n".join(
            f"<tr><td class=\"handle\">{esc(correction.id)}</td>"
            f"<td>{esc(correction.approved_by)}</td>"
            f"<td>{esc(correction.change)}</td></tr>"
            for correction in approved)
        blocks.append(
            '<p>These corrections have been approved and are waiting to be applied, which '
            'will produce a new revision of this submission.</p>'
            '<div class="tablewrap"><table>'
            '<thead><tr><th>Correction</th><th>Approved by</th>'
            '<th>What should change</th></tr></thead>'
            f'<tbody>{rows}</tbody></table></div>')

    return "\n".join(blocks)


def _short_date(stamp: Any) -> str:
    """The date part of an ISO timestamp. The time of day is noise in this table."""
    text = str(stamp or "")
    return text.split("T", 1)[0] if "T" in text else text


# Which heading each check belongs under. A curator does not think in check ids; they
# think "is the naming right", "is the sequence right", "do the host and place make
# sense". Anything unlisted falls to the end under "Other", so adding a check cannot make
# it silently disappear from the report.
CHECK_GROUPS = (
    ("Naming", (
        "name_already_in_malavi", "name_claimed_by_another_submission",
        "lineage_known", "lineage_not_in_malavi",
        "lineage_unresolved", "lineage_accession_conflict", "lineage_missing",
        "lineage_without_sequence", "sequence_without_declaration",
        "lineage_without_host_record", "accession_malformed", "accession_format",
        "name_malformed",
        "accession_collision", "accession_resolves")),
    ("Sequences", (
        "sequence_is_known_lineage", "sequence_qc", "sequence_needs_reframing",
        "sequence_stop_codon", "sequence_unplaceable")),
    ("Hosts and geography", (
        "host_geography_plausible", "host_not_in_malavi", "host_name_resolves",
        "host_missing", "host_binomial", "country_not_in_malavi", "country_missing",
        "record_without_country", "pair_already_in_malavi", "new_country",
        "prevalence_sanity", "prevalence_inconsistent", "record_without_prevalence",
        "lineage_previously_recorded")),
    ("Vectors", ("vector_missing", "vector_sanity", "vector_method")),
    ("The submission itself", (
        "headers_intact", "reference_missing", "reference_unpubl_malformed",
        "source_reprinted", "source_uncertain",
        "values_normalized", "lineage_missing", "records_unlinked")),
)


# What to call a finding, in words that say what was found rather than what was asserted.
# A check title is written as the thing being tested -- "Proposed names are free" -- and
# printed above a finding it reads as a pass. `{subjects}` is filled with the names the
# finding is about, because "a name is taken" and "TUMIG10 is taken" are different amounts
# of help at the moment a curator is deciding.
FINDING_HEADLINES = {
    "name_already_in_malavi":
        "{subjects} was listed as new, but MalAvi already has that name",
    "name_claimed_by_another_submission":
        "{subjects} was already claimed by an earlier submission still under review",
    "sequence_is_known_lineage":
        "{subjects} was listed as new, but the sequence is already in MalAvi",
    "sequence_qc": "{subjects}: unusual for a cytochrome b barcode",
    "sequence_needs_reframing": "{subjects} may not be in the standard reading frame",
    "sequence_stop_codon": "{subjects} contains a stop codon",
    "sequence_unplaceable": "{subjects} could not be placed against the alignment",
    "accession_malformed": "{subjects} has a malformed GenBank accession",
    "accession_collision": "{subjects} uses an accession already recorded elsewhere",
    "lineage_accession_conflict": "{subjects} disagrees with MalAvi about its accession",
    "lineage_without_sequence": "{subjects} was named but no sequence was given",
    "sequence_without_declaration": "{subjects} has a sequence but was never declared",
    "lineage_without_host_record": "{subjects} was named but has no host record",
    "host_name_resolves": "{subjects} is not a host name we could resolve",
    "host_geography_plausible": "{subjects}: this host and place are unusual for it",
    "pair_already_in_malavi":
        "{subjects}: MalAvi already records this parasite in this host",
    "prevalence_sanity": "{subjects} has prevalence numbers that do not add up",
    "prevalence_inconsistent": "{subjects} has prevalence numbers that disagree",
    "record_without_country": "{subjects} has no country",
    "record_without_prevalence": "{subjects} reports no numbers tested or found",
    "name_malformed": "{subjects} does not look like a name built from its host",
    "reference_missing": "The reference for this submission is incomplete",
    "reference_unpubl_malformed":
        "{subjects} does not follow MalAvi's naming for an unpublished study",
    "accession_format": "{subjects} has an accession that is not in the expected format",
    "headers_intact": "The workbook's column headers were changed or are missing",
    "host_missing": "{subjects} has a record with no host species",
    "lineage_missing": "{subjects} has a record with no lineage name",
    "vector_missing": "{subjects} has a vector record with no vector species",
    "vector_sanity": "{subjects} names a vector that is not a known genus",
}


ALLOWED_URL_HOSTS = ("docs.google.com", "drive.google.com")

# How often the chimera screen fires on a sequence that is certainly not a chimera.
#
# Measured 2026-08-20 by leave-one-out over 250 curated lineages (the ones that are a
# clean 479 bp of ACGT -- 2,027 of the release's 5,365): remove each from the reference
# and screen it as though it had just been submitted. 58 of 250 came back
# "possible_chimera", none of which is one.
#
# The two conditions are not equal partners. `parent_switches >= 2` was met by 246 of the
# 250 (median 8, max 16), so it separates almost nothing: among lineages a base or two
# apart -- the median distance to the nearest relative is 1 bp -- which one wins a 120 bp
# window is close to arbitrary. `chimera_delta >= 3` was met by exactly the 58 that were
# flagged, so the call is that term alone. And that term reduces to "the nearest single
# relative is >= 3 bp away while windows match better", which is a novelty measure --
# and novelty is what a submission is full of.
#
# The number is printed beside every chimera call, because a curator cannot weigh the
# call without it. Re-measure with curation/r (leave-one-out over the release) if the
# thresholds in malaviR's .qc_settings change.
CHIMERA_FALSE_POSITIVE_RATE = "23% (58 of 250 tested)"

# malaviR's lineage_qc calls, as its final-call chain assigns them (R/lineage_qc.R):
# a stop codon in frame wins outright, then an exact match, then a chimera flag, and
# only then the plausibility score, banded at 0.85 / 0.60 / 0.35.
#
# Each call gets its own words and its own gloss. Two of them used to share a phrase,
# which meant a chimera and a low score were indistinguishable on the page, and the
# ladder had no stated order so "review" and "Worth a close look" looked unrelated.
# If the thresholds in malaviR change, the score bands quoted here must change with them.
QC_CALLS: Dict[str, tuple] = {
    "invalid_or_strong_warning": (
        "Invalid — stop codon in reading frame",
        "The sequence translates with a stop codon. Usually the sequence was pasted "
        "outside the 479 bp window rather than anything wrong with the parasite — check "
        "the framing before reading anything else here."),
    "known_lineage": (
        "Already in MalAvi",
        "Identical to a lineage MalAvi already holds, over every position both cover."),
    "possible_chimera": (
        "Possible chimera",
        "Different stretches of the barcode have different closest relatives. A 120 bp "
        "window is slid along the sequence and asked which lineage it matches best; when "
        "that answer keeps changing, the sequence looks like a mosaic of two or more "
        "templates rather than one parasite. Not a score — this call is made on that "
        "pattern alone, and the stretches are listed under the call. Read it as a prompt "
        "to look rather than as a finding: tested by removing curated lineages from the "
        "reference and re-screening them, " + CHIMERA_FALSE_POSITIVE_RATE + " of "
        "sequences that are certainly not chimeras are called one."),
    "plausible_new_lineage": (
        "Plausible new lineage",
        "Score of 0.85 or better: nothing about the sequence itself looks wrong."),
    "review": (
        "Worth a look",
        "Score between 0.60 and 0.85. Something is mildly unusual — the notes beneath "
        "the alignment say what."),
    "strong_warning": (
        "Worth a close look",
        "Score between 0.35 and 0.60. Several unusual features at once."),
    "possible_error": (
        "Probably an error",
        "Score below 0.35. More likely a sequencing or transcription problem than a "
        "real parasite."),
}


def _qc_call_key(calls_present: Sequence[str]) -> str:
    """The key for whichever QC calls this submission actually produced."""
    seen = [c for c in QC_CALLS if c in set(calls_present)]
    if not seen:
        return ""
    items = "".join(f"<dt>{esc(QC_CALLS[c][0])}</dt><dd>{esc(QC_CALLS[c][1])}</dd>"
                    for c in seen)
    return ('<p class="sub">Each sequence carries one call from malaviR\'s quality '
            'screen. The ones in this submission mean:</p>'
            f'<dl class="tierkey">{items}</dl>')


def _safe_url(url: str) -> Optional[str]:
    """A URL fit to put in an href, or None.

    Escaping is not enough. `esc()` makes a string safe to *display*; it does nothing
    about what a scheme *does* when followed, so `javascript:...` survives escaping intact.
    Every URL in this report comes from a submitter's form answer or from configuration, so
    the rule is an allowlist rather than a blocklist: https only, and only to the two hosts
    this report ever legitimately links to.

    The CSP meta tag does not cover this. It constrains a browser; WeasyPrint ignores it,
    and the PDF is the copy curators actually open.
    """
    from urllib.parse import urlparse
    try:
        parsed = urlparse(str(url or "").strip())
    except ValueError:
        return None
    if parsed.scheme != "https" or not parsed.netloc:
        return None
    host = parsed.netloc.split("@")[-1].split(":")[0].lower()
    if host not in ALLOWED_URL_HOSTS and not any(
            host.endswith("." + allowed) for allowed in ALLOWED_URL_HOSTS):
        return None
    return parsed.geturl()


def _humanize(message: str) -> str:
    """Turn a screen's code strings into something a biologist reads without decoding.

    The sequence QC verdict arrives as tokens -- "strong_warning —
    moderately_divergent_from_known_lineages; 2_nonsynonymous_changes_vs_nearest_lineage".
    Every one of those is a phrase with the spaces taken out, so putting them back is
    almost the whole job; the rest is a handful that read badly straight.
    """
    swaps = {
        "vs nearest lineage": "compared with the nearest lineage",
        "moderately divergent from known lineages":
            "moderately divergent from known lineages",
        "possible chimera or mixed template pattern":
            "a pattern that can indicate a chimera or mixed template",
        "rare bases at sites": "bases that are rare at those positions",
        "second codon position changes": "changes at second codon positions",
        "nonsynonymous changes": "amino-acid-changing differences",
    }
    # Identifiers must survive the underscore strip below. malaviR's messages name its
    # own functions -- "see frame_to_malavi" -- and blanket-replacing underscores turned
    # that into "see frame to malavi", which reads as three words of broken English
    # rather than as something a reader can go and look up.
    text = str(message or "")
    for identifier in ("frame_to_malavi", "blast_malavi", "lineage_qc", "lineage_screen",
                       "match_taxonomy", "clean_alignment", "extract_alignment"):
        text = text.replace(identifier, "\x00" + identifier + "\x00")
    parts = text.split("\x00")
    text = "".join(part if index % 2 else part.replace("_", " ")
                   for index, part in enumerate(parts))
    text = text.replace("\x00", "")
    for code, plain in swaps.items():
        text = text.replace(code, plain)

    # The QC call is the FIRST token of the message and is translated only there.
    # Translating it wherever it appeared would rewrite ordinary prose -- "review"
    # occurs in "a curator's review" -- and that is a rename of somebody's words, not
    # a glossary. Until 2026-08-19 "strong_warning" and "possible_chimera" both became
    # "Worth a close look", losing the one thing a chimera call says, and "review" was
    # left as a raw token beside them, so nothing indicated which was worse.
    for token in sorted(QC_CALLS, key=len, reverse=True):
        spaced = token.replace("_", " ")
        if text.startswith(spaced):
            text = QC_CALLS[token][0] + text[len(spaced):]
            break
    return text


def _emphasise(escaped: str) -> str:
    """Bold the one value a curator is being asked to accept.

    Applied to already-escaped text, against a pattern this code writes itself, so a
    submitter cannot smuggle markup in through a finding message.
    """
    import re as _re
    return _re.sub(r"Suggesting ([A-Z0-9]+) for this lineage",
                   r"Suggesting <b>\1</b> for this lineage", escaped)


def _headline(result, check) -> str:
    """A sentence saying what was found. Falls back to the check title if unnamed."""
    template = FINDING_HEADLINES.get(result.check_id)
    if not template:
        return check.title if check else result.check_id
    subjects = []
    for finding in result.findings:
        subject = _subject(finding.subject)
        if subject and subject not in subjects:
            subjects.append(subject)
    if not subjects:
        return template.format(subjects="One entry")
    shown = ", ".join(subjects[:3]) + (" and others" if len(subjects) > 3 else "")
    return template.format(subjects=shown)


def _group_of(check_id: str) -> str:
    for title, ids in CHECK_GROUPS:
        if check_id in ids:
            return title
    return "Other"


def _subject(text: Any) -> str:
    """A finding's subject as a curator would name it.

    Internal row indices -- "row[0]" -- are what a program calls a row and mean nothing to
    a reader, so they are dropped in favour of the location, which the caller adds anyway.
    """
    value = str(text or "")
    return "" if re.fullmatch(r"row\[\d+\]", value) else value


def _clean_summaries(submission: Optional[Dict[str, Any]], screen: Optional[Any],
                     attention: Sequence[Any]) -> Dict[str, str]:
    """One green sentence per group, naming what came through clean."""
    submission = submission or {}
    reports = screen if isinstance(screen, list) else ([screen] if screen else [])

    flagged_in: Dict[str, set] = {}
    for report in reports:
        for issue in report.get("issues", []):
            subject = issue.get("subject")
            if subject:
                flagged_in.setdefault(_group_of(issue.get("code", "")), set()).add(
                    str(subject))
    naming_flagged = flagged_in.get("Naming", set())
    sequence_flagged = flagged_in.get("Sequences", set())

    out: Dict[str, str] = {}

    proposed = [str(e.get("lineage_name") or "")
                for e in (submission.get("proposed_lineages") or [])]
    free = [n for n in proposed if n and n not in naming_flagged]
    if free:
        out["Naming"] = (f"Proposed name{'s' if len(free) != 1 else ''} "
                         f"({', '.join(free)}) "
                         f"{'are' if len(free) != 1 else 'is'} available")

    sequences = [str(entry.get("label") or "")
                 for report in reports for entry in report.get("sequences", [])]
    clean_seq = [n for n in sequences if n and n not in sequence_flagged]
    if clean_seq:
        out["Sequences"] = (f"Sequence{'s' if len(clean_seq) != 1 else ''} "
                            f"({', '.join(clean_seq)}) look"
                            f"{'' if len(clean_seq) != 1 else 's'} new")

    records = submission.get("records") or []
    if records and not any(_group_of(r.check_id) == "Hosts and geography"
                           for r in attention):
        out["Hosts and geography"] = (
            f"All {len(records)} host record{'s' if len(records) != 1 else ''} "
            f"resolved to known hosts and places")
    return out


def _checks_section(run: CheckRun, workbook: Optional[str] = None,
                    submission: Optional[Dict[str, Any]] = None,
                    screen: Optional[Any] = None) -> str:
    attention = [r for r in run.results
                 if r.outcome in (Outcome.ERROR, Outcome.FINDING)]
    skipped = [r for r in run.results if r.outcome is Outcome.SKIP]
    passed = [r for r in run.results if r.outcome is Outcome.PASS]

    counts = run.counts()
    bits = []
    blocking = sum(1 for r in attention
                   if _severity_class(r.check_id) == "blocking")
    if blocking:
        bits.append(f'<p class="sub"><b>{blocking} of these stop the submission '
                    f'until they are resolved.</b></p>')

    if not attention:
        bits.append('<div class="note">No check found anything to raise. Read the '
                    'skipped list below before concluding the submission is clean: a '
                    'check that did not run has no opinion.</div>')

    # What is fine, said specifically. "Proposed names are available" helps nobody;
    # "Proposed names (TUMIG31) are available" tells a curator which one they can stop
    # thinking about.
    clean_by_group = _clean_summaries(submission, screen, attention)

    ordered = sorted(attention, key=lambda r: (
        [t for t, _ in CHECK_GROUPS] .index(_group_of(r.check_id))
        if _group_of(r.check_id) != "Other" else len(CHECK_GROUPS),
        r.check_id))

    # Informational findings get one line, not a card. They are things like "this lineage
    # is not in MalAvi yet", which is true of every new lineage and is already stated on
    # the first page — printed as full cards they bury the two findings that actually stop
    # the submission, which is the opposite of what this section is for.
    def _is_minor(result) -> bool:
        return (result.outcome is Outcome.FINDING
                and _severity_class(result.check_id) == "info")

    current_group = None
    minor_by_group: Dict[str, List[str]] = {}
    # Groups that actually put a card on the page. A group whose findings were all
    # informational renders nothing, so it still needs its green line.
    shown_groups: set = set()

    for result in ordered:
        group = _group_of(result.check_id)
        if _is_minor(result):
            check = CHECKS.get(result.check_id)
            label = check.title if check else result.check_id
            for finding in result.findings:
                subject = _subject(finding.subject)
                minor_by_group.setdefault(group, []).append(
                    (f'<b>{esc(subject)}</b> — ' if subject else "")
                    + esc(finding.message))
            continue
        if group != current_group:
            current_group = group
            bits.append(f'<h3 class="group">{esc(group)}</h3>')
            # The Sequences group leads with every QC call it is about to use. The calls
            # are a ladder and nothing on the page said so, so a curator meeting "Worth a
            # look" and "Worth a close look" had no way to rank them.
            if group == "Sequences":
                calls = [str((f.evidence or {}).get("call") or "")
                         for r in ordered if r.check_id == "sequence_qc"
                         for f in r.findings]
                key = _qc_call_key([c for c in calls if c])
                if key:
                    bits.append(key)
            # What came through clean in this group, named. "Proposed names are available"
            # helps nobody; "Proposed name (TUMIG31) is available" tells a curator which
            # one they can stop thinking about.
            shown_groups.add(group)
            clean = clean_by_group.get(group)
            if clean:
                bits.append(f'<div class="check clear"><div class="hd">'
                            f'<span class="title">{esc(clean)}</span>'
                            f'<span class="pill clear">fine</span></div></div>')
        check = CHECKS.get(result.check_id)
        severity = _severity_class(result.check_id)
        classes = f"check {result.outcome.value}"
        if result.outcome is Outcome.FINDING:
            classes += f" {severity}"

        head = (f'<span class="title">{esc(_headline(result, check))}</span>'
                f'<span class="pill {severity}">{severity}</span>')
        if result.outcome is Outcome.ERROR:
            head += '<span class="count">could not run</span>'

        # No assertion line on a finding card. The assertion is what the check tests for
        # ("No proposed new lineage name is already a MalAvi lineage"), and printed under a
        # headline saying one IS, it reads as a contradiction of the headline above it.
        body = ""
        if result.outcome is Outcome.ERROR:
            body += (f'<p class="asserts"><b>Error:</b> {esc(result.error)}</p>')
        else:
            items = ""
            for f in result.findings:
                subject = _subject(f.subject)
                where = _where(f.source, workbook)
                link = ""
                if current_group == "Sequences" and subject:
                    link = (f' <a class="jump" href="#aln-{esc(subject)}">'
                            f'see the alignment</a>'
                            f' <a class="jump" href="#seq-{esc(subject)}">'
                            f'see the sequence</a>')
                message = _emphasise(esc(_humanize(f.message)))
                items += (
                    f'<li><span class="subj">{esc(subject)}</span> '
                    + (f'<span class="where">— {esc(where)}</span>' if where else "")
                    + f'<br>{message}{link}'
                    + _chimera_evidence(f.evidence)
                    + '</li>')
            body += f'<ul class="finds">{items}</ul>'

        bits.append(f'<div class="{classes}"><div class="hd">{head}</div>{body}</div>')

    # Everything that ran and did not stop the submission, as a single reassuring line.
    # The individual notes were restatements of things page one already says -- "TUMIG31 is
    # not in MalAvi" is true of every new lineage -- and listing them made a clean
    # submission look busy.
    for group, clean in clean_by_group.items():
        if group not in shown_groups:
            bits.append(f'<h3 class="group">{esc(group)}</h3>'
                        f'<div class="check clear"><div class="hd">'
                        f'<span class="title">{esc(clean)}</span>'
                        f'<span class="pill clear">fine</span></div></div>')

    quiet = sum(len(lines) for lines in minor_by_group.values())
    # A check with nothing to examine has not "found nothing" -- it had no opportunity to
    # find anything, and counting it as a clean result overstates what was covered.
    examined = [r for r in passed if r.evaluated]
    vacuous = len(passed) - len(examined)
    if quiet or examined or vacuous:
        # Built as one string rather than a chain of f-strings around a conditional. The
        # previous version put `if ... else` inside an implicit concatenation, which binds
        # over the whole chain: the closing </div> lived in only one branch, so the box was
        # never closed and every section after it rendered inside a green panel.
        headline = (
            f"{len(examined) + quiet} other checks found nothing that stops this submission"
            if (len(examined) + quiet)
            else "No other check had anything of its kind to examine")
        note = ("They cover naming, sequences, accessions, hosts, localities and "
                "prevalence. The maintainer can supply the full list.")
        if vacuous:
            note += (f" {vacuous} additional test{'s were' if vacuous != 1 else ' was'} "
                     f"not relevant to this submission.")
        bits.append(
            '<div class="check clear">'
            f'<div class="hd"><span class="title">{esc(headline)}</span>'
            '<span class="pill clear">fine</span></div>'
            f'<p class="asserts">{esc(note)}</p>'
            '</div>')

    if skipped:
        bits.append("<h3>Did not run</h3>")
        bits.append('<div class="note">These checks have no opinion on this submission. '
                    'Nothing below is evidence that anything is correct.</div>')
        for result in skipped:
            check = CHECKS.get(result.check_id)
            bits.append(
                f'<div class="check skip"><div class="hd">'
                f'<span class="id">{esc(result.check_id)}</span></div>'
                + (f'<p class="asserts">{esc(check.asserts)}</p>' if check else "")
                + f'<p class="asserts"><b>Skipped:</b> {esc(result.skip_reason)}</p>'
                "</div>")


    return "\n".join(bits)


def _sequences_section(screen: Optional[Any]) -> str:
    """Per-sequence detail, from the template screen."""
    if not screen:
        return ""
    reports = screen if isinstance(screen, list) else [screen]
    sequences: List[Dict[str, Any]] = []
    for report in reports:
        sequences.extend(report.get("sequences", []) or [])
    if not sequences:
        return ""

    head = ("<tr><th>Lineage</th><th>Length</th><th>Offset</th><th>Stops</th>"
            "<th>Verdict</th><th>Nearest lineage</th><th>Distance</th>"
            "<th>Compared over</th></tr>")
    rows = []
    for entry in sequences:
        nearest = (entry.get("nearest") or [{}])[0]
        rows.append(
            "<tr>"
            f"<td class='mono'>{esc(entry.get('label'))}</td>"
            f"<td>{esc(entry.get('raw_length'))}</td>"
            f"<td>{esc(entry.get('offset'))}</td>"
            f"<td>{esc(entry.get('n_stop_codons'))}</td>"
            f"<td>{esc(entry.get('verdict'))}</td>"
            f"<td class='mono'>{esc(nearest.get('lineage'))}</td>"
            f"<td>{esc(nearest.get('distance'))}</td>"
            f"<td>{esc(nearest.get('comparable'))} positions</td>"
            "</tr>")

    # The sequences themselves, in FASTA. A curator meeting this pipeline for the first
    # time will not take its word for a lineage call, and should not have to: this is the
    # block they copy into their own aligner to check the program is right. Wrapped at 60
    # columns, which is what every tool expects and what stops it running off the page.
    fasta_lines = []
    for entry in sequences:
        raw = str(entry.get("sequence") or entry.get("raw") or "")
        if not raw:
            continue
        fasta_lines.append((entry.get("label"), [">" + str(entry.get("label"))]
                            + [raw[i:i + 60] for i in range(0, len(raw), 60)]))
    blocks = "".join(
        f'<div id="seq-{esc(label)}"><pre class="fasta">{esc(chr(10).join(lines))}</pre></div>'
        for label, lines in fasta_lines)
    fasta = ("<h3>The novel sequences submitted</h3>"
             "<p class='sub'>Copy these calls to check in your own sequence alignment software.</p>"
             + blocks) if fasta_lines else ""

    notes = []
    for entry in sequences:
        for note in entry.get("notes") or []:
            notes.append(f"<li><b>{esc(entry.get('label'))}</b> — {esc(note)}</li>")

    out = ['<h2 class="display">Sequences</h2>',
           f'<div class="tablewrap"><table>{head}{"".join(rows)}</table></div>']
    if notes:
        out.append(f'<ul class="finds">{"".join(notes)}</ul>')
    if fasta:
        out.append(fasta)
    return "\n".join(out)


def _alignment_section(figures: Optional[Sequence[Any]]) -> str:
    """The submitted sequences beside their nearest relatives, at the differing positions.

    This is the view that answers the question a distance cannot: *where* does it differ,
    and does anything else differ there too. A single change shared with three other
    lineages reads very differently from one nothing in MalAvi has ever shown.
    """
    if not figures:
        return ""

    out = ['<h2 class="display">Alignment with the nearest lineages</h2>',
           '<p class="sub">Only positions where something varies are shown; a dot is '
           'agreement with the submitted sequence. The submitted row is shaded by what '
           'each difference does, measured against the nearest lineage — the same '
           'quantities the quality screen reports as counts, put where they can be '
           'pointed at.</p>',
           '<div class="alnkey">'
           '<span class="sw k-nonsyn">A</span> changes the amino acid'
           '<span class="sw k-transversion">A</span> transversion, silent'
           '<span class="sw k-transition">A</span> transition, silent'
           '<span class="sw rare">A</span> base rare at this site '
           '(carried by 1% or less of the lineages across the whole release, all three genera, that cover it)'
           '</div>']

    for figure in figures:
        data = figure.as_dict() if hasattr(figure, "as_dict") else figure
        # The anchor the sequence findings link to, so a curator reaches the picture for
        # the sequence they were reading about rather than the top of the section.
        out.append(f'<h3 id="aln-{esc(data["label"])}">{esc(data["label"])}</h3>')

        if data.get("unavailable"):
            out.append(f'<div class="banner">{esc(data["unavailable"])}</div>')
            continue
        # Which arrangement of the bases this picture is of. Printed before the table,
        # because a reader who works that out afterwards has already misread it.
        if data.get("framing"):
            klass = "banner" if data.get("offset") else "note"
            out.append(f'<div class="{klass}">{esc(data["framing"])}</div>')

        if not data.get("positions"):
            out.append('<div class="note">Identical to the lineages shown at every '
                       'position they both cover.</div>')
            continue

        # Position numbers sit horizontally above their column. That widens each
        # column to roughly three digits, which is the reason MAX_COLUMNS is what it is.
        header = "".join(f'<th class="pos">{p}</th>' for p in data["positions"])
        columns = data.get("columns") or []
        rows = []
        for row in data["rows"]:
            cells_out = []
            for index, c in enumerate(row["cells"]):
                classes = ["b", c["state"]]
                title = ""
                # The submitted row carries the interpretation; the neighbours stay plain
                # so the eye has one thing to read per column rather than six.
                if row["is_query"] and index < len(columns):
                    column = columns[index]
                    if column.get("kind") in ("nonsynonymous", "transversion",
                                              "transition"):
                        classes.append("k-" + ("nonsyn" if column["kind"] ==
                                               "nonsynonymous" else column["kind"]))
                    if column.get("rare"):
                        classes.append("rare")
                    bits = [f"position {column['position']}",
                            f"codon position {column['codon_position']}"]
                    if column.get("amino_acid"):
                        bits.append(f"amino acid {column['amino_acid']}")
                    if column.get("share") is not None:
                        bits.append(f"this base at {column['share'] * 100:.1f}% "
                                    f"of covering lineages")
                    title = f' title="{esc("; ".join(bits))}"'
                cells_out.append(
                    f'<td class="{" ".join(classes)}"{title}>'
                    + ("." if c["state"] == "same" else esc(c["base"]))
                    + "</td>")
            cells = "".join(cells_out)
            distance = ("" if row["is_query"] else
                        f'<span class="dist">{row["distance"]} of {row["comparable"]}</span>')
            name_class = "seqname query" if row["is_query"] else "seqname"
            rows.append(f'<tr><th class="{name_class}">{esc(row["name"])}{distance}</th>'
                        f'{cells}</tr>')

        out.append('<div class="tablewrap alignment"><table class="aln">'
                   f'<tr><th class="seqname"></th>{header}</tr>{"".join(rows)}'
                   '</table></div>')

        # The caption used to be a count of differing positions, which the table above
        # already showed. What it says now is what the differences are -- which change
        # the protein, which are transversions, which bases are rare where they sit, and
        # whether they bunch into one stretch, the pattern a chimera leaves.
        caption = []
        if data.get("truncated"):
            caption.append(f'Showing the first {len(data["positions"])} of '
                           f'{data["n_differing"]} differing positions; the notes below '
                           f'describe all of them.')
        caption.extend(data.get("notes") or [])
        if caption:
            out.append('<p class="sub">' + " ".join(esc(s) for s in caption) + '</p>')
    return "\n".join(out)


def _names_section(submission: Dict[str, Any]) -> str:
    proposed = submission.get("proposed_lineages") or []
    if not proposed:
        return ""
    head = ("<tr><th>Proposed name</th><th>Parasite</th><th>Representative host</th>"
            "<th>Accessions</th><th>Where</th></tr>")
    rows = "".join(
        "<tr>"
        f"<td class='mono'>{esc(entry.get('lineage_name'))}</td>"
        f"<td>{esc(entry.get('parasite_genus'))}</td>"
        f"<td>{esc(entry.get('host_species'))}</td>"
        f"<td class='mono'>{esc(', '.join(entry.get('accessions') or []))}</td>"
        f"<td class='rowno'>{esc(_where(entry.get('source')))}</td>"
        "</tr>" for entry in proposed)
    return ('<h2 class="display">Names claimed</h2>'
            f'<div class="tablewrap"><table>{head}{rows}</table></div>')


def _normalization_section(submission: Dict[str, Any]) -> str:
    """What the system reinterpreted, beside what the submitter wrote."""
    provenance = submission.get("provenance") or {}
    changes = provenance.get("normalizations") or []
    repairs = provenance.get("header_repairs") or []
    if not changes and not repairs:
        return ""

    out = ['<h2 class="display">How the file was read</h2>',
           '<div class="note">Everything the system decided about the submitter\'s '
           'file. Values are shown as submitted beside how they were read, so a '
           'correction can always be traced back to what was actually sent.</div>']
    if changes:
        head = "<tr><th>Field</th><th>Submitted</th><th>Read as</th><th>Where</th></tr>"
        rows = "".join(
            "<tr>"
            f"<td>{esc(change.get('field'))}</td>"
            f"<td class='mono was'>{esc(change.get('submitted'))}</td>"
            f"<td class='mono'>{esc(change.get('normalized'))}</td>"
            f"<td class='rowno'>{esc(_where(change.get('source')))}</td>"
            "</tr>" for change in changes)
        out.append(f'<div class="tablewrap"><table>{head}{rows}</table></div>')
    if repairs:
        out.append('<ul class="finds">'
                   + "".join(f"<li>{esc(note)}</li>" for note in repairs)
                   + "</ul>")
    return "\n".join(out)


def _chimera_evidence(evidence: Optional[Dict[str, Any]]) -> str:
    """Which stretch of the barcode matches which lineage, under a chimera call.

    A "possible chimera" verdict is otherwise an assertion with nothing behind it. What
    malaviR actually did is slide a 120 bp window along the sequence and record the
    nearest lineage in each; the call fires when that answer keeps changing. Printing the
    runs turns the verdict into something a curator can agree or disagree with.

    The caveat is printed with them, because it is the usual explanation: among lineages
    that are themselves a base or two apart, which one wins a short window is close to
    arbitrary, and switching between near-identical relatives is much weaker evidence
    than switching between distant ones.
    """
    if not evidence or (evidence.get("call") != "possible_chimera"):
        return ""
    runs = evidence.get("chimera_runs") or []
    if not runs:
        return ""

    parts = "".join(
        f'<li><span class="mono">{esc(r.get("lineage"))}</span> — positions '
        f'{esc(_number(r.get("start")))} to {esc(_number(r.get("end")))}'
        + (f', matching exactly' if (r.get("distance") == 0)
           else f', {esc(_number(r.get("distance")))} bp away')
        + '</li>'
        for r in runs)

    switches = evidence.get("chimera_parent_switches")
    best = evidence.get("chimera_best_single")
    best_distance = evidence.get("chimera_best_single_distance")
    lead = (f'The nearest lineage changes {esc(_number(switches))} time(s) along the '
            f'barcode. No single lineage is nearest throughout'
            + (f'; the best over the whole sequence is '
               f'<span class="mono">{esc(best)}</span> at '
               f'{esc(_number(best_distance))} bp' if best else "")
            + ':')
    return ('<div class="chimera"><p>' + lead + '</p>'
            f'<ul>{parts}</ul>'
            '<p class="caveat">Weigh this against its base rate. Curated MalAvi '
            'lineages removed from the reference and re-screened as though newly '
            'submitted are called possible chimeras ' + CHIMERA_FALSE_POSITIVE_RATE
            + ' of the time. The switching on its own means almost nothing -- 98% of '
            'those lineages switched parents twice or more (median 8): among lineages that '
            'differ by a base or two, which wins a 120 bp window is close to arbitrary. '
            'What is worth weighing is whether the stretches above point at '
            '<i>distant</i> relatives, and whether the run boundaries fall in the same '
            'place.</p></div>')


def _alt_names_section(submission: Dict[str, Any]) -> str:
    """Lineages MalAvi already holds that the submitter published under another name.

    This is the Alt_Lineage_names sheet, and it is a claim about MalAvi's existing data
    rather than about the submission: "the lineage you call SGS1 is the one I published
    as P15". It needs a curator's eye more than most rows do, because accepting it makes
    two names one lineage for everyone, and it is not recoverable by re-reading the
    submission later.

    Whether MalAvi actually holds the named lineage is stated, because the commonest
    error here is a submitter naming a lineage that does not exist -- a typo, or a name
    from a paper that MalAvi indexes differently.
    """
    rows = submission.get("alternative_names") or []
    if not rows:
        return ""

    try:
        from .release_index import load_release_index
        known = {str(n).strip().upper() for n in (load_release_index().lineages or ())}
    except Exception:                                          # noqa: BLE001
        known = set()

    body = ""
    for entry in rows:
        malavi = str(entry.get("malavi_name") or "")
        alt = str(entry.get("alternative_name") or "")
        accessions = ", ".join(entry.get("accessions") or [])
        if not known:
            status = '<td class="miss">not checked</td>'
        elif malavi.strip().upper() in known:
            status = '<td class="hit">in MalAvi</td>'
        else:
            status = ('<td class="bad"><b>no such lineage in MalAvi</b></td>')
        body += (f"<tr><td class='mono'>{esc(malavi)}</td>"
                 f"<td class='mono'>{esc(alt)}</td>"
                 f"<td class='mono'>{esc(accessions)}</td>{status}"
                 f"<td class='rowno'>{esc((entry.get('source') or {}).get('row'))}</td>"
                 "</tr>")

    return ('<h2 class="display">Names for lineages already in MalAvi</h2>'
            '<p class="sub">The submitter says each lineage on the left, which MalAvi '
            'already holds, was published under the name on the right. Accepting these '
            'makes the two names one lineage for every future reader, so they are worth '
            'more attention than a row count suggests — and they are not proposals for '
            'new lineages, which are listed separately above.</p>'
            '<div class="tablewrap"><table><thead><tr>'
            '<th>MalAvi name</th><th>Published as</th><th>Accessions</th>'
            '<th>Status</th><th>Row</th></tr></thead>'
            f'<tbody>{body}</tbody></table></div>')


def _submitted_files(metadata: Optional[Dict[str, Any]]) -> str:
    """Links to what the submitter actually uploaded.

    A curator reading a rendering of a workbook will sooner or later want the workbook —
    to sort it, to search it, to open a supplementary file the appendix cannot show at
    all. The Drive ids are already in the fetched form response, so this costs nothing but
    reading them out.

    Nothing is linked for a submission that was not fetched from the form, which is the
    honest outcome: a hand-built submission has no Drive file to point at.
    """
    if not metadata:
        return ""

    links = []
    for question, answer in metadata.items():
        if question.startswith("_") or "drive.google.com" not in str(answer):
            continue
        # The question text names the file far better than the id does.
        label = str(question).split("(")[0].strip() or "Uploaded file"
        for token in str(answer).replace(",", " ").split():
            url = _safe_url(token)
            if url:
                links.append(f'<li><a href="{esc(url)}">{esc(label)}</a></li>')

    if not links:
        return ""
    return ('<h3 class="appendix-head">The uploaded files</h3>'
            '<p class="sub">Opens in Drive. You will need to be signed in with the '
            'account the folder was shared with.</p>'
            f'<ul class="files">{"".join(links)}</ul>')


def _matrix_section(submission: Dict[str, Any]) -> str:
    """A host x parasite matrix per locality.

    The flat table below this one lists every row, which is the right shape for checking a
    single record and the wrong shape for the question a curator actually asks: which
    parasites were found in which hosts, here. A matrix answers that in one look, and makes
    the two things worth noticing visible without reading -- a host with nothing in it, and
    a lineage found in only one host at one site.

    Prevalence is shown as found/tested where the study reported it, because 1/60 and 40/60
    are different claims and a tick would flatten them.
    """
    records = submission.get("records") or []
    if not records:
        return ""

    # Grouped by locality rather than by country alone: two sites in one country are two
    # samples, and merging them would invent a pattern neither of them shows.
    places: Dict[str, List[Dict[str, Any]]] = {}
    for record in records:
        where = " · ".join(filter(None, [str(record.get("country") or "").strip(),
                                         str(record.get("site") or "").strip()]))
        places.setdefault(where or "Locality not given", []).append(record)

    out = ['<h2 class="display">What was found where</h2>']

    for place in sorted(places):
        rows = places[place]
        hosts = sorted({str(r.get("host_species") or "") for r in rows} - {""})
        lineages = sorted({str(r.get("lineage_name") or "") for r in rows} - {""})
        if not hosts or not lineages:
            continue

        cells: Dict[tuple, str] = {}
        for record in rows:
            host = str(record.get("host_species") or "")
            lineage = str(record.get("lineage_name") or "")
            found, tested = record.get("number_found"), record.get("number_tested")
            if found is None and tested is None:
                value = "&check;"
            else:
                value = f"{_number(found) if found is not None else '?'}/" \
                        f"{_number(tested) if tested is not None else '?'}"
            cells[(host, lineage)] = value

        out.append(f'<h3 class="place">{esc(place)}</h3>')

        # A page fits about this many lineage columns beside the host names. A study of
        # 39 lineages ran off the right edge of the paper, silently -- the HTML scrolls,
        # the PDF just loses the columns. So the matrix is split into blocks that fit,
        # each repeating the host rows, rather than being shrunk until it is unreadable.
        per_block = 10
        blocks = [lineages[i:i + per_block] for i in range(0, len(lineages), per_block)]
        for number, block in enumerate(blocks, start=1):
            header = "".join(f"<th class='mono'>{esc(l)}</th>" for l in block)
            body = ""
            for host in hosts:
                tds = ""
                for lineage in block:
                    value = cells.get((host, lineage))
                    tds += (f"<td class='hit'>{value}</td>" if value
                            else "<td class='miss'>&middot;</td>")
                body += f"<tr><th class='rowhead'><i>{esc(host)}</i></th>{tds}</tr>"
            if len(blocks) > 1:
                out.append(f'<div class="note">Lineages {esc(block[0])}–'
                           f'{esc(block[-1])} — part {number} of {len(blocks)}; the same '
                           f'hosts are listed in each part.</div>')
            out.append(f'<div class="tablewrap"><table class="matrix">'
                       f'<tr><th></th>{header}</tr>{body}</table></div>')

    return "\n".join(out)


def _records_section(submission: Dict[str, Any]) -> str:
    records = submission.get("records") or []
    vectors = submission.get("vectors") or []
    if not records and not vectors:
        return ""

    # The handle a curator quotes to correct one of these. Generated by the same function
    # the correction path resolves with, so the label on the page and the label the form
    # accepts cannot drift apart.
    from .record_handles import handles as _handles      # noqa: PLC0415 - avoids a cycle
    handle_for = {(entry.kind, entry.index): entry.handle for entry in _handles(submission)}

    # The tier vocabulary is defined here rather than left to be inferred. The words are
    # taken from row_flags.TIERS, which is what assigns them, so the definition on the
    # page cannot drift from the rule that produced it. Only the tiers actually present
    # are defined: explaining a tier no row is in is noise in a document already long.
    from .row_flags import TIERS                      # noqa: PLC0415 - avoids a cycle
    present = [t for t in TIERS if any(
        (r.get("tier") == t) for r in list(records) + list(vectors))]
    legend = ""
    if present:
        items = "".join(f"<dt>{esc(t)}</dt><dd>{esc(TIERS[t])}</dd>" for t in present)
        legend = ('<p class="sub">The <b>Tier</b> column is this row\'s triage, in the '
                  'order worth working through:</p>'
                  f'<dl class="tierkey">{items}</dl>')

    out = ['<h2 class="display">Records</h2>',
           '<p class="sub">The <b>Ref</b> column is how you name a record on the verdict '
           'form. To correct one, quote its Ref — for example R3 — and give the field and '
           'the corrected value.</p>',
           legend]
    if records:
        head = ("<tr><th>Ref</th><th>Row</th><th>Lineage</th><th>Host</th>"
                "<th>Country</th><th>Site</th><th>Found</th><th>Tested</th><th>Tier</th>"
                "<th>Notes</th></tr>")
        rows = "".join(
            "<tr>"
            f"<td class='handle'>{esc(handle_for.get(('records', i), ''))}</td>"
            f"<td class='rowno'>{esc((r.get('source') or {}).get('row'))}</td>"
            f"<td class='mono'>{esc(r.get('lineage_name'))}</td>"
            f"<td>{esc(r.get('host_species'))}</td>"
            f"<td>{esc(r.get('country'))}</td>"
            f"<td>{esc(r.get('site'))}</td>"
            f"<td>{esc(_number(r.get('number_found')))}</td>"
            f"<td>{esc(_number(r.get('number_tested')))}</td>"
            f"<td>{esc(r.get('tier'))}{_row_flags(r)}</td>"
            f"<td class='wrapcell'>{esc(r.get('notes'))}</td>"
            "</tr>" for i, r in enumerate(records))
        out.append(f'<div class="tablewrap"><table><thead>{head}</thead>'
                   f'<tbody>{rows}</tbody></table></div>')
    if vectors:
        out.append("<h3>Vectors</h3>")
        head = ("<tr><th>Ref</th><th>Row</th><th>Lineage</th><th>Vector</th>"
                "<th>Method</th><th>Country</th><th>Tier</th><th>Notes</th></tr>")
        rows = "".join(
            "<tr>"
            f"<td class='handle'>{esc(handle_for.get(('vectors', i), ''))}</td>"
            f"<td class='rowno'>{esc((v.get('source') or {}).get('row'))}</td>"
            f"<td class='mono'>{esc(v.get('lineage_name'))}</td>"
            f"<td>{esc(v.get('vector_species'))}</td>"
            f"<td>{esc(v.get('vector_method'))}</td>"
            f"<td>{esc(v.get('country'))}</td>"
            f"<td>{esc(v.get('tier'))}{_row_flags(v)}</td>"
            f"<td class='wrapcell'>{esc(v.get('notes'))}</td>"
            "</tr>" for i, v in enumerate(vectors))
        out.append(f'<div class="tablewrap"><table><thead>{head}</thead>'
                   f'<tbody>{rows}</tbody></table></div>')
    return "\n".join(out)


def _row_flags(row: Dict[str, Any]) -> str:
    """The flag codes behind a row's tier, named under it.

    The tier vocabulary says "review" means at least one flag needs a curator's
    judgment, and until 2026-08-19 the table never said which flag. A reader was told a
    judgment was required and not what about. The codes are printed as row_flags writes
    them, because those are the words the rest of the report and the check list use.
    """
    flags = row.get("flags") or []
    codes = [str(f.get("code")) for f in flags if isinstance(f, dict) and f.get("code")]
    if not codes:
        return ""
    return ('<br><span class="rowflags">'
            + esc(", ".join(dict.fromkeys(codes)).replace("_", " "))
            + "</span>")


def _number(value: Any) -> Any:
    """Counts are carried as floats; a curator should not read '3.0 birds'."""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _workbook_section(workbook_path: Optional[Path], max_rows: int = 300) -> str:
    """Every sheet of the submitted workbook, as submitted.

    Read at render time rather than carried through the submission: the file is already
    in the submission directory, and rendering from it means the curator sees the
    workbook rather than our interpretation of it.
    """
    if not workbook_path or not Path(workbook_path).is_file():
        return ""
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    except Exception as exc:                                  # noqa: BLE001
        return ('<h2 class="display">The submitted workbook</h2>'
                f'<div class="banner">Could not be rendered: {esc(exc)}. '
                "The file itself is beside this report.</div>")

    from .template_adapter import _is_example_row

    out = ['<h2 class="display">The submitted workbook</h2>',
           '<div class="note">The submitted template tables as they arrived. Two things '
           'are left out: the READ ME sheet, which is the same instructions in every '
           'submission, and the template\'s own worked example rows where the submitter '
           'left them in place — the template invites them to, so their presence says '
           'nothing about the submission. Any row they typed over is their data and is '
           'shown.</div>']
    dropped_examples = 0
    try:
        for name in workbook.sheetnames:
            # The READ ME is instructions shipped with the template, identical in every
            # submission, and it costs a page of a report a curator has to read.
            if name.strip().casefold() == "read me":
                continue
            worksheet = workbook[name]
            rows: List[str] = []
            for index, values in enumerate(
                    worksheet.iter_rows(values_only=True), start=1):
                if not values or all(v in (None, "") for v in values):
                    continue
                if _is_example_row(name, values):
                    dropped_examples += 1
                    continue
                if len(rows) >= max_rows:
                    rows.append(f'<tr><td colspan="99" class="rowno">…more rows in the '
                                f'workbook itself</td></tr>')
                    break
                cells = "".join(f"<td>{esc(v)}</td>" for v in values)
                rows.append(f'<tr><td class="rowno">{index}</td>{cells}</tr>')
            if not rows:
                continue
            out.append(f"<h3>{esc(name)}</h3>")
            out.append(f'<div class="tablewrap"><table>{"".join(rows)}</table></div>')
    finally:
        workbook.close()
    if dropped_examples:
        out.append(f'<div class="note">{dropped_examples} unmodified worked-example '
                   f'row(s) from the blank template were left out of the tables above. '
                   f'The screen ignores them too.</div>')
    return "\n".join(out)


# ---------------------------------------------------------------------------------------
# Page
# ---------------------------------------------------------------------------------------

def render_paper_only_report(
    metadata: Optional[Dict[str, Any]] = None,
    submission_id: Optional[str] = None,
    submit_form_url: str = "",
) -> str:
    """The report for a submission that arrived as a paper with no data template.

    A legitimate way to submit — plenty of useful papers will arrive this way, and the form
    invites it — but nothing about it can be screened. There is no template to read, so no
    lineage names to check for collisions, no sequences to compare, no records to place.

    The failure this replaces was worse than useless: the screen exited with an error and
    wrote nothing, so a curator received an email about a submission and then found nothing
    to review and no explanation. This says plainly that nothing has been checked, links
    the paper, and names the one action that moves it forward — a curator reads the paper,
    fills in the template, and submits that as a new submission, which then goes through
    every check in the ordinary way.
    """
    submitter = (metadata or {}).get("_submitter") or {}
    files = _submitted_files(metadata)

    rows = [
        ("Submitted by", esc((metadata or {}).get("What is your first and last name?"))),
        ("Institution", esc((metadata or {}).get("What institution are you associated with?"))),
        ("Email", esc((metadata or {}).get("Email Address"))),
        ("Received", esc((metadata or {}).get("_fetched_at"))),
        ("Submission", esc(submission_id)),
    ]
    notes = (metadata or {}).get(
        "Please provide any relevant notes or communication here (if applicable).")
    if notes:
        rows.append(("Their notes", esc(notes)))
    cells = "".join(f"<dt>{label}</dt><dd>{value or '—'}</dd>" for label, value in rows)

    action = ""
    if submit_form_url:
        action = (f'<p><a class="verdict-link" href="{esc(submit_form_url)}">'
                  f'Submit the data from this paper</a></p>')

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta http-equiv="Content-Security-Policy"
      content="default-src 'none'; style-src 'unsafe-inline'; img-src data:;">
<title>MalAvi review — a paper with no data template</title>
<style>{_stylesheet()}</style>
</head>
<body><div class="wrap">
<h1 class="display">MalAvi submission review</h1>
<dl class="meta">{cells}</dl>

<div class="banner"><b>Nothing here has been checked.</b> This submission arrived as a
paper without a filled-in data template, so there was nothing for the automatic checks to
read: no proposed names to test for collisions, no sequences to compare against the
release, no records to place. Treat everything below as unexamined.</div>

<h2 class="display">This one is yours to complete</h2>
<p>A paper on its own is a perfectly good way to send data to MalAvi, and this is how it
gets in: a curator reads the paper, fills in the data template, and submits that as a new
submission. It then goes through every check in the ordinary way and comes back as a
normal report.</p>
<p>There is nothing to approve or reject here. Until the data is entered, this is an open
item rather than a decision.</p>
{action}

{files or '<p class="sub">No uploaded file was recorded with this submission.</p>'}

<footer>
Nothing in this submission has been added to MalAvi, and nothing in it has been checked.
</footer>
</div></body>
</html>
"""


def render_report(
    submission: Dict[str, Any],
    run: CheckRun,
    screen: Optional[Any] = None,
    workbook_path: Optional[Path] = None,
    metadata: Optional[Dict[str, Any]] = None,
    alignments: Optional[Sequence[Any]] = None,
    submission_id: Optional[str] = None,
    revision: Optional[int] = None,
    entry: Optional[Any] = None,
) -> str:
    """Return the complete HTML document for one submission.

    ``revision`` must be supplied whenever ``submission_id`` is: the verdict link carries
    it, and a curator's approval is only counted against the revision it names. Defaulting
    it to 1 meant that after any resubmission the link recorded an approval against a
    version nobody was looking at, which the ledger accepts and then never treats as
    standing -- so the submission stalls silently while the curator believes they approved
    it. Holds carry forward across revisions, so the failure was one-sided and invisible.

    ``entry`` is this submission's review-ledger entry, when there is one. It supplies the
    flags and corrections already standing, and with them the ``V1``/``C1`` ids the verdict
    form asks a curator to type -- see :func:`_standing_section`. Optional, because a report
    can legitimately be generated outside a submission tree; the section is then omitted
    rather than the report failing.
    """
    if submission_id and revision is None:
        raise ValueError(
            "render_report needs the revision when it has a submission id: the verdict "
            "link binds an approval to the revision it names, and guessing 1 records "
            "approvals against a version the curator never saw.")
    revision = 1 if revision is None else int(revision)
    reference = submission.get("reference") or {}
    title = reference.get("title") or "MalAvi submission"

    # Order matters more than any single section here. A curator asks, in this order:
    # what is this, what is wrong with it, what exactly was sent, and only then how the
    # machinery reached its opinion. The previous order answered the last question first
    # and the third one on page six.
    sections = "\n".join(filter(None, [
        _header(submission, run, metadata),
        _disposition(run),
        _summary_section(submission, screen),
        _verdict_section(submission_id, revision),
        # Directly after the verdict block, because it is what the curator needs while they
        # are filling that form in -- not six sections later, past the appendix.
        _standing_section(entry),
        _checks_section(run, (submission.get('provenance') or {}).get('workbook'),
                        submission=submission, screen=screen),
        _names_section(submission),
        _alt_names_section(submission),
        _sequences_section(screen),
        _alignment_section(alignments),
        _matrix_section(submission),
        _records_section(submission),
        _normalization_section(submission),
        '<div class="appendix">' + _submitted_files(metadata)
        + (_workbook_section(workbook_path) or '') + '</div>',
        _provenance_section(run),
    ]))

    # A restrictive policy even though this is a local file: the values on this page are
    # submitter-controlled, and defense in depth costs one line.
    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta http-equiv="Content-Security-Policy"
      content="default-src 'none'; style-src 'unsafe-inline'; img-src data:;">
<title>MalAvi review — {esc(title)}</title>
<style>{_stylesheet()}</style>
</head>
<body><div class="wrap">
{sections}
</div></body>
</html>
"""


def write_pdf(content: str, destination: Path, intake_root: Path) -> Optional[Path]:
    """Render the report to PDF beside the HTML, or return ``None`` if it cannot.

    The PDF is the copy a curator actually opens: it is delivered through the Drive
    folder the submission arrived in, and Drive renders a PDF in the browser while it
    makes any HTML file a download. One click versus a download is the difference
    between a workflow a biologist uses and one they stop using.

    **Optional by design.** WeasyPrint pulls in system libraries that will not be
    present everywhere this runs, and a missing renderer must not cost the curator their
    report. When it is absent this returns ``None``, the HTML is still written, and the
    caller says the PDF is unavailable — which is a visible degradation rather than a
    failure, and is exactly how the check suite treats a check it could not run.
    """
    try:
        from weasyprint import HTML          # noqa: PLC0415 - optional dependency
    except ImportError:
        return None

    destination = _checked_destination(destination, intake_root)
    destination.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary = tempfile.mkstemp(dir=str(destination.parent), suffix=".tmp")
    os.close(handle)
    try:
        # base_url lets relative references resolve, though a self-contained report has
        # none; it costs nothing and stops a future asset silently failing to load.
        HTML(string=content, base_url=str(destination.parent)).write_pdf(temporary)
        os.chmod(temporary, 0o600)
        os.replace(temporary, destination)
    except BaseException:
        Path(temporary).unlink(missing_ok=True)
        raise
    return destination


def _checked_destination(destination: Path, intake_root: Path) -> Path:
    """Resolve a destination and refuse anything outside the intake tree.

    Shared by the HTML and PDF writers so the two cannot drift into enforcing different
    rules about where a report may be written.
    """
    destination = Path(destination).resolve()
    root = Path(intake_root).resolve()

    if not destination.is_relative_to(root):
        raise ValueError(
            f"refusing to write a curator report outside the intake tree: "
            f"{destination} is not inside {root}")
    if destination.is_symlink() or any(p.is_symlink() for p in destination.parents
                                       if p.is_relative_to(root)):
        raise ValueError(
            f"refusing to write a curator report through a symlink: {destination}")
    if destination.exists() and not destination.is_file():
        raise ValueError(f"refusing to overwrite a non-file: {destination}")
    return destination


def write_report(content: str, destination: Path, intake_root: Path) -> Path:
    """Write the report, refusing to put it anywhere but inside the intake tree.

    This is enforced here, in the code that does the writing, rather than only in a test.
    The document holds unpublished sequences and a submitter's email address; a later
    caller passing a path outside the gitignored intake tree would put both somewhere
    they could be committed or published, and a test cannot prevent that.
    """
    destination = _checked_destination(destination, intake_root)

    # Atomic, so a reader never opens a half-written report, and owner-only, because the
    # contents are somebody else's unpublished data.
    destination.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary = tempfile.mkstemp(dir=str(destination.parent), suffix=".tmp")
    try:
        with os.fdopen(handle, "w", encoding="utf-8") as fh:
            fh.write(content)
        os.chmod(temporary, 0o600)
        os.replace(temporary, destination)
    except BaseException:
        Path(temporary).unlink(missing_ok=True)
        raise
    return destination
