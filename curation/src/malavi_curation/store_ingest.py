"""Turn an approved submission's workbook into rows for the record store.

**The gap this closes.** ``write_store`` had exactly one caller, ``release_seed``, so the
store could be created from the last externally-produced release and never added to. A
curator could approve a submission and nothing would carry its records into MalAvi. The
release gate refuses records nobody approved; this is the path by which approved records
arrive at all.

**Why it reads the workbook rather than submission.json.** ``submission.json`` keeps nine
fields per record, and the template's ``Hosts_and_Sites`` sheet has fourteen columns plus
the ``Sites`` join. ``HostAge``, ``HostStatus``, ``HostEnvironment``, ``HostSubspecies``
and ``CountryRegion`` have nowhere to go in that schema -- ``_row_notes`` sweeps them into
free prose, which is not a column source. Ingesting through it would silently drop five
columns a submitter had taken the trouble to fill in, and the loss would be invisible in
the release. So the workbook is read directly, with :mod:`template_adapter`'s own reader
rather than a second parser, so the screen and the ingest cannot disagree about what a
cell says.

Submissions that arrived as a PDF with no workbook cannot use this path. That is correct
rather than a limitation: the curator instructions already tell a curator to extract such
a paper into a template and submit it like any other, and at that point it has a workbook.

**Nothing here judges anything.** Values are mapped, joined and derived; every derivation
is a lookup this project can point at -- the workbook's own ``Sites`` sheet, MalAvi's
existing records, ``country_regions.csv``. A host species MalAvi has never seen gets no
order and family rather than a guess, and a curator fills them in. Inventing taxonomy
here would put an unsourced value into a release under a submitter's name.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from . import normalize, reference_names, sequence_check, template_adapter
from .release_store import TABLES, TableSpec, record_id, row_key

# The columns the template supplies directly, as {store column: template column}. Kept as
# data rather than as a run of assignments so the mapping can be read, and tested, as one
# object -- this table IS the contract between the submission template and the store.
HOST_COLUMN_MAP = {
    "LINEAGE_NAME": "LINEAGE_NAME",
    "SUB_SPECIES_NAME": "HostSubspecies",
    "HOST_AGE": "HostAge",
    "HOST_STATUS": "HostStatus",
    "HOST_ENVIRONMENT": "HostEnvironment",
    "COUNTRY_NAME": "Country",
    "COUNTRY_REGION_NAME": "CountryRegion",
    "SITE_NAME": "SiteName",
    "NUMBER_FOUND": "NUMBER_FOUND",
    "NUMBER_TESTED": "NUMBER_TESTED",
    "REFERENCE_NAME": "Reference",
    "COMMENT": "COMMENT",
}


def continent_index(host_rows: Sequence[Dict[str, Any]]) -> Dict[str, str]:
    """COUNTRY_NAME -> CONTINENT_NAME, from the store's own records.

    **Not from country_regions.csv.** That table answers a different question in a
    different vocabulary: ``region_for("Sweden")`` is ``EUROPE``, a region code, while
    ``CONTINENT_NAME`` in the release holds ``Europe``. Writing one into the other would
    corrupt the column the Grand Lineage Summary's twelve region columns are derived from,
    and it would look plausible in the CSV. The region columns are computed at build time
    from ``country_regions.csv`` and are not this function's business.

    A country MalAvi has never recorded gets no continent, and a curator supplies it.
    """
    index: Dict[str, str] = {}
    for row in host_rows:
        country = (row.get("COUNTRY_NAME") or "").strip()
        continent = (row.get("CONTINENT_NAME") or "").strip()
        if country and continent and continent != "Unknown":
            index.setdefault(country, continent)
    return index


def lineage_cell(header: Any, values: Any, column: str) -> str:
    """One lineage name from the workbook, in MalAvi's own casing.

    **Every** lineage name this module reads goes through here, because the screen and
    the ledger already read them this way (``check_template._lineage_cell`` and
    ``enrollment`` both apply ``normalize.lineage_name``) and three readers of one cell
    that disagree is how a name gets past a check that was looking for it.

    Until 2026-08-11 the ingest alone read them raw, which made the collision refusal
    bypassable by a typo: MalAvi holds ``TUMIG19``, a submitter types ``tumig19``, the
    screen normalizes and raises "that name is taken" and offers an alternative, and then
    ``apply_name_corrections`` looks up ``tumig19``, misses, renames nothing, and
    ``colliding_lineages`` compares ``tumig19`` against a store holding ``TUMIG19`` and
    finds no collision. The store gains a second lineage differing only in case, carrying
    a different sequence and accession, and neither the duplicate-name check nor the
    duplicate tip-label check in ``release_build`` fires -- both compare exact strings.

    ``normalize.lineage_name`` also strips internal whitespace, so ``SGS 1`` cannot reach
    ``fasta_label`` and produce an alignment id that every FASTA reader truncates at the
    space.
    """
    return normalize.lineage_name(template_adapter.cell(header, values, column)) or ""


def split_host_species(host_species: Optional[str]) -> Tuple[str, str]:
    """"Accipiter tachiro" -> ("Accipiter", "Accipiter tachiro").

    ``GENUS_NAME`` is the genus and **``SPECIES_NAME`` is the whole binomial**, because
    that is what MalAvi holds: 18,473 of the 18,493 seeded host records have a
    ``SPECIES_NAME`` beginning with their ``GENUS_NAME`` (``Pipile`` / ``Pipile
    jacutinga``). This function returned the bare epithet until 2026-08-11, against a
    docstring asserting the store "keeps genus and species in separate columns" -- which
    the seeded data plainly does not.

    Nothing caught it, and four separate things hid it:

    * ``taxonomy_index`` is keyed on the binomial, so the ``(genus, species)`` lookup
      always missed and fell through to the genus-only entry -- order and family still
      resolved, so no note was raised;
    * ``derive_summary`` counts distinct ``(GENUS_NAME, SPECIES_NAME)`` pairs, so the
      same bird arriving by submission and by seed counted as **two** host species and
      inflated ``SUM_HOST``;
    * the published ``Hosts_and_Sites`` table would have carried a column mixing
      binomials and epithets, and the site lists that column's distinct values as host
      species -- ``migratorius`` would have appeared as a bird;
    * the edition report could not show any of it, because ``release_diff._host_binomial``
      reconstructs the same binomial from either form.

    Anything beyond the first two words is still left alone -- it belongs in
    ``HostSubspecies``, which has its own column, and silently absorbing it here would
    put a subspecies into the species field where no check would see it.
    """
    parts = (host_species or "").strip().split()
    if not parts:
        return ("", "")
    if len(parts) < 2:
        # A genus with no epithet. MalAvi holds these (e.g. "Sphenisciformes spp"), and
        # the binomial column stays empty rather than repeating the genus, so that
        # "we know the genus only" is distinguishable from "the species is the genus".
        return (parts[0], "")
    return (parts[0], f"{parts[0]} {parts[1]}")


def taxonomy_index(host_rows: Sequence[Dict[str, Any]]) -> Dict[Tuple[str, str], Tuple[str, str]]:
    """MalAvi's own answer to "what order and family is this bird in?".

    Built from the store's existing records rather than from an external taxonomy,
    because the store is what the release has to stay consistent with: taking a family
    from elsewhere would put two spellings of the same clade in one table and make the
    Grand Lineage Summary's family counts disagree with themselves.

    Only rows that actually carry both values contribute, and the first spelling wins so
    the result does not depend on row order beyond that.
    """
    index: Dict[Tuple[str, str], Tuple[str, str]] = {}
    for row in host_rows:
        genus = (row.get("GENUS_NAME") or "").strip()
        species = (row.get("SPECIES_NAME") or "").strip()
        order = (row.get("ORDER_NAME") or "").strip()
        family = (row.get("FAMILY_NAME") or "").strip()
        if not genus or not order or not family:
            continue
        index.setdefault((genus, species), (order, family))
        # Genus alone, so a species new to MalAvi in a genus MalAvi knows still places.
        # A genus does not span two families in this data, and where it did the first
        # spelling would win, which is the same rule as above.
        index.setdefault((genus, ""), (order, family))
    return index


def parasite_genus_index(host_rows: Sequence[Dict[str, Any]]) -> Dict[str, str]:
    """LINEAGE_NAME -> PARASITE_GENUS, from the store's own host records.

    **Why a host record carries the genus at all.** It is a projection of the lineage's
    genus, and an exceptionless one: all 18,493 host records in the seed carry a
    PARASITE_GENUS, every one agrees with ``lineages.GENUS_NAME``, and no lineage appears
    under two spellings. A record for a lineage MalAvi already knows must therefore take
    the genus MalAvi already holds -- reading it off the submitter's spelling instead is
    how one lineage acquires two genera, and leaving it blank (which this module did until
    2026-08-10) would put the first empty value into a column that has never had one.

    Read from ``host_records`` rather than from ``lineages`` because that is the table
    whose column is being filled, so the two cannot drift apart, and because it is what
    :func:`_host_rows` is already given.
    """
    index: Dict[str, str] = {}
    for row in host_rows:
        lineage = (row.get("LINEAGE_NAME") or "").strip()
        genus = (row.get("PARASITE_GENUS") or "").strip()
        if lineage and genus:
            index.setdefault(lineage, genus)
    return index


def coordinates_by_site(workbook) -> Dict[str, str]:
    """SITE_NAME -> "lat, long" from the Sites sheet.

    The template tells submitters the site names on the two sheets must match, so this is
    reading the workbook's own join rather than inferring one. A site with no usable
    pair contributes nothing and the record simply has no coordinates.
    """
    if template_adapter.SHEET_SITES not in workbook.sheetnames:
        return {}
    header, body = template_adapter.sheet_rows(
        workbook[template_adapter.SHEET_SITES], "SITE_NAME")
    found: Dict[str, str] = {}
    for _row_number, values in body:
        name = (template_adapter.cell(header, values, "SITE_NAME") or "").strip()
        latitude = (template_adapter.cell(header, values, "LATITUDE") or "").strip()
        longitude = (template_adapter.cell(header, values, "LONGITUDE") or "").strip()
        if name and latitude and longitude:
            found.setdefault(name, f"{latitude}, {longitude}")
    return found


# How a host record lists several alternative names. Measured against the seed: all 553
# multi-valued ALT_NAME cells in the 2026-03-23 release use a bare comma with no space,
# and none has surrounding whitespace.
ALT_NAME_SEPARATOR = ","


def alt_names_by_record(workbook) -> Dict[Tuple[str, str], List[str]]:
    """(LINEAGE_NAME, REFERENCE_NAME) -> the names that study used, in the sheet's order.

    **Why a host record needs this at all when there is an ``alt_names`` table.**
    ``host_records.ALT_NAME`` is a denormalized projection of that table, not a second
    unrelated field, and the seed is exactly consistent with that reading: all 5,487 host
    records carrying an ALT_NAME reconcile to the alt_names rows for the same lineage and
    reference, with no exceptions and none missing. Populating only the table -- which is
    what this module did until 2026-08-10 -- would have made the first ingested record the
    first in MalAvi to break a consistency the whole seed has, in a column nothing checks.

    **The order is the sheet's, not sorted.** 49 of the 553 multi-valued cells in the seed
    are not in alphabetical order, so the original is a source order rather than a
    canonical one. It cannot be recovered from the store, whose rows are sorted on write,
    but for a submission it is simply the order the submitter listed them in, which is the
    same thing the seed preserved.
    """
    if template_adapter.SHEET_ALT_NAMES not in workbook.sheetnames:
        return {}
    header, body = template_adapter.sheet_rows(
        workbook[template_adapter.SHEET_ALT_NAMES], "MalAvi_Name")
    found: Dict[Tuple[str, str], List[str]] = {}
    for _row_number, values in body:
        lineage = lineage_cell(header, values, "MalAvi_Name")
        alternative = (template_adapter.cell(
            header, values, "Alternative_Name") or "").strip()
        reference = (template_adapter.cell(header, values, "Reference") or "").strip()
        if not lineage or not alternative:
            continue
        names = found.setdefault((lineage, reference), [])
        if alternative not in names:      # a name listed twice is still one name
            names.append(alternative)
    return found


def parasite_genus_by_lineage(workbook) -> Dict[str, str]:
    """LINEAGE_NAME -> ParasiteGenus, for lineages the submission declares as new.

    A row naming a lineage MalAvi already has gets nothing from this, which is right: the
    release already knows its genus, and restating it from a submitter's spelling is how
    one lineage ends up with two genera.
    """
    if template_adapter.SHEET_NEWLINEAGES not in workbook.sheetnames:
        return {}
    header, body = template_adapter.sheet_rows(
        workbook[template_adapter.SHEET_NEWLINEAGES], "LINEAGE_NAME")
    found: Dict[str, str] = {}
    for _row_number, values in body:
        name = lineage_cell(header, values, "LINEAGE_NAME")
        genus = (template_adapter.cell(header, values, "ParasiteGenus") or "").strip()
        if name and genus:
            found.setdefault(name, genus)
    return found


def host_rows_from_workbook(path: Path, submission_id: str, release: str,
                            existing_host_rows: Sequence[Dict[str, Any]] = (),
                            ) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Map ``Hosts_and_Sites`` into store rows. Returns ``(rows, notes)``.

    ``notes`` are things a person should see -- a host MalAvi cannot place, a site with no
    coordinates -- reported rather than raised, because none of them makes the row wrong
    and a curator can fill the gap once the record is in.

    Provenance is stamped here: ``_source`` is the submission id, which is what the
    release gate checks against the review ledger, and ``_added`` the release the row
    first appears in. ``RECORD_ID`` is deliberately absent -- ``assign_ids`` mints it at
    write time and never reissues an existing id.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=True)
    return _host_rows(workbook, submission_id, release, existing_host_rows)


def _host_rows(workbook, submission_id: str, release: str,
               existing_host_rows: Sequence[Dict[str, Any]] = (),
               ) -> Tuple[List[Dict[str, Any]], List[str]]:
    """The body of :func:`host_rows_from_workbook`, on an already-open workbook.

    Split out so ``tables_from_workbook`` opens the file once for all five tables rather
    than five times, which also guarantees every table is read from the same bytes.
    """
    if template_adapter.SHEET_HOSTS not in workbook.sheetnames:
        return [], [f"no {template_adapter.SHEET_HOSTS} sheet"]

    coordinates = coordinates_by_site(workbook)
    genera = parasite_genus_by_lineage(workbook)
    alternatives = alt_names_by_record(workbook)
    taxonomy = taxonomy_index(existing_host_rows)
    continents = continent_index(existing_host_rows)
    known_genera = parasite_genus_index(existing_host_rows)

    header, body = template_adapter.sheet_rows(
        workbook[template_adapter.SHEET_HOSTS], "LINEAGE_NAME")
    columns = TABLES["host_records"].columns

    rows: List[Dict[str, Any]] = []
    notes: List[str] = []
    for row_number, values in body:
        row: Dict[str, Any] = {column: "" for column in columns}
        for store_column, template_column in HOST_COLUMN_MAP.items():
            if store_column in row:
                row[store_column] = template_adapter.cell(
                    header, values, template_column) or ""

        # LINEAGE_NAME arrives through the generic map above, so it has to be normalized
        # here rather than at a dedicated read. Everything downstream -- the taxonomy
        # join, the rename, the collision refusal -- keys on it, and a host record naming
        # `tumig19` would point at a lineage row named `TUMIG19` and match nothing.
        row["LINEAGE_NAME"] = normalize.lineage_name(row["LINEAGE_NAME"]) or ""

        genus, species = split_host_species(
            template_adapter.cell(header, values, "HostSpecies"))
        row["GENUS_NAME"] = genus
        row["SPECIES_NAME"] = species

        placed = taxonomy.get((genus, species)) or taxonomy.get((genus, ""))
        if placed:
            row["ORDER_NAME"], row["FAMILY_NAME"] = placed
        elif genus:
            # `species` is the binomial and already carries the genus, so it is printed
            # alone; naming both gave "MalAvi has no record of Turdus Turdus migratorius".
            notes.append(
                f"row {row_number}: MalAvi has no record of {species or genus}"
                + ", so its order and family are blank for a curator to fill in")

        site = row.get("SITE_NAME") or ""
        if site and site in coordinates:
            row["SITE_COORDINATES"] = coordinates[site]
        elif site:
            notes.append(f"row {row_number}: site {site!r} has no coordinates on the "
                         f"Sites sheet")

        lineage = row.get("LINEAGE_NAME") or ""
        # The submission's own declaration first -- for a lineage it is introducing, that
        # is the only source there is -- then MalAvi's existing answer for a lineage it
        # already holds. Never the reverse: a submitter restating the genus of a known
        # lineage must not be able to change it here.
        if lineage in genera:
            row["PARASITE_GENUS"] = genera[lineage]
        elif lineage in known_genera:
            row["PARASITE_GENUS"] = known_genera[lineage]
        elif lineage:
            notes.append(
                f"row {row_number}: {lineage} is neither a lineage MalAvi holds nor one "
                f"the NewLineages sheet declares, so its parasite genus is blank for a "
                f"curator to fill in")

        # The names this study used for this lineage, joined the way the release joins
        # them. Keyed on the reference as well as the lineage, so a record does not
        # inherit the synonym some other paper used for the same lineage.
        alternative = alternatives.get(
            (lineage.strip(), (row.get("REFERENCE_NAME") or "").strip()))
        if alternative:
            row["ALT_NAME"] = ALT_NAME_SEPARATOR.join(alternative)

        country = row.get("COUNTRY_NAME") or ""
        if country and country in continents:
            row["CONTINENT_NAME"] = continents[country]
        elif country:
            notes.append(f"row {row_number}: MalAvi has no record of {country!r}, so "
                         f"its continent is blank for a curator to fill in")

        row["_source"] = submission_id
        row["_added"] = release
        rows.append(row)

    return rows, notes


def replace_submission_rows(spec: TableSpec,
                            existing: Sequence[Dict[str, Any]],
                            incoming: Sequence[Dict[str, Any]],
                            submission_id: str,
                            ) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    """Put a submission's rows into a table, replacing whatever it contributed before.

    **Replace, not append.** Re-ingesting is what happens after a correction, and a
    correction is a statement that the earlier version was wrong. Appending would leave
    both versions in the store, both citing the same study, and the wrong one would be
    indistinguishable from the right one in the next release -- the natural-key check
    would report a duplicate and a curator would have to work out which row to believe.

    **Only this submission's rows move.** Rows from ``seed`` or from another submission
    are never touched, however similar they look: deciding that somebody else's record is
    superseded is a curator's judgment with a reason attached, not something an importer
    does on the way past.

    **RECORD_ID survives an unchanged row.** The store promises an id is assigned once and
    never changed, because a curator decision, a correction and a published release may
    all point at it. So a replacement row whose natural key matches one this submission
    already had inherits its id, and only genuinely new or genuinely changed records get
    fresh ones. Wholesale replacement would silently reissue every id in the study.

    Returns ``(rows, counts)`` where counts reports kept / replaced / added / removed for
    the caller to print. ``rows`` preserves the order of ``existing``, with this
    submission's rows in their original positions where they persist, so the CSV diff of
    a correction shows the correction rather than a reshuffle.

    **Duplicate natural keys are preserved, not merged.** The key
    ``(lineage, species, site, reference)`` genuinely cannot separate two samplings of one
    host at one site in one paper: 678 keys in the seed are duplicated across 1,624 rows,
    302 of them byte-identical. Keying this function on the row -- as it did until
    2026-08-10 -- kept the last of each group, so one row was lost per duplicate with the
    count still reporting it written, and the survivor was a chimera carrying the first
    row's RECORD_ID and the second row's data.
    """
    mine_by_key: Dict[tuple, List[Dict[str, Any]]] = {}
    for row in existing:
        if str(row.get("_source") or "").strip() == submission_id:
            mine_by_key.setdefault(row_key(spec, row), []).append(row)

    counts = {"kept": 0, "replaced": 0, "added": 0, "removed": 0}

    # Carry the previous id and first-seen release onto a row that is still the same
    # record. _added answers "since when has MalAvi held this?", and a correction to a
    # host name does not change the answer.
    #
    # Each incoming row claims the next unclaimed previous row sharing its key, so N
    # incoming and M existing rows of one key pair up in order: min(N, M) inherit an id
    # and the surplus on either side is an addition or a removal.
    unclaimed = {key: list(rows) for key, rows in mine_by_key.items()}
    prepared: List[Dict[str, Any]] = []
    for row in incoming:
        row = dict(row)
        queue = unclaimed.get(row_key(spec, row)) or []
        previous = queue.pop(0) if queue else None
        if previous is not None:
            if previous.get("RECORD_ID"):
                row["RECORD_ID"] = previous["RECORD_ID"]
            if previous.get("_added"):
                row["_added"] = previous["_added"]
            counts["replaced" if any(
                (row.get(column) or "") != (previous.get(column) or "")
                for column in spec.columns) else "kept"] += 1
        else:
            counts["added"] += 1
        prepared.append(row)

    counts["removed"] = sum(len(rows) for rows in unclaimed.values())

    # Rebuild in place: each of this submission's surviving rows sits where it was, rows
    # it no longer claims drop out, and anything genuinely new goes on the end. Matched by
    # RECORD_ID rather than by key, so two rows sharing a key land in their own positions
    # instead of both collapsing onto the first.
    by_id = {row["RECORD_ID"]: row for row in prepared if row.get("RECORD_ID")}
    out: List[Dict[str, Any]] = []
    for row in existing:
        if str(row.get("_source") or "").strip() != submission_id:
            out.append(row)
            continue
        replacement = by_id.get(row.get("RECORD_ID"))
        if replacement is not None:
            out.append(replacement)
        # else: this submission no longer claims the row, so it goes.
    out.extend(row for row in prepared if not row.get("RECORD_ID"))
    return out, counts


def template_workbooks(directory: Path) -> List[Path]:
    """Every filled ImportMalavi template in a submission directory, in a stable order.

    A submission directory routinely holds spreadsheets that are not templates -- the
    paper's own supplementary tables travel with it -- so the files are opened and asked,
    with :func:`template_adapter.looks_like_template`, rather than selected by name. That
    is the same test the screen applies, so the ingest cannot decide a file is a
    submission that the screen never checked.

    Sorted, so that a submission carrying two templates ingests in the same order every
    run and its record ids do not depend on the order the filesystem happened to list.
    """
    import openpyxl

    found: List[Path] = []
    for path in sorted(Path(directory).rglob("*.xlsx")):
        if path.name.startswith("~$"):        # a spreadsheet lock file, not a workbook
            continue
        try:
            workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception:                     # noqa: BLE001 - openpyxl raises broadly
            # Unreadable here is not a decision to make quietly, but it is also not this
            # function's to make: the caller reports the file it could not open. Returning
            # it as a template would fail again, less clearly, further along.
            continue
        if template_adapter.looks_like_template(workbook):
            found.append(path)
    return found


def reference_names_in_workbook(path: Path) -> List[str]:
    """Every study a filled template cites, read from the workbook alone.

    **Why this exists separately from the ingest.** The store's ``_source`` column answers
    "which submission brought this row", which is how ``publish_reference`` finds the
    submissions behind a study. That answer is unavailable for exactly the submissions that
    need it most: an embargoed submission is refused by ``release_gate.admissibility``, so
    ``ingest_submissions`` never writes its rows, so the store has no provenance to read and
    the study it belongs to is unknowable from the store. That was the deadlock — the
    embargo could be set and never lifted, because the only thing that lifted it looked in
    the one place the embargoed rows were guaranteed not to be.

    So the study name is read from the submitted workbook, which exists from the moment the
    submission arrives and does not depend on anything having been published.

    Deliberately much cheaper than :func:`tables_from_workbook`: no host-name resolution, no
    geography, no taxonomy index, no notes. It answers one question, and a program that only
    needs to know which study a submission is for should not have to run the whole mapping —
    which can legitimately report problems, and would turn "which study is this" into a
    failure for reasons that have nothing to do with the question.

    Unpublished names are included and are the normal case here. The Reference *sheet*
    deliberately carries no row for an unpublished study (see :func:`reference_rows`), so
    the name is read from the record sheets, where it is always present.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=True)

    # Every sheet that names the study, and the column it names it in. The record sheets
    # spell it "Reference"; the Reference sheet spells it "REFERENCE_NAME".
    sources = (
        (template_adapter.SHEET_HOSTS, "LINEAGE_NAME", "Reference"),
        (template_adapter.SHEET_VECTORS, "LINEAGE_NAME", "Reference"),
        (template_adapter.SHEET_ALT_NAMES, "MalAvi_Name", "Reference"),
        (template_adapter.SHEET_REFERENCE, "REFERENCE_NAME", "REFERENCE_NAME"),
    )

    found = set()
    for sheet_name, key_header, column in sources:
        header, body = _sheet(workbook, sheet_name, key_header)
        if not body:
            continue
        for _row_number, values in body:
            name = (template_adapter.cell(header, values, column) or "").strip()
            if name:
                found.add(name)
    return sorted(found)


def blanked_values(spec: TableSpec,
                   before: Sequence[Dict[str, Any]],
                   after: Sequence[Dict[str, Any]],
                   submission_id: str) -> List[Dict[str, str]]:
    """Values the store held that a re-ingest would empty. One entry per lost value.

    **Why this has to be looked for.** Several columns are deliberately left blank by the
    mapping above -- a host order and family MalAvi has never seen, a continent, a
    lineage's ``SEQ_LENGTH`` -- with a note asking a curator to fill them in. The curator
    fills them in *in the store*, because that is where the record now lives. Re-ingesting
    the same workbook then maps the same blanks over the top of their work and reports it
    as ``replaced``, which is indistinguishable in the count from a correction that
    genuinely changed a value. The curator's contribution disappears with nothing said.

    Rows are paired by ``RECORD_ID`` -- the identity :func:`replace_submission_rows`
    carries across -- rather than by natural key, so two rows sharing a key are compared
    against themselves and not against each other.

    Only this submission's rows are examined, and only non-empty becoming empty is
    reported: a value the workbook *changes* is a correction, which is the point of a
    re-ingest and not a loss.
    """
    after_by_id = {record_id(row): row for row in after if record_id(row)}
    lost: List[Dict[str, str]] = []
    for row in before:
        if str(row.get("_source") or "").strip() != submission_id:
            continue
        identifier = record_id(row)
        replacement = after_by_id.get(identifier)
        if replacement is None:
            # The row is gone entirely, which replace_submission_rows already counts as a
            # removal and reports on its own terms. Counting its columns here as well
            # would report one dropped row as twenty lost values.
            continue
        for column in spec.columns:
            was = (row.get(column) or "").strip()
            now = (replacement.get(column) or "").strip()
            if was and not now:
                lost.append({"table": spec.name, "record_id": identifier,
                             "column": column, "was": was})
    return lost


# Template columns with no column in the store table they belong to. Reported rather than
# dropped in silence: a submitter filled these in, and "MalAvi has nowhere to put this"
# is a fact about MalAvi's schema that somebody should be able to see and decide about.
# Every lineage in the store is exactly this long: the cytochrome b barcode window MalAvi
# is defined by. A submitted sequence of another length is not rejected here -- rejecting
# is a check's job, not an importer's -- but it is always reported, because a lineage of a
# different length would be the first in the database.
MALAVI_WINDOW = 479

# The only two values SEQ_LENGTH takes in the store. It reads like a number and is not one.
SEQ_LENGTH_VALUES = ("Full", "Partial")

UNHOUSED = {
    "Vectors": ("CountryRegion", "No_found", "No_tested", "Comment"),
    "Alt_Lineage_names": ("GenBankNr", "Comment"),
    "Reference": ("DOI",),
}


def _sheet(workbook, name: str, key_header: str):
    """Header and body of one sheet, or ``(None, [])`` when the sheet is absent."""
    if name not in workbook.sheetnames:
        return None, []
    return template_adapter.sheet_rows(workbook[name], key_header)


def _unhoused_notes(sheet_name: str, header: Optional[List[str]],
                    body: Sequence[Any]) -> List[str]:
    """Say which filled-in columns the store cannot hold."""
    if not header or not body:
        return []
    notes = []
    for column in UNHOUSED.get(sheet_name, ()):
        if any(template_adapter.cell(header, values, column) for _n, values in body):
            notes.append(f"{sheet_name}.{column} was filled in, but the store has no "
                         f"column for it; the value is not carried into MalAvi")
    return notes


def lineage_rows(workbook, submission_id: str, release: str
                 ) -> Tuple[List[Dict[str, Any]], List[str]]:
    """NewLineages + Sequences -> the ``lineages`` table.

    ``GENUS_NAME``/``SPECIES_NAME`` here are the *parasite* morphospecies, not the host.
    The template asks only for the genus, so the species is left blank: assigning a
    lineage to a morphospecies is a taxonomic judgment recorded in ``morpho_species``,
    with a reference behind it, and inferring one from a genus would be inventing it.

    **``SEQ_LENGTH`` is not a length.** Despite the name it is categorical -- ``Full``
    (4,828 lineages in the store) or ``Partial`` (540), with no numeric value anywhere.
    This function wrote ``str(len(sequence))`` until 2026-08-10, which would have put
    ``"451"`` into a column every downstream reader compares against ``"Full"``, silently
    stopping the match for that lineage in malaviR and on the site.

    Nor can it be derived: 572 lineages marked ``Full`` hold fewer than 470 unambiguous
    bases, so the label records a curator's judgment about whether the sequence covers the
    barcode rather than a threshold. It is left blank and reported, like the order and
    family of a host MalAvi has never seen.
    """
    header, body = _sheet(workbook, template_adapter.SHEET_NEWLINEAGES, "LINEAGE_NAME")
    if not body:
        return [], []

    seq_header, seq_body = _sheet(workbook, template_adapter.SHEET_SEQUENCES,
                                  "LINEAGE_NAME")
    sequences: Dict[str, str] = {}
    for _n, values in seq_body:
        name = lineage_cell(seq_header, values, "LINEAGE_NAME")
        # normalize.sequence_pair, not a local clean-up. This read `.replace(" ", "")`
        # until 2026-08-20, which removes spaces and nothing else -- and submitters paste
        # sequences wrapped across lines. NECMON01 arrived with six newlines in the cell,
        # so ingest would have written 485 characters *including literal newlines* into a
        # CSV column and then into a FASTA record, and reported the lineage as "485 bp".
        # The adapter had always cleaned this correctly; ingest rolled its own weaker copy.
        _submitted, cleaned = normalize.sequence_pair(
            template_adapter.cell(seq_header, values, "SEQUENCE"))
        if name and cleaned:
            sequences.setdefault(name, cleaned)

    columns = TABLES["lineages"].columns
    rows, notes = [], []
    for row_number, values in body:
        name = lineage_cell(header, values, "LINEAGE_NAME")
        if not name:
            continue
        sequence = sequences.get(name, "")
        if not sequence:
            notes.append(f"NewLineages row {row_number}: {name} has no sequence on the "
                         f"Sequences sheet")
        accessions = normalize.accession_list(
            template_adapter.cell(header, values, "GENBANK_NR"))
        if len(accessions) > 1:
            # Every one of the 5,368 lineages in the store carries a single accession.
            # The submitter is not wrong -- a lineage really can have several deposits --
            # but writing the first comma-separated value MalAvi has ever held is a
            # change to what the column means, and that is a curator's call.
            notes.append(
                f"NewLineages row {row_number}: {name} lists {len(accessions)} "
                f"accessions ({', '.join(accessions)}); every lineage in MalAvi has one, "
                f"so a curator should pick the representative")
        # A partial barcode is the common case (3,340 of 5,368 stored lineages), and
        # place_sequences pads it into the window on the way in. So this reports what will
        # happen rather than demanding a decision -- it used to say "this needs trimming or
        # a curator's decision", which was alarming and wrong for a perfectly good
        # forward-primer-only read. A sequence that genuinely does not fit is refused by
        # misframed_sequences instead, which is a refusal and not a note.
        if sequence and len(sequence) < MALAVI_WINDOW:
            notes.append(
                f"NewLineages row {row_number}: {name} is {len(sequence)} bp, shorter than "
                f"the {MALAVI_WINDOW} bp barcode window; it will be padded into the window "
                f"and stored like the {PARTIAL_LINEAGES_NOTE} lineages MalAvi already holds "
                f"that cover only part of it")
        elif sequence and len(sequence) > MALAVI_WINDOW:
            notes.append(
                f"NewLineages row {row_number}: {name} is {len(sequence)} bp, longer than "
                f"the {MALAVI_WINDOW} bp barcode window; a curator has to decide how it is "
                f"trimmed")
        row = {column: "" for column in columns}
        row["LINEAGE_NAME"] = name
        row["GENBANK_ACC"] = ", ".join(accessions)
        row["GENUS_NAME"] = template_adapter.cell(header, values, "ParasiteGenus") or ""
        row["SEQUENCE"] = sequence
        # Left for a curator -- see the note in this function's docstring.
        row["SEQ_LENGTH"] = ""
        if sequence:
            notes.append(
                f"NewLineages row {row_number}: {name} needs SEQ_LENGTH set to "
                f"{' or '.join(repr(v) for v in SEQ_LENGTH_VALUES)} by a curator; it is a "
                f"judgment, not a measurement, and is left blank here")
        row["_source"] = submission_id
        row["_added"] = release
        rows.append(row)
    return rows, notes


def vector_rows(workbook, submission_id: str, release: str
                ) -> Tuple[List[Dict[str, Any]], List[str]]:
    """The Vectors sheet -> the ``vector_records`` table."""
    header, body = _sheet(workbook, template_adapter.SHEET_VECTORS, "LINEAGE_NAME")
    if not body:
        return [], []
    columns = TABLES["vector_records"].columns
    rows = []
    for _row_number, values in body:
        row = {column: "" for column in columns}
        row["LINEAGE_NAME"] = lineage_cell(header, values, "LINEAGE_NAME")
        row["VECTOR_SPECIES"] = template_adapter.cell(header, values, "VectorSpecies") or ""
        row["VECTOR_METHOD"] = template_adapter.cell(header, values, "VECTOR_METHOD") or ""
        row["COUNTRY_NAME"] = template_adapter.cell(header, values, "Country") or ""
        row["SITE_NAME"] = template_adapter.cell(header, values, "SiteName") or ""
        row["REFERENCE_NAME"] = template_adapter.cell(header, values, "Reference") or ""
        row["_source"] = submission_id
        row["_added"] = release
        rows.append(row)
    return rows, _unhoused_notes("Vectors", header, body)


def alt_name_rows(workbook, submission_id: str, release: str
                  ) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Alt_Lineage_names -> the ``alt_names`` table.

    The sheet's own columns are ``MalAvi_Name`` and ``Alternative_Name``: the name MalAvi
    uses and the name the publication used. The store keeps them as ``LINEAGE_NAME`` and
    ``ALT_NAME`` in that order, and getting them the wrong way round would make MalAvi
    answer to the paper's name and offer its own as the synonym.
    """
    header, body = _sheet(workbook, template_adapter.SHEET_ALT_NAMES, "MalAvi_Name")
    if not body:
        return [], []
    columns = TABLES["alt_names"].columns
    rows = []
    for _row_number, values in body:
        row = {column: "" for column in columns}
        row["LINEAGE_NAME"] = lineage_cell(header, values, "MalAvi_Name")
        row["ALT_NAME"] = template_adapter.cell(header, values, "Alternative_Name") or ""
        row["REFERENCE_NAME"] = template_adapter.cell(header, values, "Reference") or ""
        row["_source"] = submission_id
        row["_added"] = release
        rows.append(row)
    return rows, _unhoused_notes("Alt_Lineage_names", header, body)


def reference_rows(workbook, submission_id: str, release: str
                   ) -> Tuple[List[Dict[str, Any]], List[str]]:
    """The Reference sheet -> the ``references`` table, unless the study is unpublished.

    **An unpublished study gets no reference row, on purpose.** MalAvi cites it in the
    record tables as ``<Authors> unpubl`` and holds no row for it -- there is nothing to
    cite, and 0 of the 526 rows in references.csv are unpublished. ``publish_reference``
    adds the row when the paper appears. Writing a stub row here would create a citation
    to a publication that does not exist, and it would be indistinguishable from a real
    one two years later.

    ``STUDY_TYPE`` is left blank because the template does not ask for it; a curator sets
    it. ``VOLUME_PAGES`` is assembled from the three columns the template does ask for.
    """
    header, body = _sheet(workbook, template_adapter.SHEET_REFERENCE, "REFERENCE_NAME")
    if not body:
        return [], []
    columns = TABLES["references"].columns
    rows, notes = [], _unhoused_notes("Reference", header, body)
    for _row_number, values in body:
        name = (template_adapter.cell(header, values, "REFERENCE_NAME") or "").strip()
        if not name:
            continue
        if reference_names.is_unpublished(name):
            notes.append(f"{name!r} is unpublished, so it correctly gets no row in "
                         f"references.csv; publish_reference adds one when it appears")
            continue
        volume = (template_adapter.cell(header, values, "Volume") or "").strip()
        start = (template_adapter.cell(header, values, "StartPage") or "").strip()
        end = (template_adapter.cell(header, values, "EndPage") or "").strip()
        pages = f"{start}-{end}" if start and end else (start or end)
        row = {column: "" for column in columns}
        row["REFERENCE_NAME"] = name
        row["PUBLICATION_YEAR"] = template_adapter.cell(
            header, values, "PUBLICATION_YEAR") or ""
        row["TITLE"] = normalize.text(
            template_adapter.cell(header, values, "TITLE")) or ""
        row["JOURNAL_NAME"] = normalize.text(
            template_adapter.cell(header, values, "JOURNAL_NAME")) or ""
        row["VOLUME_PAGES"] = f"{volume}:{pages}" if volume and pages else (volume or pages)
        row["_source"] = submission_id
        row["_added"] = release
        rows.append(row)
    return rows, notes


# ---------------------------------------------------------------------------
# The name a submission was approved under
# ---------------------------------------------------------------------------
#
# A proposed lineage name that MalAvi already owns is a WARNING at screen time, not a
# block, because the report offers a free alternative and approving the submission adopts
# it (checks.py: name_already_in_malavi). The ledger records that agreement in
# `entry.name_corrections` and reserves the corrected name.
#
# Until 2026-08-11 nothing applied it here. The workbook's original name went into the
# store verbatim, so a submission approved as TUMIG32 was written as TUMIG10 -- the name
# of a different lineage MalAvi already held, with a different sequence and accession. The
# store then carried two lineages under one name, the reservation feed advertised a name
# the database did not contain, and the only thing that noticed was a warning at build
# time. Rehearsed on the demo submission on 2026-08-11: ledger `name_corrections` said
# {'TUMIG10': 'TUMIG32'} and the store received TUMIG10.
#
# So the agreed name is applied at the write, and a name that still collides afterwards is
# refused rather than written.

def lineage_name_tables() -> Tuple[str, ...]:
    """Every store table with a ``LINEAGE_NAME`` column.

    Derived from the specs rather than listed, so a table added to the store cannot be
    left behind by a rename and end up pointing at a name no lineage carries.
    """
    return tuple(name for name, spec in TABLES.items()
                 if "LINEAGE_NAME" in spec.columns)


def apply_name_corrections(tables: Dict[str, List[Dict[str, Any]]],
                           corrections: Dict[str, str],
                           ) -> Tuple[Dict[str, List[Dict[str, Any]]], List[str]]:
    """Rename lineages to the names the submission was approved under.

    ``corrections`` maps the submitter's proposed name to the name agreed at approval
    (``ledger.Entry.name_corrections``). Every table carrying ``LINEAGE_NAME`` is renamed
    together, because a rename applied to some of them would leave records pointing at a
    lineage row that no longer exists under that name.

    ``ALT_NAME`` is deliberately **not** touched. It records the name a study used, and
    the superseded proposal is a name belonging to a different lineage -- writing it in as
    a synonym would assert that MalAvi's existing TUMIG10 and this new lineage are the
    same thing, which is the opposite of what the rename was for.

    Returns the tables and one note per rename applied, because a name changing between
    what a submitter typed and what MalAvi stores is something an operator must see.
    """
    if not corrections:
        return tables, []

    # Both sides normalized, so a correction the ledger recorded as TUMIG10 still applies
    # to a row read from a workbook -- and would still apply if a future ledger writer
    # stopped upper-casing. The rows themselves are already normalized by `lineage_cell`;
    # this is the belt to that pair of braces, and it costs one dict comprehension.
    corrections = {normalize.lineage_name(proposed) or "":
                   normalize.lineage_name(agreed) or ""
                   for proposed, agreed in corrections.items()}

    applied: Dict[str, int] = {}
    for table in lineage_name_tables():
        for row in tables.get(table, []):
            name = normalize.lineage_name(row.get("LINEAGE_NAME")) or ""
            agreed = corrections.get(name)
            if agreed and agreed != name:
                row["LINEAGE_NAME"] = agreed
                applied[f"{name} -> {agreed}"] = applied.get(f"{name} -> {agreed}", 0) + 1

    notes = [f"{rename} in {count} row(s): the submitter's proposed name was already a "
             f"MalAvi lineage, and this is the name agreed when the submission was "
             f"approved" for rename, count in sorted(applied.items())]
    return tables, notes


def colliding_lineages(store: Dict[str, List[Dict[str, Any]]],
                       incoming: Sequence[Dict[str, Any]],
                       submission_id: str) -> List[str]:
    """New lineages whose names are already taken. One message per collision.

    A non-empty result means the submission must not be written: two rows under one
    ``LINEAGE_NAME`` put two different sequences behind one key, which breaks every
    downstream join, duplicates a tip label in the alignment, and makes the edition report
    unable to say what changed.

    **The submission's own rows are excluded** -- ``replace_submission_rows`` removes and
    rewrites them, so a re-ingest of a correction must not be refused for colliding with
    the version of itself it is about to replace.
    """
    # Both sides are compared in MalAvi's own casing. Comparing raw strings let `tumig19`
    # slip past a store holding `TUMIG19` -- the refusal looked for an exact match and a
    # typo was enough to defeat it.
    held: Dict[str, Dict[str, Any]] = {}
    for row in store.get("lineages", []):
        if str(row.get("_source") or "").strip() == submission_id:
            continue
        name = normalize.lineage_name(row.get("LINEAGE_NAME")) or ""
        if name:
            held.setdefault(name, row)

    messages: List[str] = []
    seen_here: Dict[str, int] = {}
    for row in incoming:
        name = normalize.lineage_name(row.get("LINEAGE_NAME")) or ""
        if not name:
            continue
        seen_here[name] = seen_here.get(name, 0) + 1
        existing = held.get(name)
        if existing is None:
            continue
        accession = str(existing.get("GENBANK_ACC") or "").strip() or "no accession"
        same_sequence = (str(existing.get("SEQUENCE") or "").strip()
                         == str(row.get("SEQUENCE") or "").strip())
        detail = ("the sequences are identical, so this is a record of the lineage MalAvi "
                  "already holds rather than a new one"
                  if same_sequence else
                  "the sequences differ, so this would put two lineages under one name")
        messages.append(
            f"{name} is already a lineage in MalAvi ({accession}) and no rename was "
            f"applied: {detail}. Record the agreed name in the ledger "
            f"(name_corrections) and ingest again.")

    for name, count in sorted(seen_here.items()):
        if count > 1:
            messages.append(
                f"{name} is declared as a new lineage {count} times in this submission, so "
                f"it would enter the store twice under one name.")
    return messages


def tables_from_workbook(path: Path, submission_id: str, release: str,
                         existing_host_rows: Sequence[Dict[str, Any]] = (),
                         ) -> Tuple[Dict[str, List[Dict[str, Any]]], List[str]]:
    """Every store table a submission's workbook supplies. Returns ``(tables, notes)``.

    ``morpho_species`` is absent because the template has no sheet for it: assigning a
    lineage to a named morphospecies is a taxonomic act with its own literature, and a
    curator records it rather than a submitter declaring it.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=True)
    tables: Dict[str, List[Dict[str, Any]]] = {}
    notes: List[str] = []

    rows, table_notes = _host_rows(workbook, submission_id, release, existing_host_rows)
    tables["host_records"] = rows
    notes.extend(table_notes)

    for name, builder in (("lineages", lineage_rows),
                          ("vector_records", vector_rows),
                          ("alt_names", alt_name_rows),
                          ("references", reference_rows)):
        rows, table_notes = builder(workbook, submission_id, release)
        tables[name] = rows
        notes.extend(table_notes)
    return tables, notes


# What MalAvi pads a partial sequence with. The store holds 79,500 "-" against 800 "N", and
# 3,340 of its 5,368 lineages carry fewer than 479 unambiguous bases -- a partial barcode is
# the common case, not an exception. sequence_check._place pads with N because it builds a
# sequence for *display* in the curator report; what goes into the store follows the store.
STORE_GAP = "-"

# How many stored lineages cover only part of the window, for a note that tells a submitter
# their partial sequence is ordinary rather than a problem. Measured 2026-08-20.
PARTIAL_LINEAGES_NOTE = "3,340"


def _ingest_reference(store: Dict[str, List[Dict[str, Any]]],
                      submission_id: str) -> Optional[Any]:
    """A registration reference built from the record store, or None if it is empty.

    Every lineage MalAvi holds occupies the 479 bp window at offset 0 by construction, so
    the store is itself a valid reference. Using it means ingest gains no dependency on the
    exported alignment file, which may be stale or absent when a submission is ingested. The
    submission's own rows are excluded, so a re-ingest does not register against the copy of
    itself it is about to replace.
    """
    held = [str(row.get("SEQUENCE") or "").strip().upper()
            for row in store.get("lineages", [])
            if str(row.get("_source") or "").strip() != submission_id]
    held = [s for s in held if len(s) == MALAVI_WINDOW]
    if not held:
        return None
    return sequence_check.Reference.from_sequences(
        [f"ref{i}" for i in range(len(held))], held, where="the record store")


def _placement(sequence: str, reference: Any) -> Tuple[Optional[int], int, int]:
    """Where `sequence` sits in the window: ``(offset, bases lost at 5', at 3')``.

    ``offset`` is None when it could not be placed at all. The two counts are how many
    *real* bases placing it would throw away -- which is the thing that distinguishes a
    partial barcode from a mis-trimmed one, and it does not depend on length.
    """
    # The default slide is +/-25, which suits a full-window barcode that is a base or two
    # out. A partial read legitimately starts much further in: a reverse-primer-only read
    # covering the last 250 bp begins at window position 230, and with the default bound it
    # came back "could not be placed at all" -- refusing a perfectly good submission. So the
    # bound allows any placement in which the sequence still fits inside the window, plus
    # the usual margin for one that does not. lineage_resolve widens it for the same reason.
    max_offset = max(sequence_check.MAX_OFFSET,
                     MALAVI_WINDOW - len(sequence) + sequence_check.MAX_OFFSET)
    offset, _mismatch = sequence_check._register(sequence, reference,
                                                 max_offset=max_offset)
    if offset is None:
        return None, 0, 0
    lost_start = -offset if offset < 0 else 0
    lost_end = max(0, max(offset, 0) + len(sequence) - lost_start - MALAVI_WINDOW)
    return offset, lost_start, lost_end


def misframed_sequences(store: Dict[str, List[Dict[str, Any]]],
                        incoming: Sequence[Dict[str, Any]],
                        submission_id: str) -> List[str]:
    """New lineages whose sequence cannot be placed without losing data. One message each.

    A non-empty result means the submission must not be written, for the same reason a name
    collision must not: the row would be wrong in a way nothing downstream could detect.

    **Short sequences are normal and are NOT refused.** 3,340 of MalAvi's 5,368 lineages
    hold fewer than 479 unambiguous bases -- ``ABSUP01`` has 326 real bases behind 153
    leading gaps. Sequencing with the forward primer only, or a read that failed at one end,
    produces a perfectly good partial barcode. It is padded into the window and stored.

    **What is refused is losing real bases.** NECMON01 (MALAVI-SUB-2026-000006, 2026-08-20)
    was *exactly* 479 bp and still wrong: it began at frame position 3, so it carried two
    bases past the end of the window, and placing it would have discarded them. The length
    check inside ``lineage_rows`` passed it in silence, and it would have entered the
    alignment two bases out of frame -- 39% identity to its own clade instead of 92%.

    A partial barcode fits inside the window and loses nothing; a mis-trimmed or
    over-length one does not fit and something has to be thrown away. That is the real
    distinction, and unlike a length or a fixed set of shapes it holds for any read.

    An earlier version of this tested membership of ``sequence_check.CANONICAL_SHAPES``,
    which would have refused every partial submission -- the majority of them. The same
    mistake in a stricter form is recorded at the ``canonical`` branch of
    ``sequence_check.check_sequence``.

    **It refuses rather than silently trimming.** The same 2-base shift is produced by a
    mis-windowed read (harmless -- re-place it) and by an indel near the 5' end (a
    sequencing error -- re-read it). A program that quietly trimmed the overhang would
    settle that question by itself and hide the second case.
    """
    reference = _ingest_reference(store, submission_id)
    if reference is None:
        # An empty store is a legitimate state (a fresh seed) and there is nothing to
        # register against; refusing every sequence would be worse than not checking.
        return []

    messages: List[str] = []
    for row in incoming:
        name = normalize.lineage_name(row.get("LINEAGE_NAME")) or ""
        sequence = str(row.get("SEQUENCE") or "").strip().upper()
        if not name or not sequence:
            continue
        offset, lost_start, lost_end = _placement(sequence, reference)
        if offset is None:
            messages.append(
                f"{name}: the sequence could not be placed against MalAvi's reading frame "
                f"at all. It may not be avian haemosporidian cytochrome b, or it may be "
                f"reverse-complemented. A curator has to look at it.")
        elif lost_start or lost_end:
            lost = " and ".join(
                part for part in (
                    f"{lost_start} base(s) at the 5' end" if lost_start else "",
                    f"{lost_end} base(s) at the 3' end" if lost_end else "") if part)
            # A negative offset means the sequence starts before the window -- untrimmed
            # primer or adapter at the 5' end. "frame position -19" is not a sentence
            # anyone can act on, so describe the overhang instead.
            where = (f"placed at frame position {offset + 1}" if offset >= 0
                     else f"starting {-offset} base(s) before the window begins")
            messages.append(
                f"{name}: {len(sequence)} bp {where} does not fit the {MALAVI_WINDOW} bp "
                f"barcode window -- storing it would discard {lost}. A partial sequence is "
                f"fine and is padded into the window; this one is mis-trimmed or is a "
                f"longer amplicon. Either the submitter re-windows it, or a curator decides "
                f"the shift is an indel and asks for the sequence to be re-read; the "
                f"screening report shows both readings.")
    return messages


def place_sequences(incoming: Sequence[Dict[str, Any]],
                    store: Dict[str, List[Dict[str, Any]]],
                    submission_id: str
                    ) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Pad each new lineage's sequence into the 479 bp window. Returns (rows, notes).

    Run **after** :func:`misframed_sequences` has refused anything that does not fit, so
    every sequence reaching here sits inside the window and padding it is arithmetic rather
    than judgment: ``offset`` gaps in front, gaps behind, nothing discarded.

    Without this a primer-trimmed haem amplicon would be stored at 478 bp, and a
    forward-primer-only read at whatever length it happened to be, in a table where all
    5,368 existing rows are exactly 479 characters wide. Nothing downstream expects a short
    row: the alignment would be ragged, and every distance computed against that lineage
    would be measured over a different set of columns from every other one.

    Padding uses :data:`STORE_GAP`, matching the store, so a partial submission is
    indistinguishable from the 3,340 partial lineages MalAvi already holds.

    The workbook is not touched. What the submitter sent stays exactly what they sent --
    this pads the copy on its way into the store, and says so in a note.
    """
    reference = _ingest_reference(store, submission_id)
    if reference is None:
        return list(incoming), []

    rows: List[Dict[str, Any]] = []
    notes: List[str] = []
    for row in incoming:
        row = dict(row)
        name = normalize.lineage_name(row.get("LINEAGE_NAME")) or ""
        sequence = str(row.get("SEQUENCE") or "").strip().upper()
        if name and sequence:
            offset, lost_start, lost_end = _placement(sequence, reference)
            if offset is not None and not lost_start and not lost_end:
                placed = STORE_GAP * offset + sequence
                placed += STORE_GAP * (MALAVI_WINDOW - len(placed))
                if placed != sequence:
                    notes.append(
                        f"{name}: {len(sequence)} bp padded into the {MALAVI_WINDOW} bp "
                        f"window at frame position {offset + 1} "
                        f"({sequence_check._assay_of(offset, len(sequence))}); "
                        f"{offset} gap(s) before, {MALAVI_WINDOW - offset - len(sequence)} "
                        f"after. No base was discarded and the workbook is unchanged.")
                row["SEQUENCE"] = placed
        rows.append(row)
    return rows, notes
