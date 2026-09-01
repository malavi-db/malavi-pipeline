"""Turn a filled ImportMalavi workbook into a submission (schemas/submission.schema.json).

This is the adapter that makes the two intake paths converge. The PDF route already
produces a ``submission.json`` through ``record_builder``; until now a template that
arrived through the Google Form produced only an ad-hoc screen report, which meant the
pre-ingest gate, the row flags and the malaviR validators -- all of which operate on a
submission -- simply never ran on the *higher-value* input. Emitting the same object here
is what lets them run, and it is refactoring rather than new checking.

What this module does NOT do is judge anything. It reads cells, applies the normalization
contract in ``normalize`` and records every change it makes. Every judgment -- is this name
taken, does this sequence already exist, is found <= tested -- belongs to a check, because
a check can report itself and a silent coercion cannot.

Two joins are performed, and both are lookups the workbook itself defines rather than
inferences:

* ``Hosts_and_Sites.SiteName`` -> ``Sites.SITE_NAME`` supplies each record's coordinates.
  The template tells submitters those strings must match, so following it is reading, not
  guessing. When it does not resolve, the record simply has no coordinates and a check can
  say so.
* ``Hosts_and_Sites.LINEAGE_NAME`` -> ``NewLineages.ParasiteGenus`` supplies the parasite
  genus, which the records sheet has no column for. A row naming a lineage already in
  MalAvi gets no genus from this, which is correct: the release knows it, and inventing it
  here would put an unsourced value in front of a curator.
"""
from __future__ import annotations

from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from . import __version__
from .normalize import (
    accession_list, clean_count, clean_genus, lineage_name, record_change,
    sequence_pair, source_ref, text,
)

# The version of schemas/submission.schema.json this adapter writes against. Bumped only
# when the contract changes in a way a reader must notice.
SCHEMA_VERSION = "1.0.0"

SHEET_NEWLINEAGES = "NewLineages"
SHEET_SEQUENCES = "Sequences"
SHEET_REFERENCE = "Reference"
SHEET_HOSTS = "Hosts_and_Sites"
SHEET_SITES = "Sites"
SHEET_ALT_NAMES = "Alt_Lineage_names"
SHEET_VECTORS = "Vectors"

# The gray italic worked example that ships on each sheet, identified by the leading cells
# of its own row. The READ ME tells submitters they may leave it in place, so it must never
# be read as submitted data.
#
# This used to be a flat list of marker strings matched anywhere in a row, and that was
# wrong in a way that quietly destroyed real submissions: the markers included "SGS1" and
# "GRW04", which are not just example values but two of the most-recorded lineages in
# MalAvi -- 571 and 187 host records respectively in the 2026-03-23 release. A submitter
# reporting a new SGS1 record had that row dropped from both the screen and the
# submission, with nothing said. Matching is therefore per sheet, on the leading cells,
# and only for the row in the example's own position (see ``_is_example_row``).
EXAMPLE_ROW_SIGNATURES = {
    SHEET_NEWLINEAGES: ("ALCPOI02",),
    SHEET_SEQUENCES: ("ALCPOI02",),
    SHEET_HOSTS: ("ALCPOI02",),
    SHEET_REFERENCE: ("Gupta et al 2019",),
    SHEET_SITES: ("Ambalapara",),
    # These two sheets' examples lead with real, common lineage names, so the second
    # column is required as well: the example's alternative name and vector species are
    # what make the row the example rather than somebody's data.
    SHEET_ALT_NAMES: ("SGS1", "P15"),
    SHEET_VECTORS: ("GRW04", "Culex pipiens"),
}


def _is_example_row(sheet_name: str, values: tuple) -> bool:
    """Does this row match the shipped worked example for this sheet?

    Compares only the leading cells the signature names, trimmed and case-insensitively.
    A submitter who has typed over the example changes those cells and their row is read
    as the data it now is.
    """
    signature = EXAMPLE_ROW_SIGNATURES.get(sheet_name)
    if not signature or len(values) < len(signature):
        return False
    for index, expected in enumerate(signature):
        cell_value = values[index]
        actual = "" if cell_value is None else str(cell_value).strip()
        if actual.casefold() != expected.casefold():
            return False
    return True

# The column order the template ships with, per sheet. This is the contract the template
# generator writes and this adapter reads; ``test_template_adapter`` asserts the two still
# agree, which is what keeps a single source of truth without importing a generator script
# into the package.
#
# It exists to repair one specific real-world damage: a submitter whose copy of the
# workbook has a *blank header cell* over a column that still holds data. Observed in the
# Shimizu 2026 submission, where the NewLineages "HostSpecies" header was gone but every
# host name was present -- and that column is the one a lineage's acronym is built from,
# so losing it silently would be expensive.
CANONICAL_HEADERS = {
    SHEET_NEWLINEAGES: ["LINEAGE_NAME", "GENBANK_NR", "ParasiteGenus", "HostSpecies",
                        "HOST_SPECIES_ID", "Reference", "COMMENT"],
    SHEET_SEQUENCES: ["LINEAGE_NAME", "SEQUENCE"],
    SHEET_REFERENCE: ["REFERENCE_NAME", "PUBLICATION_YEAR", "TITLE", "JOURNAL_NAME",
                      "Volume", "StartPage", "EndPage", "DOI"],
    SHEET_HOSTS: ["LINEAGE_NAME", "HostSpecies", "HOST_SPECIES_ID", "HostSubspecies",
                  "HostAge", "HostStatus", "HostEnvironment", "Country", "CountryRegion",
                  "SiteName", "NUMBER_FOUND", "NUMBER_TESTED", "Reference", "COMMENT"],
    SHEET_SITES: ["SITE_NAME", "Country", "LATITUDE", "LONGITUDE", "ALTITUDE(m)"],
    SHEET_ALT_NAMES: ["MalAvi_Name", "Alternative_Name", "GenBankNr", "Reference",
                      "Comment"],
    SHEET_VECTORS: ["LINEAGE_NAME", "VectorSpecies", "VECTOR_METHOD", "Country",
                    "CountryRegion", "SiteName", "No_found", "No_tested", "Reference",
                    "Comment"],
}

# Sheets whose presence means "this workbook is a filled template". A submission often
# carries supplementary spreadsheets alongside the template, and those are for the PDF
# and table-extraction path, not this one.
TEMPLATE_SHEETS = frozenset(CANONICAL_HEADERS)


def looks_like_template(workbook) -> bool:
    """Is this workbook a filled ImportMalavi template rather than a supplement?

    Requires one of the sheets that actually carries submitted data. A workbook holding
    only "Table S1" and "Author Information" is a supplementary file: adapting it yields
    an empty submission, and an empty submission put in front of a curator is a phantom
    to be dismissed rather than information.
    """
    return any(name in TEMPLATE_SHEETS for name in workbook.sheetnames)


def _repair_header(sheet_name: str, header: List[str]) -> Tuple[List[str], List[str]]:
    """Fill blank header cells from the canonical column order, conservatively.

    Only repairs when **every non-blank header already matches the canonical header at
    its own position**. That is the evidence that the sheet's layout is untouched and the
    only damage is a missing label; if a submitter has reordered, inserted or renamed a
    column, nothing is assumed and the blanks stay blank. Positional recovery on a
    reordered sheet would file one column's values under another column's name, which is
    far worse than not reading the column at all.

    Returns the header and a list of human-readable repair notes for the report.
    """
    canonical = CANONICAL_HEADERS.get(sheet_name)
    if not canonical:
        return header, []

    blanks = [index for index, name in enumerate(header)
              if not name and index < len(canonical)]
    if not blanks:
        return header, []

    for index, name in enumerate(header):
        if name and (index >= len(canonical) or name.lower() != canonical[index].lower()):
            return header, []               # layout differs; assume nothing

    repaired = list(header)
    notes: List[str] = []
    for index in blanks:
        repaired[index] = canonical[index]
        notes.append(f"{sheet_name}: column {index + 1} had no header; read as "
                     f"'{canonical[index]}' from the template's column order")
    return repaired, notes


def sheet_rows(worksheet, key_header: str,
               repairs: Optional[List[str]] = None,
               ) -> Tuple[List[str], List[Tuple[int, tuple]]]:
    """Split a template sheet into its header and its real data rows.

    Returns ``(header, [(row_number, values), ...])``. The row number is the true
    worksheet row -- the number the submitter sees down the left-hand side of Excel --
    so a finding can name the cell they need to look at. That is the whole reason this
    does not simply filter a list of values.

    Rows above the header are skipped (the sheet carries an instruction note), fully
    empty rows are skipped (the template ships 200 blank formatted rows below the
    example), and the worked example is skipped.

    When ``repairs`` is given, any blank header cell recovered from the template's
    canonical column order is appended to it as a note, so the report can tell the
    curator that a column was read positionally rather than by its label.
    """
    header: List[str] = []
    body: List[Tuple[int, tuple]] = []
    seen_header = False
    # The example sits immediately below the header. Requiring that position as well as
    # a value match means a genuine record for one of the example's lineages -- an
    # ALCPOI02 or SGS1 row further down the sheet -- is still read as data.
    at_first_data_row = True

    # ``iter_rows`` yields every row from 1 in order, so enumerating it gives the real
    # worksheet row number even though blank rows are dropped from the result.
    for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if not values or all(value in (None, "") for value in values):
            continue
        first_cell = str(values[0]).strip() if values[0] is not None else ""
        if not seen_header:
            if first_cell == key_header:
                header = [("" if cell is None else str(cell).strip()) for cell in values]
                header, notes = _repair_header(worksheet.title, header)
                if repairs is not None:
                    repairs.extend(notes)
                seen_header = True
            continue
        if at_first_data_row:
            at_first_data_row = False
            if _is_example_row(worksheet.title, values):
                continue
        body.append((row_number, values))
    return header, body


def cell(header: List[str], row: tuple, column: str) -> Optional[str]:
    """Read one named column out of a row, case-insensitively, or ``None``.

    Returns the raw cell as a string without normalization. Callers normalize, so that
    the submitted and normalized forms can both be reported.
    """
    for index, name in enumerate(header):
        if name.lower() == column.lower() and index < len(row):
            value = row[index]
            return None if value is None else str(value).strip()
    return None


def _raw_cell(header: List[str], row: tuple, column: str) -> Any:
    """The cell's underlying value, so counts keep the type openpyxl gave them."""
    for index, name in enumerate(header):
        if name.lower() == column.lower() and index < len(row):
            return row[index]
    return None


def _notes_from(pairs: List[Tuple[str, Any]]) -> Optional[str]:
    """Join the template columns MalAvi's submission schema has no field for.

    The schema fixes the record fields (``additionalProperties: false``), so anything a
    curator needs to *see* but the schema does not hold goes into notes -- the same
    convention ``record_builder._row_notes`` follows for the PDF path. Nothing submitted
    is discarded merely because the schema has no slot for it.
    """
    parts = [f"{label}: {value}" for label, value in pairs if value not in (None, "")]
    return "; ".join(parts) if parts else None


def _reference_row(header: List[str], row: tuple) -> Dict[str, Any]:
    """One Reference sheet row as the schema's reference object."""
    year_raw = _raw_cell(header, row, "PUBLICATION_YEAR")
    # clean_count returns a float because it guards individual counts, where the point is
    # rejecting non-integers. A publication year is displayed, and "2026.0" in a curator's
    # report reads as a bug in our software rather than a year.
    year = clean_count(year_raw)
    year = int(year) if year is not None else None
    return {
        "doi": text(cell(header, row, "DOI")),
        "pmid": None,                       # the template does not ask for a PMID
        "title": text(cell(header, row, "TITLE")),
        "year": year,
    }


def build_submission_from_workbook(
    workbook,
    workbook_name: str,
    submitter: Optional[Dict[str, str]] = None,
    validate: bool = True,
) -> Dict[str, Any]:
    """Build a schema-valid submission from an already-open openpyxl workbook.

    Args:
        workbook: an ``openpyxl`` workbook opened with ``data_only=True``.
        workbook_name: file name, recorded so a finding can name the file it came from.
        submitter: ``{name, email?}``; the Form's metadata supplies this when known.
        validate: jsonschema-validate the result before returning.
    """
    changes: List[Dict[str, Any]] = []
    repairs: List[str] = []

    def sheet(name: str):
        return workbook[name] if name in workbook.sheetnames else None

    # ---- proposed new lineages ------------------------------------------------------
    # Read first because the records sheet has no parasite-genus column and this is where
    # the genus for a newly proposed lineage is stated.
    proposed: List[Dict[str, Any]] = []
    genus_by_lineage: Dict[str, str] = {}
    worksheet = sheet(SHEET_NEWLINEAGES)
    if worksheet is not None:
        header, body = sheet_rows(worksheet, "LINEAGE_NAME", repairs)
        for row_number, row in body:
            submitted_name = cell(header, row, "LINEAGE_NAME")
            name = lineage_name(submitted_name)
            if not name:
                continue
            where = source_ref(sheet=SHEET_NEWLINEAGES, row=row_number,
                               file=workbook_name)
            record_change(changes, "lineage_name", submitted_name, name, where)
            genus = clean_genus(cell(header, row, "ParasiteGenus"))
            if genus:
                genus_by_lineage[name] = genus
            proposed.append({
                "lineage_name": name,
                "parasite_genus": genus,
                "host_species": text(cell(header, row, "HostSpecies")),
                "accessions": accession_list(cell(header, row, "GENBANK_NR")),
                "reference": text(cell(header, row, "Reference")),
                "notes": _notes_from([
                    ("COMMENT", text(cell(header, row, "COMMENT"))),
                    ("HOST_SPECIES_ID", text(cell(header, row, "HOST_SPECIES_ID"))),
                ]),
                "source": where,
            })

    # ---- sequences ------------------------------------------------------------------
    sequences: List[Dict[str, Any]] = []
    worksheet = sheet(SHEET_SEQUENCES)
    if worksheet is not None:
        header, body = sheet_rows(worksheet, "LINEAGE_NAME", repairs)
        for row_number, row in body:
            submitted_name = cell(header, row, "LINEAGE_NAME")
            name = lineage_name(submitted_name)
            submitted_seq, cleaned_seq = sequence_pair(_raw_cell(header, row, "SEQUENCE"))
            if not name and not cleaned_seq:
                continue
            where = source_ref(sheet=SHEET_SEQUENCES, row=row_number, file=workbook_name)
            record_change(changes, "lineage_name", submitted_name, name, where)
            sequences.append({
                "lineage_name": name,
                "sequence": submitted_seq,
                "sequence_clean": cleaned_seq,
                "source": where,
            })

    # ---- sites, for the coordinate join ---------------------------------------------
    coordinates_by_site: Dict[str, str] = {}
    worksheet = sheet(SHEET_SITES)
    if worksheet is not None:
        header, body = sheet_rows(worksheet, "SITE_NAME", repairs)
        for _row_number, row in body:
            site_name = text(cell(header, row, "SITE_NAME"))
            latitude = text(cell(header, row, "LATITUDE"))
            longitude = text(cell(header, row, "LONGITUDE"))
            # Both halves or nothing: half a coordinate pair is not a location, and
            # storing one would look like data while being useless.
            if site_name and latitude and longitude:
                coordinates_by_site[site_name.casefold()] = f"{latitude} {longitude}"

    # ---- records --------------------------------------------------------------------
    records: List[Dict[str, Any]] = []
    worksheet = sheet(SHEET_HOSTS)
    if worksheet is not None:
        header, body = sheet_rows(worksheet, "LINEAGE_NAME", repairs)
        for row_number, row in body:
            submitted_name = cell(header, row, "LINEAGE_NAME")
            name = lineage_name(submitted_name)
            host = text(cell(header, row, "HostSpecies"))
            # A row with neither a lineage nor a host is a stray -- a leftover formatted
            # row or a note the submitter typed in the margin. It is dropped here rather
            # than carried as an empty record, because row_flags would otherwise tier it
            # as "incomplete" and put a phantom in the curator's queue.
            if not name and not host:
                continue
            where = source_ref(sheet=SHEET_HOSTS, row=row_number, file=workbook_name)
            record_change(changes, "lineage_name", submitted_name, name, where)

            site = text(cell(header, row, "SiteName"))
            records.append({
                "lineage_name": name,
                "host_species": host,
                "country": text(cell(header, row, "Country")),
                "site": site,
                "coordinates": coordinates_by_site.get(site.casefold()) if site else None,
                "parasite_genus": genus_by_lineage.get(name) if name else None,
                "number_tested": clean_count(_raw_cell(header, row, "NUMBER_TESTED")),
                "number_found": clean_count(_raw_cell(header, row, "NUMBER_FOUND")),
                "notes": _notes_from([
                    ("COMMENT", text(cell(header, row, "COMMENT"))),
                    ("subspecies", text(cell(header, row, "HostSubspecies"))),
                    ("age", text(cell(header, row, "HostAge"))),
                    ("status", text(cell(header, row, "HostStatus"))),
                    ("environment", text(cell(header, row, "HostEnvironment"))),
                    ("region", text(cell(header, row, "CountryRegion"))),
                    ("HOST_SPECIES_ID", text(cell(header, row, "HOST_SPECIES_ID"))),
                    ("reference", text(cell(header, row, "Reference"))),
                ]),
                # source_scope answers "is this the paper's own data or reprinted from
                # someone else's?", which is a question about extraction from a
                # publication. A submitter filling in the template is telling us these
                # are their records; there is no textual evidence here to classify, and
                # asserting "focal" would be inventing a judgment the workbook never
                # made. Left unset for the curator.
                "source_scope": None,
                "source": where,
            })

    # ---- vectors --------------------------------------------------------------------
    vectors: List[Dict[str, Any]] = []
    worksheet = sheet(SHEET_VECTORS)
    if worksheet is not None:
        header, body = sheet_rows(worksheet, "LINEAGE_NAME", repairs)
        for row_number, row in body:
            submitted_name = cell(header, row, "LINEAGE_NAME")
            name = lineage_name(submitted_name)
            species = text(cell(header, row, "VectorSpecies"))
            if not name and not species:
                continue
            where = source_ref(sheet=SHEET_VECTORS, row=row_number, file=workbook_name)
            record_change(changes, "lineage_name", submitted_name, name, where)
            vectors.append({
                "lineage_name": name,
                "vector_species": species,
                "vector_method": text(cell(header, row, "VECTOR_METHOD")),
                "country": text(cell(header, row, "Country")),
                "site": text(cell(header, row, "SiteName")),
                "notes": _notes_from([
                    ("Comment", text(cell(header, row, "Comment"))),
                    ("region", text(cell(header, row, "CountryRegion"))),
                    ("found", text(cell(header, row, "No_found"))),
                    ("tested", text(cell(header, row, "No_tested"))),
                    ("reference", text(cell(header, row, "Reference"))),
                ]),
                "source_scope": None,
                "source": where,
            })

    # ---- references -----------------------------------------------------------------
    # The schema holds one reference per submission, which is the ordinary case: one
    # workbook describes one study. A submitter who lists several is not wrong, so the
    # extras are kept in provenance where a curator can see them rather than dropped.
    reference: Dict[str, Any] = {"doi": None, "pmid": None, "title": None, "year": None}
    extra_references: List[Dict[str, Any]] = []
    reference_keys: List[str] = []
    worksheet = sheet(SHEET_REFERENCE)
    if worksheet is not None:
        header, body = sheet_rows(worksheet, "REFERENCE_NAME", repairs)
        for index, (_row_number, row) in enumerate(body):
            key = text(cell(header, row, "REFERENCE_NAME"))
            if key:
                reference_keys.append(key)
            parsed = _reference_row(header, row)
            if index == 0:
                reference = parsed
            else:
                extra_references.append(parsed)

    # ---- accessions -----------------------------------------------------------------
    # Every accession the workbook names, from the new lineages and from any alternative
    # names for lineages MalAvi already holds. Collected top-level because that is where
    # the gate's format and INSDC-resolution checks read them.
    all_accessions: List[str] = []
    for entry in proposed:
        for accession in entry["accessions"]:
            if accession not in all_accessions:
                all_accessions.append(accession)
    worksheet = sheet(SHEET_ALT_NAMES)
    alternative_names: List[Dict[str, Any]] = []
    if worksheet is not None:
        header, body = sheet_rows(worksheet, "MalAvi_Name", repairs)
        for row_number, row in body:
            malavi_name = lineage_name(cell(header, row, "MalAvi_Name"))
            alternative = text(cell(header, row, "Alternative_Name"))
            if not malavi_name and not alternative:
                continue
            accessions = accession_list(cell(header, row, "GenBankNr"))
            for accession in accessions:
                if accession not in all_accessions:
                    all_accessions.append(accession)
            alternative_names.append({
                "malavi_name": malavi_name,
                "alternative_name": alternative,
                "accessions": accessions,
                "source": source_ref(sheet=SHEET_ALT_NAMES, row=row_number,
                                     file=workbook_name),
            })

    submission: Dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "submitter": submitter or {"name": "template_submission"},
        "reference": reference,
        "accessions": sorted(all_accessions),
        "records": records,
        "vectors": vectors,
        "sequences": sequences,
        "proposed_lineages": proposed,
        # Built above and, until 2026-08-19, dropped on the floor: only the accessions
        # off these rows were kept, and the declarations themselves reached nothing. A
        # submitter saying "the lineage you call SGS1 is the one I published as P15" is
        # making a claim about MalAvi's own data, and no curator was ever shown it.
        "alternative_names": alternative_names,
        "provenance": {
            "source": "template",
            "tool_version": __version__,
            "extracted_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "needs_review": True,
            "workbook": workbook_name,
            "normalizations": changes,
            # Columns read by position because their header cell was blank. Kept beside
            # the normalizations because it is the same kind of fact: something the
            # system decided about the submitter's file, which the curator gets to see.
            "header_repairs": repairs,
            # Nested extra keys are permitted by the schema, and these are exactly the
            # things a curator needs in view but MalAvi's own tables have no column for.
            "reference_keys": reference_keys,
            "extra_references": extra_references,
            "alternative_names": alternative_names,
            "sheets_present": [name for name in workbook.sheetnames],
        },
    }

    if validate:
        import jsonschema  # a real dependency: this path runs by default

        from .record_builder import _submission_schema

        jsonschema.validate(submission, _submission_schema())

    return submission


def build_submission_from_path(path, submitter: Optional[Dict[str, str]] = None,
                               validate: bool = True) -> Optional[Dict[str, Any]]:
    """Open a workbook and build its submission, or ``None`` if it is not a template.

    ``path`` is a ``pathlib.Path``. Returning ``None`` for a supplementary spreadsheet is
    deliberate: those belong to the table-extraction path, and emitting an empty
    submission for each would put phantoms in the curator's queue.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        if not looks_like_template(workbook):
            return None
        return build_submission_from_workbook(
            workbook, path.name, submitter=submitter, validate=validate)
    finally:
        workbook.close()
