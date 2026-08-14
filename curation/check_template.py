#!/usr/bin/env python3
# @title Screen a filled submission template deterministically
# @purpose Read an ImportMalavi template workbook, cross-check what the submitter
#   wrote, and run every declared sequence against the pinned MalAvi release.
# @why A submitted template is the highest-value input the project receives, and
#   the failures that matter are not typos: a sequence pasted outside the MalAvi
#   reading frame enters the alignment shifted, and a sequence already in MalAvi
#   must never be handed a new name. Both are caught here before a curator reads
#   a single row.
# @input curation/intake/submissions/<dir>/ (or a single .xlsx)
# @input docs/assets/downloads/malavi_alignment_<release>.fasta
# @output printed report + <submission>/screen.json
# @program python3
# @program openpyxl
# @critical-var SHEET_SEQUENCES
# @critical-var SHEET_NEWLINEAGES
"""Deterministic screen of a filled submission template.

Rule-based throughout: same workbook in, same report out. Nothing is inferred
probabilistically and nothing is auto-ingested -- the output is a report for a
curator, exactly as the pre-ingest gate requires.

Checks performed
----------------
* every sequence is registered to the MalAvi reading frame (offset reported)
* no submitted sequence is already a named MalAvi lineage
* stop codons in the correct frame
* nearest named lineages, with the distance and how many positions were compared
* the proposed lineage name is not already taken
* NewLineages, Sequences and Hosts_and_Sites agree on which lineages exist
* accessions are well formed
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parent / "src"))
from malavi_curation.config import load_config, repo_root          # noqa: E402
from malavi_curation.sequence_check import (                       # noqa: E402
    Reference, check_sequence, clean, default_alignment_path,
)

# The workbook-reading primitives, the sheet names and the worked-example markers all
# live in the adapter now, so that the screen below and the submission built from the
# same workbook can never disagree about which rows are data. A divergence there would
# mean a curator was screened on one set of rows and shown another.
from malavi_curation.alignment_figure import build_figures         # noqa: E402
from malavi_curation.checks import render_console, run_checks      # noqa: E402
from malavi_curation.form_metadata import submitter_from_metadata  # noqa: E402
from malavi_curation import enrollment                             # noqa: E402
from malavi_curation import normalize                              # noqa: E402
from malavi_curation import reference_names                        # noqa: E402
from malavi_curation.submission_id import submission_id_for        # noqa: E402
from malavi_curation.naming import suggest_name                    # noqa: E402
from malavi_curation.report_html import (                          # noqa: E402
    render_paper_only_report, render_report, write_pdf, write_report,
)
from malavi_curation.template_adapter import (                     # noqa: E402
    SHEET_HOSTS, SHEET_NEWLINEAGES, SHEET_REFERENCE, SHEET_SEQUENCES,
    build_submission_from_path, cell as _cell, sheet_rows,
)

ACCESSION_RE = re.compile(r"^[A-Z]{1,2}[0-9]{5,6}(\.[0-9]+)?$")


def _header_and_body(ws, key_header: str) -> Tuple[List[str], List[tuple]]:
    """Header plus data rows, with the row numbers dropped.

    ``sheet_rows`` carries the true worksheet row number for every row so that a
    finding can point at the cell a submitter needs to fix. This screen reports by
    lineage name rather than by row, so it takes the values alone; the row numbers
    reach the curator through the submission the adapter builds.
    """
    header, body = sheet_rows(ws, key_header)
    return header, [values for _row_number, values in body]


def _lineage_cell(hdr, row, column: str = "LINEAGE_NAME") -> str:
    """Read a lineage name from the workbook in MalAvi's own casing.

    MalAvi lineage names are upper case throughout the database, so a submitter typing
    ``sgs1``, ``Sgs1`` or ``SGS 1`` means ``SGS1``. Three consumers of this same cell used
    to disagree: the submission record normalized it (normalize.lineage_name), the name
    reservation feed upper-cased it again, and the screen compared the raw string against
    the release's upper-case names. So `sgs1` failed to match `SGS1` and the blocking
    "this name is already taken" check did not fire -- while the record built from the same
    row was stored as `SGS1`.

    The same drift split the screen's own bookkeeping: `TUMIG19` on one sheet and
    `tumig19` on another were two different keys, which spuriously raised BOTH
    `lineage_without_sequence` and `sequence_without_declaration` for what is one lineage.
    """
    return normalize.lineage_name(_cell(hdr, row, column)) or ""


def standing_claims_for(root: Path, cfg: dict, target: Path,
                        known: Optional[set]) -> Dict[str, str]:
    """Names claimed by the OTHER submissions in the queue, for the screen to respect.

    Excluded from the answer: the submission being screened (a re-run must not find that
    this submission already claimed its own names) and anything in
    ``submissions.exclude``, which is how test and withdrawn submissions are kept out of
    the reservation feed. Using the same exclusion list here keeps the screen and the
    public feed answering the same question.
    """
    submissions_cfg = cfg.get("submissions") or {}
    inbox = root / submissions_cfg.get("inbox_dir", "curation/intake/submissions")
    skip = [entry["id"] for entry in (submissions_cfg.get("exclude") or [])
            if entry.get("id")]
    # `target` is either a submission directory or a single workbook inside one. Either
    # way the directory to exclude is the one directly under the inbox.
    try:
        relative = (target if target.is_dir() else target.parent).resolve().relative_to(
            inbox.resolve())
        if relative.parts:
            skip.append(relative.parts[0])
    except (ValueError, OSError):
        # Screening something outside the inbox -- a one-off workbook a curator was sent
        # directly. Nothing to exclude; every claim in the queue still applies.
        pass
    return enrollment.standing_claims(inbox, exclude=skip, known=known)


def screen(workbook: Path, ref: Reference, known_lineages: Optional[set],
           claimed_elsewhere: Optional[Dict[str, str]] = None) -> dict:
    """Screen one workbook.

    ``known_lineages`` is the set of names MalAvi already owns, or **None** when the
    database snapshot has not been generated. The distinction matters more than it looks:
    an empty set and a missing snapshot are indistinguishable to `in`, so passing `set()`
    for "we could not look" made the highest-stakes blocking check in the project --
    "is this name already taken?" -- report a clean pass for a name MalAvi already has.
    None makes the check report as skipped, with a reason, which is what every other
    unavailable check in this pipeline does.

    ``claimed_elsewhere`` maps an upper-cased name to the submission directory that
    claimed it first -- see :func:`enrollment.standing_claims`. A released name and a
    claimed name are different findings and are reported separately: the first is settled
    and the submitter must rename, the second is a queue position that a curator may need
    to check against arrival dates. Both make the name unavailable for the free-name
    search below, which is the failure that mattered: without this, two submitters were
    offered the same replacement name and both were told it was confirmed.
    """
    import openpyxl
    wb = openpyxl.load_workbook(workbook, data_only=True)
    out: dict = {"workbook": workbook.name, "sequences": [], "issues": [], "lineages": {},
                 "skipped": []}

    if known_lineages is None:
        out["skipped"].append({
            "code": "name_already_in_malavi",
            "reason": "the malaviR database snapshot has not been generated "
                      "(run curation/r/gate_reference.R)"})

    def issue(sev: str, code: str, msg: str, subject: str = None):
        """Record one issue.

        ``code`` is a stable identifier for *which* check raised this, so the
        check registry can key on it instead of matching message text. Message
        wording is free to change; a code is not.

        ``subject`` is what the issue is *about* -- usually a lineage name. It is
        stated rather than parsed back out of the message, so a report can list
        the lineages a check flagged without every reader re-deriving them.
        """
        entry = {"severity": sev, "code": code, "message": msg}
        if subject:
            entry["subject"] = subject
        out["issues"].append(entry)

    # ---- declared new lineages ------------------------------------------
    declared: Dict[str, dict] = {}
    if SHEET_NEWLINEAGES in wb.sheetnames:
        hdr, body = _header_and_body(wb[SHEET_NEWLINEAGES], "LINEAGE_NAME")
        for r in body:
            name = _lineage_cell(hdr, r)
            if not name:
                continue
            accs = [a.strip() for a in re.split(r"[,;\s]+", _cell(hdr, r, "GENBANK_NR") or "") if a.strip()]
            declared[name] = {"genus": _cell(hdr, r, "ParasiteGenus"), "accessions": accs}
            # A name that does not follow from its host is a typo until shown otherwise.
            # Nothing else looks at the SHAPE of a proposed name: an invented acronym is
            # not in the release, so the collision check passes it, and the name goes into
            # a paper and into GenBank where it is very hard to undo.
            host_for_name = _cell(hdr, r, "HostSpecies")
            if host_for_name:
                shape = suggest_name(host_for_name, [])
                if shape.ok:
                    acronyms = [option.acronym for option in shape.options]
                    stem = re.match(r"^([A-Z]+)\d+$", name)
                    if not stem:
                        issue("warn", "name_malformed",
                              f"{name} is not an acronym followed by a number.", name)
                    elif stem.group(1) not in acronyms:
                        issue("warn", "name_malformed",
                              f"{name} does not follow from {host_for_name}: that host "
                              f"gives {' or '.join(acronyms)}, not {stem.group(1)}.", name)

            if known_lineages is not None and name in known_lineages:
                # The suggestion is appended later, once the free name is worked out.
                issue("warn", "name_already_in_malavi",
                      f"proposed name {name} is ALREADY a MalAvi lineage name.", name)
            elif claimed_elsewhere and name.upper() in claimed_elsewhere:
                # Not "already in MalAvi" -- nobody has been granted this yet. It is a
                # queue position: an earlier submission asked for it first, and priority
                # goes by the date the submission arrived. Reported separately so the
                # curator can check the dates rather than tell a submitter their name is
                # taken by a release that does not contain it.
                issue("warn", "name_claimed_by_another_submission",
                      f"proposed name {name} was already claimed by submission "
                      f"{claimed_elsewhere[name.upper()]}, which is still in the queue. "
                      f"Priority goes to whichever arrived first -- check the dates in "
                      f"docs/assets/data/reserved_names.json before granting it.", name)
            for a in accs:
                if not ACCESSION_RE.match(a.upper()):
                    issue("warn", "accession_malformed",
                          f"{name}: '{a}' is not a well-formed nucleotide accession.", name)
    out["lineages"] = declared

    # ---- sequences -------------------------------------------------------
    seqs: Dict[str, str] = {}
    if SHEET_SEQUENCES in wb.sheetnames:
        hdr, body = _header_and_body(wb[SHEET_SEQUENCES], "LINEAGE_NAME")
        for r in body:
            name = _lineage_cell(hdr, r)
            raw = _cell(hdr, r, "SEQUENCE")
            if name and raw:
                seqs[name] = clean(raw)

    for name, s in seqs.items():
        res = check_sequence(s, ref, label=name)
        entry = res.as_dict()
        # The cleaned sequence itself, so the report can print it as FASTA. A curator
        # meeting this pipeline will want to verify a lineage call in their own aligner
        # rather than take the program's word for it, and without this the report had
        # nothing to give them -- the block silently rendered empty.
        entry["sequence"] = s
        out["sequences"].append(entry)
        if res.verdict == "known_lineage":
            issue("error", "sequence_is_known_lineage",
                  f"{name}: sequence is IDENTICAL to existing lineage "
                  f"{res.nearest[0][0]} — do not assign a new name.", name)
        if "needs_reframing" in res.flags:
            issue("warn", "sequence_needs_reframing", f"{name}: {res.notes[0]}", name)
        if "contains_stop_codon" in res.flags:
            issue("warn", "sequence_stop_codon",
                  f"{name}: {res.n_stop_codons} stop codon(s) in the reference frame.", name)
        if res.verdict == "unplaceable":
            issue("error", "sequence_unplaceable",
                  f"{name}: could not be registered against the reference alignment.", name)

    # ---- cross-sheet agreement ------------------------------------------
    for name in declared:
        if name not in seqs:
            issue("error", "lineage_without_sequence",
                  f"{name} is declared in {SHEET_NEWLINEAGES} but has no sequence.", name)
    for name in seqs:
        if name not in declared:
            issue("warn", "sequence_without_declaration",
                  f"{name} has a sequence but is not declared in {SHEET_NEWLINEAGES}.", name)

    if SHEET_HOSTS in wb.sheetnames:
        hdr, body = _header_and_body(wb[SHEET_HOSTS], "LINEAGE_NAME")
        host_rows = 0
        used = set()
        for r in body:
            name = _lineage_cell(hdr, r)
            if not name:
                continue
            host_rows += 1
            used.add(name)
            if not _cell(hdr, r, "Country"):
                issue("warn", "record_without_country",
                      f"{name}: host record has no Country.", name)
            tested, found = _cell(hdr, r, "NUMBER_TESTED"), _cell(hdr, r, "NUMBER_FOUND")
            if not tested and not found:
                issue("info", "record_without_prevalence",
                      f"{name}: no prevalence (NUMBER_FOUND / NUMBER_TESTED) given.", name)
        out["n_host_records"] = host_rows
        for name in declared:
            if name not in used:
                issue("warn", "lineage_without_host_record",
                      f"{name} is a new lineage with no row in {SHEET_HOSTS}.", name)

    if SHEET_REFERENCE in wb.sheetnames:
        hdr, body = _header_and_body(wb[SHEET_REFERENCE], "REFERENCE_NAME")
        names = [_cell(hdr, r, "REFERENCE_NAME") for r in body]
        out["references"] = names
        if not body:
            # Still an error, and still for an unpublished study. What a submission needs
            # is not a *publication* but a citation key: every record row points at one,
            # and records with nothing to point at cannot be loaded. An unpublished study
            # supplies it as "<Authors> unpubl" and leaves the year, journal and pages
            # blank -- see malavi_curation.reference_names.
            issue("error", "reference_missing",
                  "no Reference row: every submission needs a reference name. If the "
                  "study is not published yet, name it '<Authors> unpubl' (for example "
                  "'Barrow et al unpubl') and leave the year, journal and pages blank.")
        # The unpublished marker has to be spelled MalAvi's way or a curator filtering
        # unpublished records will quietly miss this study. Warning, not blocking: the
        # name is a curator's to settle, and a misspelling is fixed in a correction
        # rather than by sending the whole submission back.
        for name in names:
            problem = reference_names.problem_with(name)
            if problem:
                issue("warn", "reference_unpubl_malformed", problem, name)
    return out


def offer_free_names(reports: List[dict], known: Optional[set],
                     claimed_elsewhere: Optional[Dict[str, str]],
                     submissions: List[dict]) -> None:
    """Record a free alternative for every proposed name that is not available.

    Mutates each report in place, adding ``name_suggestions`` and ``not_new_lineages``.

    Extracted from main() so it can be tested directly: the thing it must never do is
    offer submission B a name submission A has already been offered, and that is worth a
    test that does not need a whole run to reach it.
    """
    # A proposed name MalAvi already owns gets a free alternative recorded here, on the
    # screen itself, rather than computed when a report happens to be rendered. It has to
    # be durable: a curator approves a submission *including* this correction, so what was
    # suggested is part of what was agreed, and the name that is finally reserved and
    # released has to be traceable to the moment it was offered.
    for report_dict in reports:
        # Both codes want a free alternative offered. A name another submission claimed
        # first is just as unavailable to this submitter as one the release already holds
        # -- the difference is only who they have to be told about.
        taken = {i.get("subject") for i in report_dict.get("issues", [])
                 if i.get("code") in ("name_already_in_malavi",
                                      "name_claimed_by_another_submission")
                 and i.get("subject")}

        # A name being taken is NOT on its own a reason to suggest a new one. If the
        # sequence under that name is the lineage that already holds it, the submission is
        # a record of a known lineage, not a new one -- and renaming it would create a
        # duplicate of something MalAvi already has. Offering a rename here produced two
        # pages of the same report giving opposite instructions.
        already_a_known_lineage = {
            i.get("subject") for i in report_dict.get("issues", [])
            if i.get("code") == "sequence_is_known_lineage" and i.get("subject")}
        # Which MalAvi lineage each one actually matched. Recording only the proposed name
        # made the report say "this sequence is the TUMIG50 MalAvi already has" about a
        # name MalAvi has never held -- and told the curator to file it under a lineage
        # that does not exist, which is how a duplicate gets created.
        matched = {}
        for entry_seq in report_dict.get("sequences", []):
            label = entry_seq.get("label")
            if label in already_a_known_lineage:
                nearest = (entry_seq.get("nearest") or [{}])[0]
                matched[label] = nearest.get("lineage") or ""
        report_dict["not_new_lineages"] = matched

        if not taken or known is None:
            continue
        suggestions = {}
        # Everything that is not free: what the release owns, what THIS submission is
        # claiming (so two taken names cannot both be offered the same free number), and
        # what other submissions in the queue have already claimed. The last of those was
        # missing, and it is the one that could hand two submitters the same name.
        claimed_here = set(known) | set(claimed_elsewhere or {}) | {
            str(n) for n in (report_dict.get("lineages") or {})}
        for name, entry in (report_dict.get("lineages") or {}).items():
            if name not in taken or name in already_a_known_lineage:
                continue
            host = (entry or {}).get("host_species") or ""
            if not host:
                # The workbook's NewLineages sheet carries the host; fall back to the
                # first host record for this lineage rather than guessing an acronym.
                for sub in submissions:
                    for record in sub.get("records") or []:
                        if record.get("lineage_name") == name and record.get("host_species"):
                            host = record["host_species"]
                            break
                    if host:
                        break
            proposal = suggest_name(host, sorted(claimed_here))
            if proposal.ok and proposal.proposal:
                suggestions[name] = proposal.proposal
                claimed_here.add(proposal.proposal)
        if suggestions:
            report_dict["name_suggestions"] = suggestions
            # Put the alternative in the finding a curator actually reads, rather than
            # only on the summary page.
            for entry_issue in report_dict.get("issues", []):
                free_name = suggestions.get(entry_issue.get("subject"))
                if free_name and entry_issue.get("code") in (
                        "name_already_in_malavi", "name_claimed_by_another_submission"):
                    entry_issue["message"] += (
                        f" Suggesting {free_name} for this lineage.")
            for taken_name, free_name in suggestions.items():
                print(f"   [suggestion] {taken_name} is taken -> {free_name}")


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("target", help="a submission directory or a single .xlsx template")
    ap.add_argument("--alignment", help="override the reference alignment FASTA")
    ap.add_argument("--online", action="store_true",
                    help="allow checks that need the network (INSDC accession lookup)")
    ap.add_argument("--no-r", action="store_true",
                    help="skip the malaviR validators; they are then reported as "
                         "skipped, never as passed")
    ap.add_argument("--redact", action="store_true",
                    help="print check ids and counts but no submitted values. Use "
                         "anywhere the output is captured somewhere more people can "
                         "read than should see an unpublished submission, GitHub "
                         "Actions logs above all")
    args = ap.parse_args(argv)

    root = repo_root()
    cfg = load_config()
    release = cfg["malaviR"]["release"]

    aln = Path(args.alignment) if args.alignment else default_alignment_path(root, release)
    if aln is None or not aln.is_file():
        print(f"No reference alignment for release {release}. "
              f"Run export/build_downloads.R first.", file=sys.stderr)
        return 1
    ref = Reference.from_fasta(aln)

    # None, not an empty set, when the snapshot is absent -- see screen()'s docstring.
    known = None
    snap = root / "curation" / "src" / "malavi_curation" / "data" / "db_snapshot.json"
    if snap.is_file():
        known = {normalize.lineage_name(n) or ""
                 for n in json.loads(snap.read_text()).get("lineages", [])}
        known.discard("")

    target = Path(args.target)
    books = ([target] if target.suffix.lower() == ".xlsx"
             else sorted(p for p in target.rglob("*.xlsx") if not p.name.startswith("~$")))
    if not books:
        # A paper with no template is a legitimate submission, not an error. Exiting here
        # used to leave a curator with an email about a submission, nothing to review, and
        # no explanation. Write them a report that says what it is and what to do.
        if not target.is_dir():
            print(f"No .xlsx template found under {target}", file=sys.stderr)
            return 1

        metadata_path = target / "metadata.json"
        metadata = (json.loads(metadata_path.read_text(encoding="utf-8"))
                    if metadata_path.is_file() else {})
        public_id = (submission_id_for(target.parent, target.name)
                     if target.parent.name == "submissions" else None)

        print("No data template in this submission — writing the paper-only report.")
        page = render_paper_only_report(
            metadata=metadata, submission_id=public_id,
            submit_form_url=(cfg.get("submissions") or {}).get("form_url", ""))
        intake_root = root / "curation" / "intake"
        print(f"wrote {write_report(page, target / 'report.html', intake_root=intake_root)}")
        pdf = write_pdf(page, target / "report.pdf", intake_root=intake_root)
        if pdf:
            print(f"wrote {pdf}")
        # Not 0: nothing was screened, and a zero exit is how this pipeline says "clean".
        # Not 2 either: nothing is blocking, because nothing was examined. A distinct code
        # so a caller can tell "a curator must do something" from "a check failed".
        return 3

    print("== malavi_rebuild :: check_template ==")
    print(f"release   : {release}")
    print(f"reference : {len(ref.names)} lineages x {ref.width} bp")
    if known is None:
        print("known names: UNAVAILABLE -- the name-collision check will be SKIPPED, "
              "not passed.\n           Run: Rscript curation/r/gate_reference.R "
              "> curation/src/malavi_curation/data/db_snapshot.json\n")
    else:
        print(f"known names: {len(known)}\n")

    # The submitter block comes from the Google Form answers when this is a fetched
    # submission directory. A workbook handed over by other means simply has none, which
    # the adapter handles.
    submitter = None
    metadata = None
    if target.is_dir():
        metadata_path = target / "metadata.json"
        if metadata_path.is_file():
            try:
                metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
                submitter = submitter_from_metadata(metadata)
            except (ValueError, OSError) as exc:
                # Never let unreadable metadata stop the screen: the sequences and names
                # are the part that must not go unchecked, and a missing submitter is
                # something the curator report can say out loud.
                print(f"   [warn] could not read {metadata_path.name}: {exc}")

    # What other submissions in the queue have already claimed. Read once, here, rather
    # than per workbook. The directory being screened is excluded so it cannot collide
    # with itself on a re-run: check_template.py is re-run after every correction, and by
    # then this submission's own names are in its own screen.json.
    claimed_elsewhere = standing_claims_for(root, cfg, target, known)
    if claimed_elsewhere:
        print(f"{len(claimed_elsewhere)} name(s) are claimed by other submissions in the "
              f"queue and are treated as unavailable.\n")

    reports = []
    submissions = []
    build_failures: List[str] = []
    worst = 0
    for book in books:
        rep = screen(book, ref, known, claimed_elsewhere)
        reports.append(rep)

        # The same workbook, as a submission conforming to schemas/submission.schema.json.
        # This is what lets the pre-ingest gate, the row flags and the malaviR validators
        # -- all of which take a submission -- run on a Form template at all. A workbook
        # that cannot be adapted must not take the screen down with it: the screen is the
        # part that protects a name from being assigned twice.
        try:
            built = build_submission_from_path(book, submitter=submitter)
            if built is not None:
                submissions.append(built)
            else:
                # Not a template at all -- a supplementary spreadsheet travelling with the
                # submission. Expected, and not a failure.
                print(f"   [note] {book.name}: not a data template; skipped.")
        except Exception as exc:                       # noqa: BLE001
            # Loud, and it counts. This used to be a warning that let the run continue and
            # exit zero, so a missing dependency on the scheduled runner produced no
            # checks.json, no report, and a green tick.
            print(f"   [ERROR] {book.name}: could not build a submission record: "
                  f"{type(exc).__name__}: {exc}", file=sys.stderr)
            build_failures.append(book.name)

        print(f"--- {book.name} ---")
        if not rep["lineages"] and not rep["sequences"]:
            print("   no new lineages or sequences declared (records-only submission)\n")
            continue
        for s in rep["sequences"]:
            near = s["nearest"][0] if s["nearest"] else None
            print(f"   {s['label']}: {s['raw_length']} bp, offset {s['offset']}, "
                  f"{s['n_stop_codons']} stops -> {s['verdict']}")
            if near:
                print(f"       nearest {near['lineage']} (distance {near['distance']} "
                      f"over {near['comparable']} positions)")
        for i in rep["issues"]:
            print(f"   [{i['severity']}] {i['message']}")
            worst = max(worst, {"info": 0, "warn": 1, "error": 2}[i["severity"]])
        for sk in rep.get("skipped", []):
            # A blocking check that could not run makes the whole run incomplete, and an
            # incomplete run must not exit 0. This is deliberately not how --no-r behaves:
            # skipping the R validators is an opt-in the operator asked for, whereas a
            # missing snapshot is an unnoticed gap in the one check that stops a duplicate
            # lineage name reaching a paper and GenBank.
            print(f"   [SKIPPED] {sk['code']}: {sk['reason']}")
            worst = max(worst, 2)
        print()

    offer_free_names(reports, known, claimed_elsewhere, submissions)

    if build_failures:
        # An unreadable workbook means the report below describes less than was submitted,
        # or nothing at all. Zero would say "clean".
        print(f"\n{len(build_failures)} workbook(s) could not be read: "
              f"{', '.join(build_failures)}", file=sys.stderr)
        worst = max(worst, 2)

    if target.is_dir():
        # screen.json keeps its exact shape: build_name_reservations.py reads the claimed
        # names out of it, so what gets reserved stays exactly what a curator was shown.
        out = target / "screen.json"
        out.write_text(json.dumps(reports, indent=2), encoding="utf-8")
        print(f"wrote {out}")

        if submissions:
            # One workbook is the ordinary case and gets a plain object; several are
            # written as a list rather than silently merged, because rows from two
            # workbooks are two submissions and combining them would attribute one
            # submitter's records to the other's reference.
            payload = submissions[0] if len(submissions) == 1 else submissions
            out = target / "submission.json"
            out.write_text(json.dumps(payload, indent=2, ensure_ascii=False),
                           encoding="utf-8")
            print(f"wrote {out}")

            # The full check suite over the same submission: the gate, the row flags,
            # the malaviR validators and this screen, all reporting in one vocabulary.
            # Anything that could not run says so rather than passing quietly.
            run = run_checks(submissions[0], screen=reports,
                             check_online=args.online, run_r=not args.no_r)
            print()
            print(render_console(run, heading="== checks ==", redact=args.redact))
            out = target / "checks.json"
            out.write_text(json.dumps(run.as_dict(), indent=2, ensure_ascii=False),
                           encoding="utf-8")
            print(f"\nwrote {out}")

            # The curator's actual document. Written into the submission directory,
            # which is gitignored: it carries unpublished sequences and the submitter's
            # email, and write_report refuses any destination outside that tree.
            # The submitted sequences beside their nearest relatives. Built here
            # because this is where the reference alignment is already loaded, and from
            # the SAME registration the distances came from -- a figure drawn by a
            # separate aligner could illustrate a different comparison than the one
            # reported a few lines above it.
            registered = {s["label"]: s.get("registered")
                          for rep in reports for s in rep.get("sequences", [])}
            alignments = build_figures(
                reports, registered, dict(zip(ref.names, ref.seqs)))

            # The opaque identifier for this submission, minted once and never changed.
            # It is what the verdict link carries, so a curator's answer attaches to this
            # submission rather than to whatever they typed.
            public_id = submission_id_for(target.parent, target.name) \
                if target.parent.name == "submissions" else None
            # The revision this report describes, and the flags and corrections already
            # standing on it. Read from the ledger where one exists, so the verdict link
            # names the version the curator is actually looking at -- and so the report can
            # print the V1/C1 ids the verdict form asks them to type, which exist nowhere a
            # curator can see.
            revision = 1
            entry = None
            if public_id:
                try:
                    from malavi_curation.ledger import load as load_ledger
                    entry = load_ledger(target.parent).get(public_id)
                    if entry:
                        revision = entry.revision
                except Exception as exc:                      # noqa: BLE001
                    print(f"   [warn] could not read the review ledger ({exc}); "
                          f"the report will say revision 1 and will not list any "
                          f"standing flags")

            report = render_report(
                submissions[0], run,
                screen=reports,
                workbook_path=books[0] if books else None,
                metadata=metadata,
                submission_id=public_id,
                revision=revision,
                entry=entry,
                alignments=alignments,
            )
            intake_root = root / "curation" / "intake"
            written = write_report(report, target / "report.html",
                                   intake_root=intake_root)
            print(f"wrote {written}")

            # The PDF is the copy a curator opens, because it is delivered through Drive
            # and Drive renders a PDF in the browser while it makes HTML a download.
            # Its absence is reported, never fatal: a missing renderer must not cost the
            # curator their report.
            pdf = write_pdf(report, target / "report.pdf", intake_root=intake_root)
            if pdf:
                print(f"wrote {pdf}")
            else:
                print("   [note] no PDF written — install the 'report' extra "
                      "(pip install -e 'curation[report]') for the Drive-ready copy")
            if run.incomplete:
                # An incomplete run must not read as a pass at the shell either.
                worst = max(worst, 2)
    return 0 if worst < 2 else 2


if __name__ == "__main__":
    sys.exit(main())
