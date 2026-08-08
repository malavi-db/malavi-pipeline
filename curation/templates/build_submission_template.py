# @title Build the MalAvi community data-submission Excel template
# @purpose Generate ImportMalavi_Template_<VERSION>.xlsx: the workbook community members
#          fill in to submit new lineages and records to MalAvi.
# @why MalAvi has always accepted submissions as a filled Excel workbook (Staffan Bensch's
#      "ImportMalavi_Template", in circulation since 2012). The community knows that file,
#      so the rebuilt submission path keeps its sheet names and layout rather than
#      inventing a new format. This script regenerates it reproducibly so the template is
#      version-controlled as code instead of an opaque binary.
# @input curation/templates/ (no data inputs; vocabularies are pinned constants below)
# @output curation/templates/ImportMalavi_Template_2026-07.xlsx
# @program python
# @program openpyxl
# @critical-var TEMPLATE_VERSION
# @critical-var MALAVIR_RELEASE
# @critical-var MIN_SEQUENCE_LENGTH_BP
# @critical-var GREEN
"""Build the MalAvi data-submission Excel template.

Design rules, and why each one is here:

1. **Keep Staffan's sheet names.** ``NewLineages``, ``Sequences``, ``Reference``,
   ``Hosts_and_Sites``, ``Sites``, ``Alt_Lineage_names`` and ``Vectors`` are the names
   used by the workbook the community has been filling in for over a decade. Renaming
   them would break that familiarity for no benefit.

2. **Drop the red columns.** In the legacy template, green headers (#92D050) marked
   "submitter fills this in" and red headers (#FF0000) marked MalAvi-internal database
   index numbers that only the curator can supply. Roughly half the columns were red,
   which made the file look far more intimidating than it is. Those columns are removed
   here; the curator's tooling adds the index numbers on ingest. The single exception is
   HOST_SPECIES_ID, which the legacy instructions explicitly invite submitters to fill
   from the NCBI taxonomy browser, so it is kept and marked optional.

3. **Constrain what can be constrained.** Free-text spelling variants are the single
   biggest source of cleanup work. Every column with a known controlled vocabulary gets
   an Excel dropdown, so the submitter picks a valid value instead of typing one.

4. **Add the columns the database has but the template lacked.** The legacy workbook had
   no way to record HOST_ENVIRONMENT (Wild / Captivity) even though MalAvi stores it and
   the Grand Lineage Summary depends on it to exclude non-natural records. It is added
   here. DOI is likewise added to the Reference sheet.

Nothing in this file is MalAvi data; it only writes an empty, annotated workbook.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------------------

# Stamped into the workbook and its filename so a returned submission can be traced back
# to the template revision it was filled on.
TEMPLATE_VERSION = "2026-07"

# The malaviR release whose controlled vocabularies the dropdowns below were read from.
# If this is bumped, re-derive the vocabularies rather than assuming they are unchanged.
MALAVIR_RELEASE = "2026-03-23"

# Submission cutoff for new lineage sequences, in base pairs, over the 479 bp MalAvi
# cytochrome b barcode window. Sequences shorter than this should be re-sequenced before
# submission rather than entered as new lineages. Note this is a rule for NEW submissions
# only: 1,109 of the 5,365 lineages already in the 2026-03-23 release are shorter than
# this, and are of course retained.
MIN_SEQUENCE_LENGTH_BP = 470

# The full MalAvi barcode window length, for the instructions text.
BARCODE_WINDOW_BP = 479

# Header fills. The green is Staffan's exact legacy green so the template still "looks
# like" the file people are used to. Amber marks optional-but-helpful columns.
GREEN = PatternFill("solid", start_color="FF92D050", end_color="FF92D050")
AMBER = PatternFill("solid", start_color="FFFFD966", end_color="FFFFD966")
HEADER_FONT = Font(bold=True, color="FF000000")

# Style for the single inline example row that each data sheet carries. It is visually
# distinct (gray italic) and is matched verbatim by the submission validator, which
# ignores it if left untouched and warns if it was edited into looking like real data.
EXAMPLE_FILL = PatternFill("solid", start_color="FFF2F2F2", end_color="FFF2F2F2")
EXAMPLE_FONT = Font(italic=True, color="FF808080")

# --------------------------------------------------------------------------------------
# Controlled vocabularies
#
# Every value below was read from the bundled malaviR release named in MALAVIR_RELEASE
# (hosts_and_sites), so a submitter picking from these dropdowns cannot introduce a value
# the database has never seen. "Unknown" is a real stored value, not a placeholder.
# --------------------------------------------------------------------------------------

PARASITE_GENERA = ["Plasmodium", "Haemoproteus", "Leucocytozoon"]
HOST_AGE_VALUES = [
    "Adult",
    "Juvenile",
    "Nestling",
    "Adult + Juvenile",
    "Adult + Nestling",
    "Unknown",
]
HOST_STATUS_VALUES = ["Resident", "Migratory", "Unknown"]
HOST_ENVIRONMENT_VALUES = ["Wild", "Captivity"]


# --------------------------------------------------------------------------------------
# Sheet definitions
#
# Each sheet is (sheet name, note shown above the header, list of columns). A column is
# (header, width, optional, help text, example value). "optional" only controls the
# header color -- amber rather than green -- it does not change validation.
# --------------------------------------------------------------------------------------

SHEETS = [
    (
        "NewLineages",
        "One row per NEW lineage (a sequence differing by >=1 bp from every sequence "
        "already in MalAvi). Use ONE representative host species per lineage.",
        [
            ("LINEAGE_NAME", 16, False,
             "Proposed name: 5-6 letter acronym from the host's scientific name plus a "
             "two-digit number, e.g. ALCPOI02 for Alcippe poioicephala. Must be unique.",
             "ALCPOI02"),
            ("GENBANK_NR", 14, True,
             "GenBank accession, if the sequence has been deposited. Leave blank if not "
             "yet submitted -- MalAvi accepts lineages before GenBank deposit.",
             "MK493368"),
            ("ParasiteGenus", 16, False, "Parasite genus.", "Haemoproteus"),
            ("HostSpecies", 26, False,
             "Scientific name of the host the lineage was first found in.",
             "Alcippe poioicephala"),
            ("HOST_SPECIES_ID", 16, True,
             "Optional: NCBI taxonomy ID for the host "
             "(https://www.ncbi.nlm.nih.gov/taxonomy). The curator fills this if blank.",
             "181645"),
            ("Reference", 22, False,
             "Citation key matching a row on the Reference sheet. A preliminary title is "
             "fine for unpublished work.",
             "Gupta et al 2019"),
            ("COMMENT", 40, True,
             "Anything useful: voucher numbers, mixed infections, caveats.", ""),
        ],
    ),
    (
        "Sequences",
        "The sequence for every new lineage listed on the NewLineages sheet. "
        f"Sequences should span the {BARCODE_WINDOW_BP} bp MalAvi cytochrome b window and "
        f"be at least {MIN_SEQUENCE_LENGTH_BP} bp of unambiguous sequence. You may send a "
        "FASTA file instead of filling this sheet.",
        [
            ("LINEAGE_NAME", 16, False, "Must match a name on the NewLineages sheet.",
             "ALCPOI02"),
            ("SEQUENCE", 90, False,
             "Nucleotide sequence, new lineages only. Do not include primer sequences.",
             "ATGCATGCTACTGGTGCTACATTTGT..."),
        ],
    ),
    (
        "Reference",
        "The publication (or manuscript in preparation) the data come from. "
        "One row per reference.",
        [
            ("REFERENCE_NAME", 22, False,
             "Citation key used across the other sheets, e.g. 'Gupta et al 2019'. Use the "
             "SAME string everywhere -- the validator cross-checks this.",
             "Gupta et al 2019"),
            ("PUBLICATION_YEAR", 16, False, "Year of publication, or expected year.",
             "2019"),
            ("TITLE", 60, False, "Article title. A preliminary title is fine.",
             "Geographical and host species barriers..."),
            ("JOURNAL_NAME", 26, False, "Journal name.", "Proc. R. Soc. B"),
            ("Volume", 10, True, "Volume.", "286"),
            ("StartPage", 12, True, "First page or article number.", "20190439"),
            ("EndPage", 10, True, "Last page, if a page range.", ""),
            ("DOI", 26, True,
             "DOI if the paper has one, e.g. 10.1098/rspb.2019.0439. Helps us avoid "
             "entering the same paper twice.",
             "10.1098/rspb.2019.0439"),
        ],
    ),
    (
        "Hosts_and_Sites",
        "The records themselves. ONE ROW per lineage / host species / site combination. "
        "This is normally the largest sheet.",
        [
            ("LINEAGE_NAME", 16, False,
             "New (from NewLineages) or an existing MalAvi lineage name.", "ALCPOI02"),
            ("HostSpecies", 26, False, "Host scientific name.", "Alcippe poioicephala"),
            ("HOST_SPECIES_ID", 16, True, "Optional NCBI taxonomy ID.", "181645"),
            ("HostSubspecies", 20, True, "Leave blank if not applicable.", ""),
            ("HostAge", 18, True, "Age class of the sampled birds.", "Unknown"),
            ("HostStatus", 16, True, "Migratory behavior of the host at this site.",
             "Resident"),
            ("HostEnvironment", 18, True,
             "Free-living birds or held birds? Records marked Captivity are excluded "
             "from the Grand Lineage Summary.",
             "Wild"),
            ("Country", 18, False, "Country of sampling.", "India"),
            ("CountryRegion", 20, True, "Sub-national region. Blank if not applicable.",
             "Western Ghats"),
            ("SiteName", 20, False, "Must match a SITE_NAME on the Sites sheet.",
             "Ambalapara"),
            ("NUMBER_FOUND", 15, True,
             "How many individuals were infected with THIS lineage here. Must be <= "
             "NUMBER_TESTED.",
             "3"),
            ("NUMBER_TESTED", 15, True,
             "How many individuals of this host species were screened here.", "25"),
            ("Reference", 22, False, "Citation key from the Reference sheet.",
             "Gupta et al 2019"),
            ("COMMENT", 40, True, "e.g. 'additionally 1 mixed infection'.", ""),
        ],
    ),
    (
        "Sites",
        "One row per sampling locality named on the Hosts_and_Sites sheet.",
        [
            ("SITE_NAME", 22, False,
             "Locality name, spelled exactly as on Hosts_and_Sites.", "Ambalapara"),
            ("Country", 18, False, "Country.", "India"),
            ("LATITUDE", 18, False,
             "Degrees & decimal minutes, e.g. 11°28.20000'. Decimal degrees "
             "(11.47) are also accepted -- the validator converts and reports what it "
             "read back to you.",
             "11°56.40000'"),
            ("LONGITUDE", 18, False,
             "Degrees & decimal minutes, e.g. 076°07.80000'. Use a minus sign or W "
             "for western longitudes.",
             "075°56.40000'"),
            ("ALTITUDE(m)", 14, True,
             "Metres above sea level, if known. Leave blank and it can be derived from "
             "the coordinates.",
             "1498"),
        ],
    ),
    (
        "Alt_Lineage_names",
        "Use this when a lineage already named in MalAvi appears under a DIFFERENT name "
        "in your publication or in a GenBank record. This is how MalAvi keeps synonyms "
        "traceable -- please do fill it in, it saves a great deal of untangling later.",
        [
            ("MalAvi_Name", 16, False, "The established MalAvi lineage name.", "SGS1"),
            ("Alternative_Name", 20, False, "The name used in the publication/GenBank.",
             "P15"),
            ("GenBankNr", 16, True, "Accession carrying the alternative name.",
             "AF495571"),
            ("Reference", 22, False, "Citation key from the Reference sheet.",
             "Gupta et al 2019"),
            ("Comment", 40, True, "Notes.", ""),
        ],
    ),
    (
        "Vectors",
        "Lineages detected in arthropod vectors. Leave the sheet empty if you have none.",
        [
            ("LINEAGE_NAME", 16, False, "Lineage detected.", "GRW04"),
            ("VectorSpecies", 24, False, "Vector scientific name.", "Culex pipiens"),
            ("VECTOR_METHOD", 22, True,
             "How the detection was made, e.g. PCR of abdomen, salivary glands, "
             "dissection.",
             "PCR"),
            ("Country", 18, False, "Country.", "Japan"),
            ("CountryRegion", 20, True, "Sub-national region.", ""),
            ("SiteName", 20, True, "Locality.", "Tokyo"),
            ("No_found", 12, True, "Vectors found infected with this lineage.", "2"),
            ("No_tested", 12, True, "Vectors screened.", "150"),
            ("Reference", 22, False, "Citation key from the Reference sheet.",
             "Gupta et al 2019"),
            ("Comment", 40, True, "Notes.", ""),
        ],
    ),
]

# Dropdown vocabularies, keyed by (sheet name, column header).
DROPDOWNS = {
    ("NewLineages", "ParasiteGenus"): PARASITE_GENERA,
    ("Hosts_and_Sites", "HostAge"): HOST_AGE_VALUES,
    ("Hosts_and_Sites", "HostStatus"): HOST_STATUS_VALUES,
    ("Hosts_and_Sites", "HostEnvironment"): HOST_ENVIRONMENT_VALUES,
}

# Number of blank formatted rows to leave below the example row on each data sheet.
BLANK_ROWS = 200


def build_readme_sheet(workbook: Workbook) -> None:
    """Write the leading instructions sheet.

    The legacy workbook carried no instructions -- they lived in a separate PDF that got
    detached from the file as it was forwarded around. Embedding them here means the
    workbook is self-explanatory however it reaches the submitter.
    """
    worksheet = workbook.create_sheet("READ ME")
    worksheet.sheet_properties.tabColor = "92D050"

    lines = [
        ("MalAvi data submission template", "title"),
        (f"Template version {TEMPLATE_VERSION}", "subtle"),
        ("", None),
        ("Thank you for contributing to MalAvi. Fill in what you have; a curator "
         "reviews every submission before anything is added to the database.", None),
        ("", None),
        ("There are two stages, and you can use either or both.", "head"),
        ("", None),
        ("STAGE 1 - BEFORE you publish: reserve your lineage names", "head"),
        ("Name your new lineages and register them before your manuscript is accepted "
         "and before you deposit in GenBank. That way the same names appear in your "
         "paper, in GenBank and in MalAvi, and nobody else can claim them meanwhile.",
         None),
        ("  - Fill the NewLineages and Reference sheets (a preliminary title is fine).",
         None),
        ("  - Give the sequences, either on the Sequences sheet or as a FASTA file whose "
         "names match your proposed lineage names.", None),
        ("  - We check names and sequences and tell you about any changes needed, so you "
         "can adjust before depositing in GenBank.", None),
        ("  - Send us the accession numbers once you have deposited.", None),
        ("", None),
        ("Submitting before publication is safe: only the sequence and the lineage name "
         "become visible. Without host species and locality the data are of no use to "
         "anyone else, so there is no risk of being scooped.", None),
        ("", None),
        ("STAGE 2 - WHEN your paper is accepted: submit the records", "head"),
        ("  - Fill Hosts_and_Sites (one row per lineage / host species / site), Sites, "
         "and Vectors if you have vector data.", None),
        ("  - Fill Alt_Lineage_names if your paper calls an existing MalAvi lineage by a "
         "different name.", None),
        ("  - Include a PDF of the publication if you can. It is only used to check the "
         "records against the paper, is shared with the curator alone, and is never "
         "posted anywhere.", None),
        ("", None),
        ("How to fill the sheets", "head"),
        ("  - GREEN headers are for you. AMBER headers are optional but helpful.", None),
        ("  - Hover any header cell for a note explaining what goes in that column.",
         None),
        ("  - Row 2 of each sheet is a gray italic EXAMPLE. Leave it or delete it; the "
         "validator ignores it either way.", None),
        ("  - Use the same reference key everywhere, e.g., Gupta et al. 2019", None),
        ("  - Database index numbers are no longer your problem. Earlier versions of "
         "this template had red columns for them; the curator now fills those in.", None),
        ("", None),
        ("What counts as a new lineage", "head"),
        (f"A sequence differing at one or more positions, across the {BARCODE_WINDOW_BP} "
         "bp MalAvi cytochrome b window, from every sequence already in MalAvi.", None),
        (f"Sequences shorter than {MIN_SEQUENCE_LENGTH_BP} bp of unambiguous sequence "
         "should be re-sequenced before being submitted as new lineages -- a short "
         "sequence cannot be shown to be genuinely distinct.", None),
        ("Naming: a 5-6 letter acronym from the scientific name of the first host it was "
         "found in, plus a two-digit number, e.g. Alcippe poioicephala -> ALCPOI02, "
         "Turdus migratorius -> TUMIG19. Check the Grand Lineage Summary, or use the "
         "name checker on the MalAvi website, to find a free number.", None),
        ("", None),
        ("Questions are welcome -- please ask rather than guess.", "subtle"),
    ]

    for row_index, (text, style) in enumerate(lines, start=1):
        cell = worksheet.cell(row=row_index, column=1, value=text)
        if style == "title":
            cell.font = Font(bold=True, size=16)
        elif style == "head":
            cell.font = Font(bold=True, size=12)
        elif style == "subtle":
            cell.font = Font(italic=True, color="FF808080")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    worksheet.column_dimensions["A"].width = 105


def build_data_sheet(workbook: Workbook, name: str, note: str, columns: list) -> None:
    """Write one data sheet: a note row, a colored header row, and one example row."""
    worksheet = workbook.create_sheet(name)

    # Row 1 is a plain-language note about what belongs on this sheet.
    note_cell = worksheet.cell(row=1, column=1, value=note)
    note_cell.font = Font(italic=True, color="FF404040")
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 2)
    )
    worksheet.row_dimensions[1].height = 30

    # Row 2 is the header. Green = please fill in; amber = optional.
    for column_index, (header, width, optional, help_text, _example) in enumerate(
        columns, start=1
    ):
        cell = worksheet.cell(row=2, column=column_index, value=header)
        cell.fill = AMBER if optional else GREEN
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # The per-column help text rides along as a cell comment, so the instructions
        # cannot get separated from the column they describe.
        if help_text:
            cell.comment = Comment(help_text, "MalAvi")
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    # Row 3 is the example row, styled so it cannot be mistaken for real data.
    for column_index, (_header, _w, _o, _h, example) in enumerate(columns, start=1):
        cell = worksheet.cell(row=3, column=column_index, value=example or None)
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT

    # Attach dropdowns over the data range for any column with a controlled vocabulary.
    first_data_row = 4
    last_data_row = first_data_row + BLANK_ROWS
    for column_index, (header, *_rest) in enumerate(columns, start=1):
        vocabulary = DROPDOWNS.get((name, header))
        if not vocabulary:
            continue
        letter = get_column_letter(column_index)
        validation = DataValidation(
            type="list",
            formula1='"{}"'.format(",".join(vocabulary)),
            allow_blank=True,
            showDropDown=False,  # False means "show the dropdown arrow" in the file format
        )
        validation.error = "Pick one of: " + ", ".join(vocabulary)
        validation.errorTitle = "Value not recognized"
        worksheet.add_data_validation(validation)
        # Cover the example row too, so the example value itself validates.
        validation.add(f"{letter}3:{letter}{last_data_row}")

    # Freeze the note + header rows so the headers stay visible while scrolling.
    worksheet.freeze_panes = "A3"


def main() -> None:
    """Assemble the workbook and write it next to this script."""
    workbook = Workbook()
    # Drop the default sheet openpyxl creates; every sheet here is built explicitly.
    workbook.remove(workbook.active)

    build_readme_sheet(workbook)
    for name, note, columns in SHEETS:
        build_data_sheet(workbook, name, note, columns)

    output_path = (
        Path(__file__).resolve().parent
        / f"ImportMalavi_Template_{TEMPLATE_VERSION}.xlsx"
    )
    workbook.save(output_path)
    print(f"Wrote {output_path}")
    print(f"  sheets: {', '.join(workbook.sheetnames)}")


if __name__ == "__main__":
    main()
