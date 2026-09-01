"""Tests for building a release out of the record store.

The store is the database now, so these tests guard the step that lets anyone else read
it. Two kinds of property matter and they are tested separately:

* **The derivation rules are the release's own.** Every one of them was recovered by
  measuring against the 2026-03-23 release rather than reasoned from first principles, so
  the tests pin the measurement, not the reasoning. If someone later "simplifies"
  SUM_HOST to a row count or drops the vector table from the region derivation, a test
  fails with the reason attached.
* **The emitted archive is readable by malaviR.** Column order, blank-versus-NA, the file
  names inside the ZIP and the alignment's rectangularity are all things that break
  silently downstream, months later, in somebody else's analysis.
"""
from __future__ import annotations

import csv
import zipfile

import pytest

from malavi_curation.country_regions import REGION_COLUMNS, region_for
from malavi_curation.release_build import (
    GRAND_LINEAGE_SUMMARY_COLUMNS, RELEASE_TABLE_FILES, build_fasta, build_release,
    derive_summary, diff_against_release, fasta_label,
)

# A minimal store, written out in full so each test can see exactly what it asserts on.
REGIONS = {"Sweden": "EUROPE", "Brazil": "SOUTH_AMERICA", "United States": "NORTH_AMERICA",
           "Kenya": "SUB_SAHARAN_AFRICA"}


def _lineage(name, sequence="ACGT", genus="Haemoproteus", species="", acc=""):
    """One lineage row.

    ``species`` is the FULL BINOMIAL, as it is in the store: all 238 lineages that carry
    one hold e.g. "Leucocytozoon toddi", never a bare epithet. This fixture accepted an
    epithet until 2026-08-10, which is why test_label_appends_a_morphospecies_when_there_
    is_one asserted the intended label format against data the store never contains, and
    the real output -- with a duplicated genus and a space in it -- shipped unchecked.
    """
    return {"LINEAGE_NAME": name, "GENBANK_ACC": acc, "SEQ_LENGTH": "Full",
            "GENUS_NAME": genus, "SPECIES_NAME": species, "SEQUENCE": sequence}


def _host(lineage, order="Passeriformes", family="Paridae", genus="Parus",
          species="Parus major", country="Sweden", country_region="", site="S",
          reference="Ref 2020"):
    return {"LINEAGE_NAME": lineage, "ORDER_NAME": order, "FAMILY_NAME": family,
            "GENUS_NAME": genus, "SPECIES_NAME": species, "COUNTRY_NAME": country,
            "COUNTRY_REGION_NAME": country_region, "SITE_NAME": site,
            "REFERENCE_NAME": reference}


def _vector(lineage, vector_species="Culex pipiens", country="Brazil"):
    return {"LINEAGE_NAME": lineage, "VECTOR_SPECIES": vector_species,
            "COUNTRY_NAME": country, "VECTOR_METHOD": "PCR", "SITE_NAME": "",
            "REFERENCE_NAME": "Ref 2020"}


def _store(lineages=None, hosts=None, vectors=None):
    return {"lineages": lineages or [], "host_records": hosts or [],
            "vector_records": vectors or [], "morpho_species": [], "references": [],
            "alt_names": []}


class TestTallies:
    def test_sum_host_counts_binomials_not_rows(self):
        """The same host found by three studies is three rows and one host.

        Measured: counting rows disagrees with the 2026-03-23 release far more often
        than counting distinct binomials does.
        """
        store = _store([_lineage("A")], [
            _host("A", reference="Ref 2020"),
            _host("A", reference="Ref 2021"),
            _host("A", reference="Ref 2022"),
        ])
        assert derive_summary(store, REGIONS)[0]["SUM_HOST"] == "1"

    def test_sum_vectors_counts_species_not_rows(self):
        store = _store([_lineage("A")], vectors=[
            _vector("A", "Culex pipiens"), _vector("A", "Culex pipiens"),
            _vector("A", "Culex quinquefasciatus"),
        ])
        assert derive_summary(store, REGIONS)[0]["SUM_VECTORS"] == "2"

    def test_blank_taxonomy_is_not_counted_as_a_taxon(self):
        """A host record with no family recorded must not become a family of its own."""
        store = _store([_lineage("A")], [
            _host("A", family="Paridae"), _host("A", family=""),
        ])
        assert derive_summary(store, REGIONS)[0]["SUM_FAMILY"] == "1"

    def test_zero_tallies_are_blank_but_passeriformes_is_explicit(self):
        """The release's own habit: blank tallies, an explicit 0/1 flag."""
        row = derive_summary(_store([_lineage("A")]), REGIONS)[0]
        assert row["SUM_HOST"] == "" and row["SUM_VECTORS"] == ""
        assert row["PASSERIFORMES"] == "0"

    def test_passeriformes_is_set_by_any_host_record(self):
        store = _store([_lineage("A")], [
            _host("A", order="Anseriformes"), _host("A", order="Passeriformes"),
        ])
        assert derive_summary(store, REGIONS)[0]["PASSERIFORMES"] == "1"


class TestRegions:
    def test_regions_come_from_vector_records_too(self):
        """614 lineages in the seeded release have no host record at all.

        Reading host records only would strip the geography from every one of them.
        """
        store = _store([_lineage("A")], vectors=[_vector("A", country="Brazil")])
        row = derive_summary(store, REGIONS)[0]
        assert row["SOUTH_AMERICA"] == "1"

    def test_a_lineage_carries_every_region_its_records_touch(self):
        store = _store([_lineage("A")],
                       [_host("A", country="Sweden"), _host("A", country="Kenya")],
                       [_vector("A", country="Brazil")])
        row = derive_summary(store, REGIONS)[0]
        flagged = {c for c in REGION_COLUMNS if row[c] == "1"}
        assert flagged == {"EUROPE", "SUB_SAHARAN_AFRICA", "SOUTH_AMERICA"}

    def test_hawaii_replaces_north_america_rather_than_joining_it(self):
        """Measured on the release: FREMIN01's only US records are Hawaiian and it
        carries HAWAI with no NORTH_AMERICA."""
        store = _store([_lineage("A")],
                       [_host("A", country="United States", country_region="Hawaii")])
        row = derive_summary(store, REGIONS)[0]
        assert row["HAWAI"] == "1"
        assert row["NORTH_AMERICA"] == ""

    def test_a_country_outside_the_table_sets_no_region(self):
        """Not UNKNOWN_REGION: an unmapped country is a gap somebody must close, and
        filing it under a real region leaves nothing to notice."""
        store = _store([_lineage("A")], [_host("A", country="Atlantis")])
        row = derive_summary(store, REGIONS)[0]
        assert all(row[c] == "" for c in REGION_COLUMNS)
        assert region_for("Atlantis", "", REGIONS) is None


class TestAlignment:
    def test_label_carries_the_genus_prefix(self):
        assert fasta_label(_lineage("PADOM01", genus="Plasmodium")) == "P_PADOM01"
        assert fasta_label(_lineage("PADOM01", genus="Leucocytozoon")) == "L_PADOM01"

    def test_label_appends_a_morphospecies_when_there_is_one(self):
        label = fasta_label(_lineage("ALARV01", genus="Haemoproteus",
                                     species="Haemoproteus tartakovskyi"))
        assert label == "H_ALARV01_Haemoproteus_tartakovskyi"

    def test_no_label_contains_whitespace(self):
        """FASTA readers truncate the sequence id at the first space.

        238 released labels did, so their ids silently lost everything after the genus.
        """
        for species in ("Leucocytozoon toddi", "Haemoproteus tartakovskyi", ""):
            label = fasta_label(_lineage("X01", genus="Leucocytozoon", species=species))
            assert " " not in label

    def test_the_genus_is_not_repeated(self):
        label = fasta_label(_lineage("ACCFRA01", genus="Leucocytozoon",
                                     species="Leucocytozoon toddi"))
        assert label == "L_ACCFRA01_Leucocytozoon_toddi"

    def test_an_unassigned_genus_contributes_no_slash(self):
        """A / in a FASTA id is trouble.

        GENUS_NAME was literally "N/A" on 7 Accipitridae lineages until 2026-08-20, when
        COR-000031 set them to Haemoproteus (Harl et al. 2024). No row carries "N/A" now,
        so this guards the handling rather than a value currently in the store -- which is
        the point: the next unrecognized genus must not acquire a slash either.
        """
        label = fasta_label(_lineage("ACCNIS06", genus="N/A",
                                     species="Haemoproteus nisi"))
        assert label == "ACCNIS06_Haemoproteus_nisi"

    def test_an_unknown_genus_gets_no_invented_prefix(self):
        """Better a visibly odd label than a sequence silently filed under P_."""
        assert fasta_label(_lineage("X01", genus="Babesia")) == "X01"

    def test_lineages_without_a_sequence_are_omitted_not_blank(self):
        """An alignment row of nothing is indistinguishable from a row of gaps."""
        fasta = build_fasta([_lineage("A", sequence="ACGT"), _lineage("B", sequence="")])
        assert fasta.count(">") == 1
        assert ">H_B" not in fasta

    def test_sequences_are_wrapped(self):
        fasta = build_fasta([_lineage("A", sequence="A" * 150)], wrap=60)
        body = [line for line in fasta.splitlines() if not line.startswith(">")]
        assert [len(line) for line in body] == [60, 60, 30]


class TestArchive:
    @pytest.fixture
    def built(self, tmp_path):
        store = _store([_lineage("A", sequence="ACGT"), _lineage("B", sequence="ACGT")],
                       [_host("A"), _host("B", country="Brazil")],
                       [_vector("A")])
        return build_release(store, "2026-08-07", tmp_path, REGIONS), tmp_path

    def test_the_zip_holds_five_tables_and_one_alignment(self, built):
        report, tmp_path = built
        with zipfile.ZipFile(report["archive"]) as bundle:
            names = sorted(n.split("/")[-1] for n in bundle.namelist())
        assert names == sorted(
            [f"{stem}_2026-08-07.xlsx" for stem in RELEASE_TABLE_FILES.values()]
            + ["MalAvi_2026-08-07.fas"])

    def test_everything_sits_in_a_folder_named_for_the_release(self, built):
        """process_release.R unzips to a temp dir and globs by prefix."""
        report, _ = built
        with zipfile.ZipFile(report["archive"]) as bundle:
            assert all(n.startswith("MalAvi_2026-08-07/") for n in bundle.namelist())

    def test_the_summary_keeps_the_releases_column_order(self, built, tmp_path):
        import openpyxl
        report, _ = built
        path = tmp_path / "MalAvi_2026-08-07" / "GrandLineageSummary_2026-08-07.xlsx"
        sheet = openpyxl.load_workbook(path).active
        header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        assert tuple(header) == GRAND_LINEAGE_SUMMARY_COLUMNS

    def test_blank_cells_are_empty_not_empty_strings(self, built, tmp_path):
        """readxl turns an empty cell into NA and an empty string into "".

        Every downstream is.na() stops matching if this regresses.
        """
        import openpyxl
        path = tmp_path / "MalAvi_2026-08-07" / "GrandLineageSummary_2026-08-07.xlsx"
        sheet = openpyxl.load_workbook(path).active
        values = [cell.value for row in sheet.iter_rows(min_row=2) for cell in row]
        assert "" not in values

    def test_provenance_columns_never_reach_the_release(self, built, tmp_path):
        """RECORD_ID, _source and _added are ours; the release format is Staffan's."""
        import openpyxl
        for stem in RELEASE_TABLE_FILES.values():
            path = tmp_path / "MalAvi_2026-08-07" / f"{stem}_2026-08-07.xlsx"
            sheet = openpyxl.load_workbook(path).active
            header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
            assert not {"RECORD_ID", "_source", "_added"} & set(header)

    def test_a_ragged_alignment_is_reported(self, tmp_path):
        store = _store([_lineage("A", sequence="ACGT"),
                        _lineage("B", sequence="ACGTACGT")])
        report = build_release(store, "2026-08-07", tmp_path, REGIONS)
        assert any("ragged" in w for w in report["warnings"])

    def test_a_repeated_lineage_name_is_reported(self, tmp_path):
        """TUPHI01 is already like this in the seeded release."""
        store = _store([_lineage("A"), _lineage("A", acc="X1")])
        report = build_release(store, "2026-08-07", tmp_path, REGIONS)
        assert any("more than once" in w for w in report["warnings"])

    def test_an_unmapped_country_is_reported(self, tmp_path):
        store = _store([_lineage("A")], [_host("A", country="Atlantis")])
        report = build_release(store, "2026-08-07", tmp_path, REGIONS)
        assert "Atlantis" in report["unmapped_countries"]
        assert any("Atlantis" in w for w in report["warnings"])


class TestDiff:
    def _reference(self, tmp_path, rows):
        path = tmp_path / "grand_lineage_summary_2026-03-23.csv"
        with open(path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=GRAND_LINEAGE_SUMMARY_COLUMNS,
                                    extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({c: row.get(c, "") for c in GRAND_LINEAGE_SUMMARY_COLUMNS})
        return path

    def test_a_changed_derived_value_is_reported_per_column(self, tmp_path):
        reference = self._reference(tmp_path, [{"LINEAGE_NAME": "A", "SUM_HOST": "24"}])
        diff = diff_against_release([{**{c: "" for c in GRAND_LINEAGE_SUMMARY_COLUMNS},
                                      "LINEAGE_NAME": "A", "SUM_HOST": "23"}], reference)
        assert diff["changed_lineages"] == 1
        assert diff["by_column"]["SUM_HOST"]["changed"] == 1
        assert diff["by_column"]["SUM_HOST"]["examples"][0] == {
            "lineage": "A", "was": "24", "now": "23"}

    def test_primary_facts_are_not_diffed(self, tmp_path):
        """A corrected accession is a curated record change with a reason attached, and
        belongs in the record diff, not in the derived-column report."""
        reference = self._reference(tmp_path, [{"LINEAGE_NAME": "A", "GENBANK_ACC": "X1"}])
        diff = diff_against_release([{**{c: "" for c in GRAND_LINEAGE_SUMMARY_COLUMNS},
                                      "LINEAGE_NAME": "A", "GENBANK_ACC": "X2"}], reference)
        assert diff["changed_lineages"] == 0

    def test_new_and_retired_lineages_are_separated_from_changes(self, tmp_path):
        reference = self._reference(tmp_path, [{"LINEAGE_NAME": "OLD"}])
        diff = diff_against_release([{**{c: "" for c in GRAND_LINEAGE_SUMMARY_COLUMNS},
                                      "LINEAGE_NAME": "NEW"}], reference)
        assert diff["only_in_build"] == ["NEW"]
        assert diff["only_in_reference"] == ["OLD"]
        assert diff["lineages_compared"] == 0


class TestTheRecordChecks:
    """Referential and arithmetic faults that shipped silently in the seeded store.

    Warnings rather than refusals: each is a curator's decision about somebody's data.
    What they must not do is go unnoticed, which is what they did until 2026-08-10.
    """

    def _warnings(self, tmp_path, store):
        return build_release(store, "2026-01-01", tmp_path, region_map=REGIONS)["warnings"]

    def test_a_citation_with_no_reference_row_is_reported(self, tmp_path):
        store = {"lineages": [_lineage("L1")],
                 "host_records": [_host("L1", reference="Kakogawa et al 2019")],
                 "references": [], "vector_records": [], "morpho_species": [],
                 "alt_names": []}
        warnings = self._warnings(tmp_path, store)
        assert any("Kakogawa et al 2019" in w and "no row in references.csv" in w
                   for w in warnings)

    def test_an_unpublished_citation_is_not_reported(self, tmp_path):
        """"<Authors> unpubl" has no reference row BY DESIGN -- there is nothing to cite."""
        store = {"lineages": [_lineage("L1")],
                 "host_records": [_host("L1", reference="Ellis et al unpubl")],
                 "references": [], "vector_records": [], "morpho_species": [],
                 "alt_names": []}
        assert not any("references.csv" in w for w in self._warnings(tmp_path, store))

    def test_a_record_with_no_reference_at_all_is_reported(self, tmp_path):
        store = {"lineages": [_lineage("L1")],
                 "host_records": [_host("L1", reference="")],
                 "references": [], "vector_records": [], "morpho_species": [],
                 "alt_names": []}
        assert any("no REFERENCE_NAME at all" in w for w in self._warnings(tmp_path, store))

    def test_a_duplicated_reference_name_is_reported(self, tmp_path):
        """It is the join key, so a duplicate fans out every join on it."""
        reference = {"REFERENCE_NAME": "Pramual et al 2020", "PUBLICATION_YEAR": "2020",
                     "TITLE": "T", "JOURNAL_NAME": "J", "VOLUME_PAGES": "1:1",
                     "STUDY_TYPE": "Vector screening"}
        store = {"lineages": [_lineage("L1")], "host_records": [],
                 "references": [reference, dict(reference, TITLE="t")],
                 "vector_records": [], "morpho_species": [], "alt_names": []}
        assert any("appear more than once in references.csv" in w
                   for w in self._warnings(tmp_path, store))

    def test_more_found_than_tested_is_reported(self, tmp_path):
        host = _host("L1", reference="Ref 2020")
        host.update({"RECORD_ID": "HST-000867", "NUMBER_FOUND": "2", "NUMBER_TESTED": "1"})
        store = {"lineages": [_lineage("L1")], "host_records": [host],
                 "references": [], "vector_records": [], "morpho_species": [],
                 "alt_names": []}
        warnings = self._warnings(tmp_path, store)
        assert any("more infections than birds tested" in w and "HST-000867" in w
                   for w in warnings)

    def test_a_numerator_with_no_denominator_is_not_reported(self, tmp_path):
        """1,691 rows are like this. It is what Staffan received, not a fault."""
        host = _host("L1", reference="Ref 2020")
        host.update({"NUMBER_FOUND": "2", "NUMBER_TESTED": ""})
        store = {"lineages": [_lineage("L1")], "host_records": [host],
                 "references": [], "vector_records": [], "morpho_species": [],
                 "alt_names": []}
        assert not any("more infections" in w for w in self._warnings(tmp_path, store))
